import json
import logging
import os
import re
import subprocess
import tempfile
import time  # For session timeout tracking
import uuid
from datetime import datetime

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth import logout as django_logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.core.serializers.json import DjangoJSONEncoder
from django.db import transaction
from django.db.models import Case, Count, F, Q, Sum, When
from django.http import (
    HttpRequest,
    HttpResponse,
    HttpResponseForbidden,
    JsonResponse,
    StreamingHttpResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.decorators.http import require_GET
from openpyxl import load_workbook
from weasyprint import HTML

from administration.models import Department, Faculty, InventoryYear, SystemSettings, UserProfile
from inventory.models import FacultyItemStock, Item, ItemCategory, SubWarehouse
from inventory.utils import _has_related_transactions, create_pdf_report

from .forms import (
    DepartmentForm,
    EmployeeForm,
    FacultyDepartmentForm,
    FacultyForm,
    UserAdminForm,
    UserPasswordChangeForm,
    SessionSettingsForm,
)

logger = logging.getLogger(__name__)


DEFAULT_PASSWORD = "pun121fate121"


@login_required
def department_list(request):
    """List departments for current faculty - OPTIMIZED."""

    if not hasattr(request.user, "profile") or not request.user.profile.faculty:
        messages.error(request, "ليس لديك كليّة مرتبطة بحسابك.")
        return redirect("home")

    faculty = request.user.profile.faculty

    departments = (
        Department.objects.filter(faculty=faculty)
        .select_related("faculty")
        .order_by("name")
    )

    context = {
        "departments": departments,
        "faculty": faculty,
    }
    return render(request, "administration/department_list.html", context)


@login_required
def department_create(request):
    """Create department for current faculty"""
    # if (
    #     not request.user.profile.is_inventory_manager
    #     or not request.user.profile.is_inventory_employee
    # ):
    #     return HttpResponseForbidden("ليس لديك صلاحية إضافة الأقسام")

    if not hasattr(request.user, "profile") or not request.user.profile.faculty:
        messages.error(request, "ليس لديك كليّة مرتبطة بحسابك.")
        return redirect("home")

    if request.method == "POST":
        form = DepartmentForm(request.POST)
        if form.is_valid():
            department = form.save(commit=False)
            department.faculty = request.user.profile.faculty
            department.save()
            messages.success(request, "تم إضافة القسم بنجاح")
            return redirect("department_list")
        else:
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
    else:
        form = DepartmentForm()

    context = {
        "form": form,
        "title": "إضافة قسم جديد",
        "faculty": request.user.profile.faculty,
    }
    return render(request, "administration/department_form.html", context)


@login_required
def department_edit(request, department_id):
    """Edit department (only if it belongs to current faculty)"""
    if not request.user.profile.is_inventory_manager:
        return HttpResponseForbidden("ليس لديك صلاحية تعديل الأقسام")

    if not hasattr(request.user, "profile") or not request.user.profile.faculty:
        messages.error(request, "ليس لديك كليّة مرتبطة بحسابك.")
        return redirect("home")

    department = get_object_or_404(Department, id=department_id)

    # Verify department belongs to current faculty
    if department.faculty != request.user.profile.faculty:
        return HttpResponseForbidden("ليس لديك صلاحية تعديل هذا القسم.")

    if request.method == "POST":
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تحديث القسم بنجاح")
            return redirect("department_list")
        else:
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
    else:
        form = DepartmentForm(instance=department)

    context = {
        "form": form,
        "title": "تعديل القسم",
        "faculty": request.user.profile.faculty,
    }
    return render(request, "administration/department_form.html", context)


@login_required
def department_delete(request, department_id):
    """Delete department (only if it belongs to current faculty)"""
    if not request.user.profile.is_inventory_manager:
        return HttpResponseForbidden("ليس لديك صلاحية حذف الأقسام")

    if not hasattr(request.user, "profile") or not request.user.profile.faculty:
        messages.error(request, "ليس لديك كليّة مرتبطة بحسابك.")
        return redirect("home")

    department = get_object_or_404(Department, id=department_id)

    # Verify department belongs to current faculty
    if department.faculty != request.user.profile.faculty:
        return HttpResponseForbidden("ليس لديك صلاحية حذف هذا القسم.")

    if request.method == "POST":
        department_name = department.name
        department.delete()
        messages.success(request, f"تم حذف القسم '{department_name}' بنجاح")
        return redirect("department_list")

    context = {
        "object": department,
        "faculty": request.user.profile.faculty,
    }
    return render(request, "administration/department_confirm_delete.html", context)


@login_required
def employee_list(request):
    """List employees with 'is_user' status for current faculty"""
    # if not request.user.profile.is_inventory_manager:
    #     return HttpResponseForbidden("ليس لديك صلاحية إدارة الموظفين")

    if not hasattr(request.user, "profile") or not request.user.profile.faculty:
        messages.error(request, "ليس لديك كليّة مرتبطة بحسابك.")
        return redirect("home")

    employees = (
        User.objects.filter(
            profile__faculty=request.user.profile.faculty,
            profile__is_user=True,
            is_active=True,
        )
        .select_related("profile__department")
        .order_by("first_name", "last_name")
    )

    context = {
        "employees": employees,
        "faculty": request.user.profile.faculty,
    }
    return render(request, "administration/employee_list.html", context)


@login_required
def employee_create(request):
    """Create employee with default password '1'"""

    if not hasattr(request.user, "profile") or not request.user.profile.faculty:
        messages.error(request, "ليس لديك كليّة مرتبطة بحسابك.")
        return redirect("home")

    if request.method == "POST":
        form = EmployeeForm(request.POST, user_faculty=request.user.profile.faculty)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Generate unique username
                    uuid_part = uuid.uuid4().hex[:8].lower()
                    username = f"user-{uuid_part}"
                    counter = 0
                    while User.objects.filter(username=username).exists():
                        counter += 1
                        uuid_part = uuid.uuid4().hex[:8].lower()
                        username = f"user-{uuid_part}-{counter}"

                    # Create User (first_name from form.cleaned_data)
                    user = User.objects.create_user(
                        username=username,
                        first_name=form.cleaned_data["first_name"],  # From form
                        last_name="",  # Always empty per your logic
                        password="1",  # Default password
                        is_active=True,
                    )

                    # Create UserProfile using get_or_create (prevents duplicates)
                    profile, created = UserProfile.objects.get_or_create(
                        user=user,
                        defaults={
                            "faculty": request.user.profile.faculty,
                            "department": form.cleaned_data.get("department"),
                            "phone": form.cleaned_data.get("phone"),
                            "is_user": True,
                            "is_inventory_manager": False,
                            "is_inventory_employee": False,
                            "is_administration_manager": False,
                            "is_faculty_manager": False,
                        },
                    )

                    # Update if profile already existed
                    if not created:
                        profile.faculty = request.user.profile.faculty
                        profile.department = form.cleaned_data.get("department")
                        profile.phone = form.cleaned_data.get("phone")
                        profile.save()

                    messages.success(
                        request, f"تم إنشاء الموظف '{user.first_name}' بنجاح"
                    )
                    return redirect("employee_list")

            except Exception as e:
                logger.error(f"Error creating employee: {str(e)}", exc_info=True)
                messages.error(request, f"حدث خطأ أثناء إنشاء الموظف: {str(e)}")
        else:
            # Log form errors for debugging
            for field, errors in form.errors.items():
                logger.error(f"Form error - {field}: {errors}")
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
    else:
        form = EmployeeForm(user_faculty=request.user.profile.faculty)

    context = {
        "form": form,
        "title": "إضافة موظف جديد",
        "faculty": request.user.profile.faculty,
    }
    return render(request, "administration/employee_form.html", context)


@login_required
def employee_edit(request, employee_id):
    """Edit employee"""
    if not request.user.profile.is_inventory_manager:
        return HttpResponseForbidden("ليس لديك صلاحية تعديل الموظفين")

    if not hasattr(request.user, "profile") or not request.user.profile.faculty:
        messages.error(request, "ليس لديك كليّة مرتبطة بحسابك.")
        return redirect("home")

    employee = get_object_or_404(User, id=employee_id)

    # Prevent editing superusers
    if employee.is_superuser:
        messages.error(
            request, "لا يمكن تعديل بيانات المستخدمين المشرفين (Superusers)."
        )
        return redirect("employee_list")

    # Verify employee belongs to current faculty
    if (
        not hasattr(employee, "profile")
        or employee.profile.faculty != request.user.profile.faculty
    ):
        return HttpResponseForbidden("ليس لديك صلاحية تعديل هذا الموظف.")

    if not employee.profile.is_user:
        messages.error(request, "لا يمكن تعديل موظفين آخرين غير الموظفين العاديين.")
        return redirect("employee_list")

    if request.method == "POST":
        form = EmployeeForm(
            request.POST,
            instance=employee.profile,
            user_faculty=request.user.profile.faculty,
        )
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Update User fields (first_name from form)
                    employee.first_name = form.cleaned_data["first_name"]
                    employee.save()

                    # Update UserProfile fields
                    profile = form.save(commit=False)
                    profile.user = employee
                    profile.faculty = request.user.profile.faculty
                    profile.save()

                messages.success(
                    request, f"تم تحديث بيانات الموظف '{employee.first_name}' بنجاح"
                )
                return redirect("employee_list")
            except Exception as e:
                logger.error(f"Error updating employee: {str(e)}", exc_info=True)
                messages.error(request, f"خطأ في النظام: {str(e)}")
        else:
            for field, errors in form.errors.items():
                logger.error(f"Form error - {field}: {errors}")
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
    else:
        form = EmployeeForm(
            instance=employee.profile,
            user_faculty=request.user.profile.faculty,
        )

    context = {
        "form": form,
        "title": "تعديل بيانات الموظف",
        "faculty": request.user.profile.faculty,
        "editing": True,
        "employee": employee,
    }
    return render(request, "administration/employee_form.html", context)


def _set_user_type(request):
    if request.user.profile.is_user:
        request.session["user_type"] = "user"
    elif request.user.profile.is_inventory_employee:
        request.session["user_type"] = "inventory_employee"
    elif request.user.profile.is_administration_manager:
        request.session["user_type"] = "administration_manager"
    elif request.user.profile.is_faculty_manager:
        request.session["user_type"] = "faculty_manager"
    elif request.user.profile.is_inventory_manager:
        request.session["user_type"] = "inventory_manager"
    else:
        request.session["user_type"] = "unknown"

    return request.session["user_type"]


@login_required
def show_user_home(request):
    return render(request, "administration/user_home.html")


@login_required
def index(request):
    if request.session["user_type"] == "user":
        return show_user_home(request)

    elif request.session["user_type"] == "inventory_employee":
        return show_inventory_employee_home(request)

    elif request.session["user_type"] == "administration_manager":
        return show_administration_manager_home(request)

    elif request.session["user_type"] == "faculty_manager":
        return show_faculty_manager_home(request)

    elif request.session["user_type"] == "inventory_manager":
        return show_inventory_manager_home(request)

    elif request.session["user_type"] == "admin":
        return show_admin_home(request)
    return render(request, "404.html")


def user_login(request: HttpRequest) -> HttpResponse:
    """Handle user authentication with proper session management."""

    # Redirect if already authenticated
    if request.user.is_authenticated:
        return redirect("home")

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()

        # Authenticate user (Django handles session internally)
        logged_user = authenticate(request, username=username, password=password)

        if logged_user is not None and logged_user.is_active:
            # Login the user (Django creates session, sets cookies)
            login(request, logged_user)

            # Reset session timeout tracker on fresh login
            request.session["last_activity"] = time.time()

            # Set user type for non-superusers
            if not logged_user.is_superuser:
                user_type = _set_user_type(request)
                if user_type == "unknown":
                    # Clean logout if user type cannot be determined
                    logout(request)
                    messages.error(request, "بيانات المستخدم غير صحيحة أو غير مكتملة")
                    return redirect("login")
                request.session["user_type"] = user_type
            else:
                request.session["user_type"] = "admin"

            # Redirect to intended page or home
            next_url = request.GET.get("next") or request.POST.get("next") or "home"
            return redirect(next_url)
        else:
            # Clear any partial session data on failed login
            if "username" in request.session:
                del request.session["username"]
            messages.error(request, "اسم المستخدم أو كلمة المرور غير صحيحة")
            return redirect("login")

    # GET request: show login form
    context = {
        "next": request.GET.get("next", "home"),  # Preserve redirect target
    }
    return render(request, "account/login.html", context)


@login_required
def user_logout(request: HttpRequest) -> HttpResponse:
    """Handle user logout with complete session cleanup."""

    # Clear custom session keys before Django logout
    keys_to_clear = ["username", "user_type", "last_activity"]
    for key in keys_to_clear:
        if key in request.session:
            del request.session[key]

    # Django's logout() flushes session, deletes cookie, redirects
    django_logout(request)

    # Redirect to login with clear message
    messages.info(request, "تم تسجيل الخروج بنجاح")
    return redirect("login")


@login_required
def profile(request, user_id):
    """
    Display the profile of a user.
    """

    user = get_object_or_404(UserProfile, pk=user_id)
    context = {"user": user}
    return render(request, "administration/profile.html", context)


@login_required
def password_change_view(request):
    """Handle password change with custom form and messages."""
    if request.method == "POST":
        form = UserPasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            form.save()
            messages.success(
                request, "تم تغيير كلمة المرور بنجاح. يرجى تسجيل الدخول مرة أخرى."
            )
            # Clear session to force re-login
            request.session.flush()
            return redirect("login")
        else:
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
    else:
        form = UserPasswordChangeForm(user=request.user)

    context = {
        "form": form,
        "page_title": "تغيير كلمة المرور",
    }
    return render(request, "administration/password_change.html", context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def system_settings(request):
    settings_obj = SystemSettings.get()
    open_year = InventoryYear.get_open_year()
    years = InventoryYear.objects.all()

    if request.method == "POST":
        form = SessionSettingsForm(request.POST, instance=settings_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "تم حفظ إعدادات الجلسة بنجاح.")
            return redirect("system_settings")
        messages.error(request, "تعذر حفظ إعدادات الجلسة. راجع القيم المدخلة.")
    else:
        form = SessionSettingsForm(instance=settings_obj)

    return render(
        request,
        "administration/system_settings.html",
        {
            "settings_obj": settings_obj,
            "form": form,
            "open_year": open_year,
            "years": years,
            "page_title": "إعدادات النظام",
        },
    )


@login_required
@user_passes_test(lambda u: u.is_superuser)
def close_inventory_year(request, year_id):
    if request.method != "POST":
        return redirect("system_settings")

    year = get_object_or_404(InventoryYear, id=year_id)
    try:
        next_year = year.close_and_open_next()
        messages.success(
            request,
            f"تم إغلاق سنة {year.year} وفتح سنة {next_year.year} بنجاح.",
        )
    except Exception as exc:
        messages.error(request, f"تعذر إغلاق السنة: {exc}")
    return redirect("system_settings")


@login_required
def show_faculty_manager_home(request):
    """Faculty manager dashboard - shows data for user's faculty only."""
    user_profile = getattr(request.user, "profile", None)

    # Check for is_faculty_manager (correct role)
    if not (user_profile and user_profile.is_faculty_manager):
        return HttpResponseForbidden("ليس لديك صلاحية الوصول.")

    # Delegate to inventory manager view (which accepts both roles)
    return show_inventory_manager_home(request)


@login_required
def show_administration_manager_home(request):
    """Administration manager dashboard - FULLY OPTIMIZED."""
    if not (
        hasattr(request.user, "profile")
        and getattr(request.user.profile, "is_administration_manager", False)
    ):
        return HttpResponseForbidden("ليس لديك صلاحية الوصول إلى هذه الصفحة.")

    faculty_stats_data = (
        FacultyItemStock.objects.values(
            "faculty__id",
            "faculty__name",
        )
        .annotate(
            item_count=Count("item_id", distinct=True),
            sub_warehouse_count=Count("sub_warehouse_id", distinct=True),
            category_count=Count("item__category_id", distinct=True),
            out_of_stock=Count(Case(When(cached_quantity=0, then=1))),
            low_stock=Count(
                Case(
                    When(
                        cached_quantity__lte=F("limit_quantity"),
                        cached_quantity__gt=0,
                        then=1,
                    )
                )
            ),
            normal=Count(Case(When(cached_quantity__gt=F("limit_quantity"), then=1))),
        )
        .order_by("faculty__name")
    )

    faculty_ids = [d["faculty__id"] for d in faculty_stats_data]
    faculty_cache = {f.id: f for f in Faculty.objects.filter(id__in=faculty_ids)}

    faculty_stats = []
    all_low_stock = all_out_of_stock = all_normal = 0

    def sanitize_label(name):
        return (
            str(name)
            .replace('"', "")
            .replace("'", "")
            .replace("\\", "")
            .replace("\n", " ")
            .strip()
            or "كلية غير معروفة"
        )

    for data in faculty_stats_data:
        faculty = faculty_cache.get(data["faculty__id"])
        if not faculty:
            continue

        out_of_stock = data["out_of_stock"] or 0
        low_stock = data["low_stock"] or 0
        normal = data["normal"] or 0

        all_low_stock += low_stock
        all_out_of_stock += out_of_stock
        all_normal += normal

        faculty_stats.append(
            {
                "faculty": faculty,
                "faculty_name_safe": sanitize_label(data["faculty__name"]),
                "warehouses": data["sub_warehouse_count"],
                "categories": data["category_count"],
                "items": data["item_count"],
                "low_stock": low_stock,
                "out_of_stock": out_of_stock,
            }
        )

    if not faculty_stats:
        empty = json.dumps(
            {
                "labels": ["لا توجد كليات"],
                "datasets": [{"data": [1], "backgroundColor": ["#ced4da"]}],
            },
            ensure_ascii=False,
        )
        chart_items = chart_warehouses = chart_stock = empty
    else:
        chart_items = json.dumps(
            {
                "labels": [f["faculty_name_safe"] for f in faculty_stats],
                "datasets": [
                    {
                        "data": [f["items"] for f in faculty_stats],
                        "backgroundColor": "#4e73df",
                    }
                ],
            },
            ensure_ascii=False,
        )

        chart_warehouses = json.dumps(
            {
                "labels": [f["faculty_name_safe"] for f in faculty_stats],
                "datasets": [
                    {
                        "data": [f["warehouses"] for f in faculty_stats],
                        "backgroundColor": [
                            "#FF6384",
                            "#36A2EB",
                            "#FFCE56",
                            "#4BC0C0",
                            "#9966FF",
                            "#FF9F40",
                        ][: len(faculty_stats)],
                    }
                ],
            },
            ensure_ascii=False,
        )

        chart_stock = json.dumps(
            {
                "labels": ["غير متوفر", "احتياطي منخفض", "متوفر"],
                "datasets": [
                    {
                        "data": [all_out_of_stock, all_low_stock, all_normal],
                        "backgroundColor": ["#e74a3b", "#f6c23e", "#1cc88a"],
                    }
                ],
            },
            ensure_ascii=False,
        )

    context = {
        "faculty_stats": faculty_stats,
        "total_faculties": len(faculty_stats),  # No COUNT query
        "total_warehouses": sum(f["warehouses"] for f in faculty_stats),
        "total_items": sum(f["items"] for f in faculty_stats),
        "total_low_stock": sum(f["low_stock"] for f in faculty_stats),
        "chart_items_per_faculty": chart_items,
        "chart_warehouses_per_faculty": chart_warehouses,
        "chart_stock_status": chart_stock,
    }
    return render(request, "administration/administration_manager_home.html", context)


@login_required
def show_inventory_manager_home(request):
    """Inventory manager dashboard - FULLY OPTIMIZED (NO N+1 QUERIES)."""
    user_profile = getattr(request.user, "profile", None)
    if not (
        user_profile
        and (user_profile.is_inventory_manager or user_profile.is_faculty_manager)
    ):
        return HttpResponseForbidden("ليس لديك صلاحية الوصول.")

    user = request.user
    faculty = user.profile.faculty
    if not faculty:
        return redirect("home")

    selected_sub_warehouse_id = request.GET.get("sub_warehouse_id")
    selected_sub_warehouse = None

    # Cache sub_warehouses list
    all_sub_warehouses_list = list(
        SubWarehouse.objects.filter(item_stocks__faculty=faculty)
        .select_related("warehouse")
        .distinct()
        .order_by("warehouse__name", "name")
    )
    sub_warehouses_count = len(all_sub_warehouses_list)

    if selected_sub_warehouse_id:
        try:
            selected_sub_warehouse = next(
                (
                    sw
                    for sw in all_sub_warehouses_list
                    if sw.id == int(selected_sub_warehouse_id)
                ),
                None,
            )
        except (ValueError, TypeError):
            pass

    # Initialize context
    per_warehouse_data = []
    low_stock_items = []
    items_count = categories_count = low_stock_count = out_of_stock_count = 0
    chart_category_json = chart_stock_json = None

    if selected_sub_warehouse:
        # Add select_related("category") to avoid N+1
        items_qs = (
            Item.objects.filter(category__sub_warehouse=selected_sub_warehouse)
            .select_related("category")
            .only("id", "name", "category_id", "limit_quantity", "category__name")
        )
        items_list = list(items_qs)
        item_ids = [item.id for item in items_list]

        # Fetch FacultyItemStock in ONE query
        stock_map = {}
        if item_ids:
            stocks = FacultyItemStock.objects.filter(
                faculty=faculty,
                sub_warehouse=selected_sub_warehouse,
                item_id__in=item_ids,
            ).only("item_id", "cached_quantity", "limit_quantity")
            stock_map = {s.item_id: s for s in stocks}

        # Build combined data with pre-fetched category names
        combined_data = []
        category_ids = set()
        oos = low = norm = 0

        for item in items_list:
            stock = stock_map.get(item.id)
            qty = stock.cached_quantity if stock else 0
            limit = stock.limit_quantity if stock else item.limit_quantity

            if item.category_id:
                category_ids.add(item.category_id)

            if qty == 0:
                oos += 1
            elif qty <= limit:
                low += 1
            else:
                norm += 1

            combined_data.append(
                {
                    "item": item,
                    "item_name": item.name,
                    "category_name": item.category.name
                    if item.category
                    else "بدون فئة",
                    "category_id": item.category_id,
                    "quantity": qty,
                    "limit_quantity": limit,
                    "has_stock_record": stock is not None,
                }
            )

        items_count = len(combined_data)
        categories_count = len(category_ids)
        low_stock_count = low + oos
        out_of_stock_count = oos

        # Category chart data
        from collections import defaultdict

        category_counts = defaultdict(int)
        for data in combined_data:
            if data["category_id"]:
                category_counts[data["category_id"]] += 1

        # Get category names in ONE query (already cached from select_related, but safe fallback)
        category_names = {
            cat.id: cat.name
            for cat in ItemCategory.objects.filter(id__in=category_ids).only(
                "id", "name"
            )
        }

        category_data = {
            category_names.get(cat_id, f"Category {cat_id}"): count
            for cat_id, count in category_counts.items()
            if count > 0
        }

        # Build charts
        category_colors = [
            "rgba(78, 115, 223, 0.8)",
            "rgba(54, 162, 235, 0.8)",
            "rgba(255, 206, 86, 0.8)",
        ]
        cat_labels = list(category_data.keys())
        cat_values = list(category_data.values())

        chart_category_json = json.dumps(
            {
                "type": "bar",
                "labels": cat_labels,
                "datasets": [
                    {
                        "label": "عدد الأصناف",
                        "data": cat_values,
                        "backgroundColor": category_colors[: len(cat_labels)],
                        "borderWidth": 1,
                    }
                ],
            },
            ensure_ascii=False,
        )

        chart_stock_json = json.dumps(
            {
                "type": "doughnut",
                "labels": ["نافد", "احتياطي منخفض", "متوفر"],
                "datasets": [
                    {
                        "label": "حالة المخزون",
                        "data": [oos, low, norm],
                        "backgroundColor": [
                            "rgba(231, 74, 59, 0.8)",
                            "rgba(246, 194, 58, 0.8)",
                            "rgba(28, 200, 138, 0.8)",
                        ],
                        "borderWidth": 1,
                        "cutout": "60%",
                    }
                ],
            },
            ensure_ascii=False,
        )

        # Low stock items - pass simple dicts, NOT Item objects
        low_stock_items = [
            {
                "item_name": data["item_name"],
                "category_name": data["category_name"],
                "quantity": data["quantity"],
                "limit_quantity": data["limit_quantity"],
                "has_stock_record": data["has_stock_record"],
            }
            for data in combined_data
            if data["quantity"] <= data["limit_quantity"]
        ]

        per_warehouse_data.append(
            {
                "warehouse": selected_sub_warehouse,
                "items_count": items_count,
                "categories_count": categories_count,
                "low_stock_count": low + oos,
                "out_of_stock_count": oos,
                "chart_category_json": chart_category_json,
                "chart_stock_json": chart_stock_json,
            }
        )

    context = {
        "user_faculty": faculty,
        "sub_warehouses": all_sub_warehouses_list,
        "selected_sub_warehouse": selected_sub_warehouse,
        "selected_sub_warehouse_id": selected_sub_warehouse_id,
        "sub_warehouses_count": sub_warehouses_count,
        "categories_count": categories_count,
        "items_count": items_count,
        "low_stock_count": low_stock_count,
        "out_of_stock_count": out_of_stock_count,
        "per_warehouse_data": per_warehouse_data,
        "is_manager_view": True,
        "low_stock_items": low_stock_items,
        "page_title": "لوحة تحكم مسؤول المخزون",
        "dashboard_type": "manager",
    }
    return render(request, "administration/inventory_manager_home.html", context)


@login_required
def show_inventory_employee_home(request):
    """Inventory employee dashboard - FULLY OPTIMIZED (NO N+1 QUERIES)."""
    user = request.user
    if not (hasattr(user, "profile") and user.profile.faculty):
        return HttpResponseForbidden("ليس لديك كليّة مرتبطة بحسابك.")

    faculty = user.profile.faculty
    selected_sub_warehouse_id = request.GET.get("sub_warehouse_id")

    # Cache sub_warehouses list
    all_sub_warehouses_list = list(
        SubWarehouse.objects.filter(item_stocks__faculty=faculty)
        .select_related("warehouse")
        .distinct()
        .order_by("name")
    )
    warehouses_count = len(all_sub_warehouses_list)

    selected_sub_warehouse = None
    if selected_sub_warehouse_id:
        try:
            selected_sub_warehouse = next(
                (
                    sw
                    for sw in all_sub_warehouses_list
                    if sw.id == int(selected_sub_warehouse_id)
                ),
                None,
            )
        except (ValueError, TypeError):
            pass

    items_count = categories_count = low_stock_count = out_of_stock_count = 0
    chart_category_json = chart_stock_json = None
    low_stock_items = []

    if selected_sub_warehouse:
        # Add select_related("category")
        items_qs = (
            Item.objects.filter(category__sub_warehouse=selected_sub_warehouse)
            .select_related("category")
            .only(
                "id",
                "name",
                "code",
                "category_id",
                "limit_quantity",
                "unit",
                "category__name",
            )
        )
        items_list = list(items_qs)
        item_ids = [item.id for item in items_list]

        # Fetch FacultyItemStock in ONE query
        stock_map = {}
        if item_ids:
            stocks = FacultyItemStock.objects.filter(
                faculty=faculty,
                sub_warehouse=selected_sub_warehouse,
                item_id__in=item_ids,
            ).only("item_id", "cached_quantity", "limit_quantity")
            stock_map = {s.item_id: s for s in stocks}

        # Build combined data with pre-extracted values
        combined_data = []
        oos = low = norm = 0
        category_ids = set()

        for item in items_list:
            stock = stock_map.get(item.id)
            qty = stock.cached_quantity if stock else 0
            limit = stock.limit_quantity if stock else item.limit_quantity

            if item.category_id:
                category_ids.add(item.category_id)

            if qty == 0:
                oos += 1
            elif qty <= limit:
                low += 1
            else:
                norm += 1

            combined_data.append(
                {
                    "item_name": item.name,
                    "item_code": item.code or "-",
                    "category_name": item.category.name
                    if item.category
                    else "بدون فئة",
                    "category_id": item.category_id,
                    "quantity": qty,
                    "limit_quantity": limit,
                    "unit": item.get_unit_display(),
                    "has_stock_record": stock is not None,
                }
            )

        items_count = len(combined_data)
        categories_count = len(category_ids)
        low_stock_count = low + oos
        out_of_stock_count = oos

        # Category chart data
        from collections import defaultdict

        category_counts = defaultdict(int)
        for data in combined_data:
            if data["category_id"]:
                category_counts[data["category_id"]] += 1

        category_names = {
            cat.id: cat.name
            for cat in ItemCategory.objects.filter(id__in=category_ids).only(
                "id", "name"
            )
        }
        category_data = {
            category_names.get(cat_id, f"Category {cat_id}"): count
            for cat_id, count in category_counts.items()
            if count > 0
        }

        # Build charts
        category_colors = [
            "rgba(78, 115, 223, 0.8)",
            "rgba(54, 162, 235, 0.8)",
            "rgba(255, 206, 86, 0.8)",
        ]
        sorted_cats = sorted(category_data.items(), key=lambda x: x[1], reverse=True)
        cat_labels = [name for name, _ in sorted_cats]
        cat_values = [count for _, count in sorted_cats]

        chart_category_json = json.dumps(
            {
                "type": "bar",
                "labels": cat_labels,
                "datasets": [
                    {
                        "label": "عدد الأصناف",
                        "data": cat_values,
                        "backgroundColor": category_colors[: len(cat_labels)],
                        "borderWidth": 1,
                    }
                ],
            },
            ensure_ascii=False,
        )

        chart_stock_json = json.dumps(
            {
                "type": "doughnut",
                "labels": ["نافد", "احتياطي منخفض", "متوفر"],
                "datasets": [
                    {
                        "label": "حالة المخزون",
                        "data": [oos, low, norm],
                        "backgroundColor": [
                            "rgba(231, 74, 59, 0.8)",
                            "rgba(246, 194, 58, 0.8)",
                            "rgba(28, 200, 138, 0.8)",
                        ],
                        "borderWidth": 1,
                        "cutout": "60%",
                    }
                ],
            },
            ensure_ascii=False,
            cls=DjangoJSONEncoder,
        )

        # Low stock items - pass dicts, NOT Item objects
        low_stock_items = [
            {
                "item_name": data["item_name"],
                "item_code": data["item_code"],
                "category_name": data["category_name"],
                "quantity": data["quantity"],
                "limit_quantity": data["limit_quantity"],
                "unit": data["unit"],
            }
            for data in combined_data
            if data["quantity"] <= data["limit_quantity"]
        ]

    if not selected_sub_warehouse:
        items_count = (
            FacultyItemStock.objects.filter(faculty=faculty)
            .values("item_id")
            .distinct()
            .count()
        )

    context = {
        "user_faculty": faculty,
        "sub_warehouses": all_sub_warehouses_list,
        "selected_sub_warehouse": selected_sub_warehouse,
        "selected_sub_warehouse_id": selected_sub_warehouse_id,
        "chart_category_json": chart_category_json,
        "chart_stock_json": chart_stock_json,
        "items_count": items_count,
        "categories_count": categories_count,
        "low_stock_count": low_stock_count,
        "out_of_stock_count": out_of_stock_count,
        "warehouses_count": warehouses_count,
        "low_stock_items": low_stock_items,
    }
    return render(request, "administration/inventory_employee_home.html", context)


@login_required
def administration_item_search(request):
    """Administration manager dashboard - Search items across ALL faculties. OPTIMIZED."""
    user_profile = getattr(request.user, "profile", None)
    if not (user_profile and getattr(user_profile, "is_administration_manager", False)):
        return HttpResponseForbidden("ليس لديك صلاحية الوصول إلى هذه الصفحة.")

    search_query = request.GET.get("q", "").strip()
    faculty_filter = request.GET.get("faculty_id", "")
    category_filter = request.GET.get("category_id", "")

    # Base queryset with select_related
    items_qs = Item.objects.select_related(
        "category",
        "category__sub_warehouse",  # ForeignKey chain
    ).order_by("name")

    if search_query:
        items_qs = items_qs.filter(
            Q(name__icontains=search_query) | Q(code__icontains=search_query)
        )
    if faculty_filter and faculty_filter.isdigit():
        items_qs = items_qs.filter(faculty_stocks__faculty_id=faculty_filter).distinct()
    if category_filter and category_filter.isdigit():
        items_qs = items_qs.filter(category_id=category_filter)

    # Pagination
    paginator = Paginator(items_qs, 20)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    # Get dropdowns (cached)
    faculties = list(Faculty.objects.all().order_by("name"))
    categories = list(ItemCategory.objects.all().order_by("name"))

    # Get all FacultyItemStock for page items in ONE query
    items_with_faculty_qty = []
    if search_query or faculty_filter or category_filter:
        # Get all item IDs on this page
        page_item_ids = [item.id for item in page_obj]

        all_stocks = (
            FacultyItemStock.objects.filter(item_id__in=page_item_ids)
            .select_related("faculty", "sub_warehouse")
            .order_by("faculty__name")
        )

        # Group stocks by item_id in Python
        stocks_by_item = {}
        for stock in all_stocks:
            if stock.item_id not in stocks_by_item:
                stocks_by_item[stock.item_id] = []
            stocks_by_item[stock.item_id].append(stock)

        # Build response from cached data
        for item in page_obj:
            stocks = stocks_by_item.get(item.id, [])
            total_qty = sum(s.cached_quantity for s in stocks)

            faculty_qty_data = [
                {
                    "faculty_name": s.faculty.name,
                    "faculty_id": s.faculty.id,
                    "quantity": s.cached_quantity,
                    "sub_warehouse": s.sub_warehouse.name if s.sub_warehouse else "-",
                }
                for s in stocks
            ]

            items_with_faculty_qty.append(
                {
                    "item": item,
                    "total_quantity": total_qty,
                    "faculty_count": len(stocks),
                    "faculty_qty_data": faculty_qty_data,
                }
            )

    context = {
        "search_query": search_query,
        "faculties": faculties,
        "categories": categories,
        "selected_faculty_id": faculty_filter if faculty_filter.isdigit() else "",
        "selected_category_id": category_filter if category_filter.isdigit() else "",
        "items_with_faculty_qty": items_with_faculty_qty,
        "page_obj": page_obj,
        "total_items": paginator.count,
    }
    return render(request, "administration/administration_item_search.html", context)


@login_required
def administration_item_search_api(request):
    """
    API endpoint for AJAX search - returns JSON with item quantities per faculty.
    Optimized for fast response with minimal queries.
    """
    user_profile = getattr(request.user, "profile", None)
    if not (user_profile and getattr(user_profile, "is_administration_manager", False)):
        return JsonResponse({"error": "غير مسموح"}, status=403)

    search_query = request.GET.get("q", "").strip()

    if not search_query or len(search_query) < 2:
        return JsonResponse(
            {"items": [], "message": "يرجى إدخال حرفين على الأقل للبحث"}
        )

    # Search items (global catalog)
    items = (
        Item.objects.filter(
            Q(name__icontains=search_query) | Q(code__icontains=search_query)
        )
        .select_related("category", "category__sub_warehouse")  # ForeignKey
        .order_by("name")[:50]
    )  # Limit to 50 results

    # Get all faculties once
    faculties = list(Faculty.objects.all().order_by("name"))

    # Get all FacultyItemStock for these items in ONE query
    all_stocks = FacultyItemStock.objects.filter(item__in=items).select_related(
        "faculty", "sub_warehouse"
    )

    # Group stocks by item_id
    stocks_by_item = {}
    for stock in all_stocks:
        if stock.item_id not in stocks_by_item:
            stocks_by_item[stock.item_id] = []
        stocks_by_item[stock.item_id].append(stock)

    # Build response
    items_data = []
    for item in items:
        stocks = stocks_by_item.get(item.id, [])
        total_qty = sum(s.cached_quantity for s in stocks)

        faculty_qty = []
        for faculty in faculties:
            faculty_stock = next(
                (s for s in stocks if s.faculty_id == faculty.id), None
            )
            faculty_qty.append(
                {
                    "faculty_id": faculty.id,
                    "faculty_name": faculty.name,
                    "quantity": faculty_stock.cached_quantity if faculty_stock else 0,
                }
            )

        items_data.append(
            {
                "item_id": item.id,
                "item_name": item.name,
                "item_code": item.code or "-",
                "category": item.category.name if item.category else "-",
                "category_sub_warehouse": item.category.sub_warehouse.name
                if item.category and item.category.sub_warehouse
                else "-",
                "total_quantity": total_qty,
                "faculty_count": len([s for s in stocks if s.cached_quantity > 0]),
                "faculty_qty": faculty_qty,
            }
        )

    return JsonResponse(
        {
            "items": items_data,
            "total_results": len(items_data),
            "search_query": search_query,
        }
    )


@login_required
def admin_charts_data(request):
    """
    Returns chart data as JSON for the administration manager dashboard.
    OPTIMIZED: Single query with annotations.
    """
    if not (
        hasattr(request.user, "profile")
        and getattr(request.user.profile, "is_administration_manager", False)
    ):
        return JsonResponse({"error": "غير مسموح"}, status=403)

    faculty_stats_data = (
        FacultyItemStock.objects.values(
            "faculty__id",
            "faculty__name",
        )
        .annotate(
            item_count=Count("item_id", distinct=True),
            sub_warehouse_count=Count("sub_warehouse_id", distinct=True),
            out_of_stock=Count(Case(When(cached_quantity=0, then=1))),
            low_stock=Count(
                Case(
                    When(
                        cached_quantity__lte=F("limit_quantity"),
                        cached_quantity__gt=0,
                        then=1,
                    )
                )
            ),
            normal=Count(Case(When(cached_quantity__gt=F("limit_quantity"), then=1))),
        )
        .order_by("faculty__name")
    )

    faculty_stats = []
    all_low_stock = all_out_of_stock = all_normal = 0

    def sanitize_label(name):
        return (
            str(name)
            .replace('"', "")
            .replace("'", "")
            .replace("\\", "")
            .replace("\n", " ")
            .strip()
            or "كلية غير معروفة"
        )

    for data in faculty_stats_data:
        # faculty = Faculty.objects.get(id=data["faculty__id"])

        out_of_stock = data["out_of_stock"]
        low_stock = data["low_stock"]
        normal = data["normal"]

        all_low_stock += low_stock
        all_out_of_stock += out_of_stock
        all_normal += normal

        faculty_stats.append(
            {
                "faculty_name_safe": sanitize_label(data["faculty__name"]),
                "items": data["item_count"],
                "warehouses": data["sub_warehouse_count"],
            }
        )

    # Prepare chart data
    items_chart = {
        "labels": [f["faculty_name_safe"] for f in faculty_stats],
        "datasets": [
            {
                "data": [f["items"] for f in faculty_stats],
                "backgroundColor": "#4e73df",
            }
        ],
    }

    warehouses_chart = {
        "labels": [f["faculty_name_safe"] for f in faculty_stats],
        "datasets": [
            {
                "data": [f["warehouses"] for f in faculty_stats],
                "backgroundColor": [
                    "#FF6384",
                    "#36A2EB",
                    "#FFCE56",
                    "#4BC0C0",
                    "#9966FF",
                    "#FF9F40",
                ][: len(faculty_stats)],
            }
        ],
    }

    stock_chart = {
        "labels": ["غير متوفر", "احتياطي منخفض", "متوفر"],
        "datasets": [
            {
                "data": [all_out_of_stock, all_low_stock, all_normal],
                "backgroundColor": ["#e74a3b", "#f6c23e", "#1cc88a"],
            }
        ],
    }

    return JsonResponse(
        {
            "items_chart": items_chart,
            "warehouses_chart": warehouses_chart,
            "stock_chart": stock_chart,
        }
    )


@login_required
def show_admin_home(request):
    """Admin home dashboard."""

    return redirect("admin_user_list")
    users = (
        User.objects.only("id", "username", "first_name", "last_name", "is_active")
        .prefetch_related("profile")
        .all()
    )

    if hasattr(request.user, "profile") and request.user.profile.faculty:
        faculty = request.user.profile.faculty
        low_stock_items = (
            FacultyItemStock.objects.filter(
                faculty=faculty, cached_quantity__lte=F("limit_quantity")
            )
            .select_related("item", "item__category", "sub_warehouse")
            .order_by("cached_quantity")
        )
    else:
        low_stock_items = FacultyItemStock.objects.none()
    db_engine = settings.DATABASES.get("default", {}).get("ENGINE", "")
    is_postgres = "postgresql" in db_engine.lower()
    context = {
        "low_stock_items": low_stock_items,
        "users": users,
        "is_postgres": is_postgres,
    }
    return render(request, "administration/index.html", context)


@login_required
def export_low_stock_items_pdf(request):
    """
    Export low stock items PDF.
    - Mandatory: Faculty isolation (user's faculty only)
    - Optional: Filter by selected sub_warehouse (must belong to user's faculty)
    """
    user = request.user

    # Step 1: Faculty isolation (mandatory)
    if not (hasattr(user, "profile") and user.profile.faculty):
        return HttpResponseForbidden("ليس لديك كليّة مرتبطة بحسابك.")

    faculty = user.profile.faculty
    selected_sub_warehouse_id = request.GET.get("sub_warehouse_id")
    selected_sub_warehouse = None

    try:
        # Step 2: Validate sub_warehouse belongs to user's faculty
        # FIX 1: Use .filter().first() instead of get_object_or_404()
        #        because the JOIN creates duplicate rows
        if selected_sub_warehouse_id and selected_sub_warehouse_id.isdigit():
            selected_sub_warehouse = (
                SubWarehouse.objects.filter(
                    id=selected_sub_warehouse_id,
                    item_stocks__faculty=faculty,  # Ensures faculty isolation
                )
                .select_related("warehouse")
                .distinct()
                .first()
            )

            # If no sub_warehouse found, return 404
            if not selected_sub_warehouse:
                return HttpResponseForbidden(
                    "المخزن الفرعي غير موجود أو لا ينتمي لكليتك."
                )

        # Step 3: Build queryset with faculty isolation
        low_stock_items = FacultyItemStock.objects.filter(
            faculty=faculty,  # Mandatory
            cached_quantity__lte=F("limit_quantity"),
        )

        # Step 4: Apply sub_warehouse filter if selected
        if selected_sub_warehouse:
            low_stock_items = low_stock_items.filter(
                sub_warehouse=selected_sub_warehouse
            )

        # Step 5: Optimize queries (NO .only() conflict)
        low_stock_items = (
            low_stock_items.select_related(
                "item",
                "item__category",
                "sub_warehouse",
                "sub_warehouse__warehouse",
            ).order_by("item__category__name", "item__name")
        )[:500]  # Limit to prevent timeout

        # Step 6: Evaluate query
        low_stock_items_list = list(low_stock_items)
        total_count = len(low_stock_items_list)

        if total_count == 0:
            messages.warning(request, "لا توجد أصناف منخفضة المخزون.")

        # Step 7: Render template
        html_string = render_to_string(
            "administration/reports/low_stock_items_report.html",
            {
                "low_stock_items": low_stock_items_list,
                "total_count": total_count,
                "current_date": datetime.now(),
                "user": user,
                "faculty": faculty,
                "selected_sub_warehouse": selected_sub_warehouse,
            },
            request=request,
        )

        # Step 8: Generate PDF
        html = HTML(string=html_string, base_url=request.build_absolute_uri("/"))
        response = HttpResponse(content_type="application/pdf")

        # Dynamic filename
        if selected_sub_warehouse:
            filename = f"low_stock_{selected_sub_warehouse.name}_{datetime.now().strftime('%Y%m%d')}.pdf"
        else:
            filename = (
                f"low_stock_{faculty.name}_{datetime.now().strftime('%Y%m%d')}.pdf"
            )

        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        html.write_pdf(response)

        # FIX 2: Remove emoji from logger (Windows console encoding issue)
        logger.info(
            f"[SUCCESS] PDF exported | Faculty: {faculty.name} | "
            f"Sub-Warehouse: {selected_sub_warehouse.name if selected_sub_warehouse else 'All'} | "
            f"Items: {total_count}"
        )
        return response

    except Exception as e:
        # FIX 2: Remove emoji from logger (Windows console encoding issue)
        logger.error(
            f"[ERROR] PDF export failed | Faculty: {faculty.name} | Error: {str(e)}",
            exc_info=True,
        )
        messages.error(request, "حدث خطأ أثناء توليد ملف PDF.")

        # Safe redirect
        referrer = request.META.get("HTTP_REFERER")
        if referrer:
            return redirect(referrer)
        else:
            return redirect("home")


@login_required
@require_GET
def get_subwarehouse_charts(request, subwarehouse_id):
    """API endpoint to get chart data for a specific sub-warehouse."""
    try:
        # Get the sub-warehouse (global)
        sub_warehouse = get_object_or_404(SubWarehouse, id=subwarehouse_id)

        # Get user's faculty for isolation
        faculty = getattr(getattr(request.user, "profile", None), "faculty", None)
        if not faculty:
            return JsonResponse({"error": "حسابك غير مرتبط بكليّة"}, status=403)

        # Filter FacultyItemStock by faculty + sub_warehouse
        stocks = FacultyItemStock.objects.filter(
            faculty=faculty,
            sub_warehouse=sub_warehouse,
        ).select_related("item", "item__category")

        # Category chart data - items count per category
        # Note: ItemCategory.sub_warehouse is ForeignKey, so we filter by that
        category_data = (
            stocks.values("item__category__name")
            .annotate(item_count=Count("item_id", distinct=True))
            .order_by("-item_count")
        )

        categories = [item["item__category__name"] for item in category_data]
        category_counts = [item["item_count"] for item in category_data]

        # Stock status chart data
        total_items = stocks.count()
        if total_items == 0:
            return JsonResponse(
                {"error": "لا توجد أصناف في هذا المخزن الفرعي"}, status=404
            )

        # Calculate stock status percentages
        low_stock_items = stocks.filter(
            cached_quantity__lte=F("limit_quantity")
        ).count()
        out_of_stock_items = stocks.filter(cached_quantity=0).count()
        available_items = total_items - low_stock_items - out_of_stock_items

        # Prepare chart data
        chart_category = {
            "labels": categories,
            "datasets": [
                {
                    "label": "عدد الأصناف",
                    "data": category_counts,
                    "backgroundColor": [
                        "rgba(54, 162, 235, 0.8)",
                        "rgba(75, 192, 192, 0.8)",
                        "rgba(255, 206, 86, 0.8)",
                        "rgba(153, 102, 255, 0.8)",
                        "rgba(255, 159, 64, 0.8)",
                        "rgba(201, 203, 207, 0.8)",
                        "rgba(255, 99, 132, 0.8)",
                        "rgba(54, 162, 235, 0.6)",
                        "rgba(75, 192, 192, 0.6)",
                        "rgba(255, 206, 86, 0.6)",
                    ],
                    "borderColor": [
                        "rgba(54, 162, 235, 1)",
                        "rgba(75, 192, 192, 1)",
                        "rgba(255, 206, 86, 1)",
                        "rgba(153, 102, 255, 1)",
                        "rgba(255, 159, 64, 1)",
                        "rgba(201, 203, 207, 1)",
                        "rgba(255, 99, 132, 1)",
                        "rgba(54, 162, 235, 0.8)",
                        "rgba(75, 192, 192, 0.8)",
                        "rgba(255, 206, 86, 0.8)",
                    ],
                    "borderWidth": 1,
                }
            ],
        }

        chart_stock = {
            "labels": ["متوفر", "منخفض الكمية", "نافد"],
            "datasets": [
                {
                    "label": "حالة المخزون",
                    "data": [available_items, low_stock_items, out_of_stock_items],
                    "backgroundColor": [
                        "rgba(40, 167, 69, 0.8)",  # Green for available
                        "rgba(255, 193, 7, 0.8)",  # Yellow for low stock
                        "rgba(220, 53, 69, 0.8)",  # Red for out of stock
                    ],
                    "borderColor": [
                        "rgba(40, 167, 69, 1)",
                        "rgba(255, 193, 7, 1)",
                        "rgba(220, 53, 69, 1)",
                    ],
                    "borderWidth": 1,
                }
            ],
        }

        # Calculate total quantity
        total_quantity = stocks.aggregate(total=Sum("cached_quantity"))["total"] or 0

        # Prepare response data
        response_data = {
            "sub_warehouse_name": sub_warehouse.name,
            "items_count": total_items,
            "total_quantity": total_quantity,
            "low_stock_count": low_stock_items,
            "out_of_stock_count": out_of_stock_items,
            "chart_category": chart_category,
            "chart_stock": chart_stock,
        }

        return JsonResponse(response_data, encoder=DjangoJSONEncoder)

    except SubWarehouse.DoesNotExist:
        return JsonResponse({"error": "المخزن الفرعي غير موجود"}, status=404)
    except Exception as e:
        return JsonResponse({"error": f"حدث خطأ: {str(e)}"}, status=500)


@login_required
@require_GET
def get_subwarehouse_charts_api(request, subwarehouse_id):
    """API endpoint to get chart data for a specific sub-warehouse."""
    try:
        # Get user's faculty for security
        if not hasattr(request.user, "profile") or not request.user.profile.faculty:
            return JsonResponse({"error": "حسابك غير مرتبط بكليّة"}, status=403)

        faculty = request.user.profile.faculty

        # Get the sub-warehouse (global)
        sub_warehouse = get_object_or_404(SubWarehouse, id=subwarehouse_id)

        # Filter FacultyItemStock by faculty + sub_warehouse
        stocks = FacultyItemStock.objects.filter(
            faculty=faculty,
            sub_warehouse=sub_warehouse,
        ).select_related("item", "item__category")

        # Category chart data - count by category name
        category_counts = {}
        for stock in stocks:
            cat_name = stock.item.category.name if stock.item.category else "بدون فئة"
            category_counts[cat_name] = category_counts.get(cat_name, 0) + 1

        # Stock status data
        oos = low = norm = 0
        for stock in stocks:
            qty = stock.cached_quantity
            if qty <= 0:
                oos += 1
            elif qty <= stock.limit_quantity:
                low += 1
            else:
                norm += 1

        # Prepare chart data
        chart_category = {
            "labels": list(category_counts.keys()),
            "datasets": [
                {
                    "label": "عدد الأصناف",
                    "data": list(category_counts.values()),
                    "backgroundColor": [
                        "rgba(78, 115, 223, 0.8)",
                        "rgba(54, 162, 235, 0.8)",
                        "rgba(255, 206, 86, 0.8)",
                        "rgba(153, 102, 255, 0.8)",
                        "rgba(255, 159, 64, 0.8)",
                        "rgba(75, 192, 192, 0.8)",
                        "rgba(201, 203, 207, 0.8)",
                        "rgba(255, 99, 132, 0.8)",
                        "rgba(199, 199, 199, 0.8)",
                        "rgba(83, 102, 199, 0.8)",
                    ],
                    "borderColor": [
                        "rgba(78, 115, 223, 1)",
                        "rgba(54, 162, 235, 1)",
                        "rgba(255, 206, 86, 1)",
                        "rgba(153, 102, 255, 1)",
                        "rgba(255, 159, 64, 1)",
                        "rgba(75, 192, 192, 1)",
                        "rgba(201, 203, 207, 1)",
                        "rgba(255, 99, 132, 1)",
                        "rgba(199, 199, 199, 1)",
                        "rgba(83, 102, 199, 1)",
                    ],
                    "borderWidth": 1,
                }
            ],
        }

        chart_stock = {
            "labels": ["نافد", "احتياطي منخفض", "متوفر"],
            "datasets": [
                {
                    "label": "حالة المخزون",
                    "data": [oos, low, norm],
                    "backgroundColor": [
                        "rgba(231, 74, 59, 0.8)",
                        "rgba(246, 194, 58, 0.8)",
                        "rgba(28, 200, 138, 0.8)",
                    ],
                    "borderColor": [
                        "rgba(231, 74, 59, 1)",
                        "rgba(246, 194, 58, 1)",
                        "rgba(28, 200, 138, 1)",
                    ],
                    "borderWidth": 1,
                }
            ],
        }

        # Get low stock items
        low_stock_items = stocks.filter(
            cached_quantity__lte=F("limit_quantity")
        ).order_by("cached_quantity")

        low_stock_items_data = [
            {
                "id": stock.item.id,
                "name": stock.item.name,
                "category": stock.item.category.name
                if stock.item.category
                else "بدون فئة",
                "quantity": stock.cached_quantity,
                "limit": stock.limit_quantity,
                "status": "نافد" if stock.cached_quantity <= 0 else "منخفض",
            }
            for stock in low_stock_items
        ]

        response_data = {
            "sub_warehouse_name": sub_warehouse.name,
            "warehouse_name": sub_warehouse.warehouse.name,
            "items_count": stocks.count(),
            "categories_count": len(category_counts),
            "low_stock_count": low + oos,
            "out_of_stock_count": oos,
            "chart_category": chart_category,
            "chart_stock": chart_stock,
            "low_stock_items": low_stock_items_data,
        }

        return JsonResponse(response_data)

    except SubWarehouse.DoesNotExist:
        return JsonResponse({"error": "المخزن الفرعي غير موجود"}, status=404)
    except Exception as e:
        return JsonResponse({"error": f"حدث خطأ: {str(e)}"}, status=500)


@login_required
def export_users_roles_report(request):
    """Export users roles to PDF using WeasyPrint with optional filtering."""

    # Get users
    users = (
        User.objects.only(
            "id",
            "username",
            "first_name",
            "last_name",
            "is_active",
            "is_staff",
            "is_superuser",
        )
        .prefetch_related("profile")
        .all()
    )

    context = {
        "users": users,
        "current_date": datetime.now(),
    }

    return create_pdf_report(
        request=request,
        query=users,
        template="administration/reports/export_users_roles_report.html",
        context=context,
        pdf_filename="users_roles_report.pdf",
        is_page_break=False,
    )


def is_superuser(user):
    return user.is_superuser


@login_required
@user_passes_test(is_superuser)
def admin_user_list(request):
    users = User.objects.select_related(
        "profile__faculty", "profile__department"
    ).order_by("first_name")
    db_engine = settings.DATABASES["default"]["ENGINE"]
    is_postgres = "postgresql" in db_engine.lower()
    return render(
        request,
        "administration/admin_user_list.html",
        {"users": users, "is_postgres": is_postgres},
    )


@login_required
@user_passes_test(is_superuser)
def admin_user_create(request):
    if request.method == "POST":
        form = UserAdminForm(request.POST)
        if form.is_valid():
            user = User.objects.create_user(
                username=form.cleaned_data["username"],
                password=form.cleaned_data.get("password") or "TempPass123!",
                first_name=form.cleaned_data.get("first_name", ""),
                last_name=form.cleaned_data.get("last_name", ""),
                email=form.cleaned_data.get("email", ""),
                is_active=form.cleaned_data.get("is_active", True),
                is_staff=form.cleaned_data.get("is_staff", False),
                is_superuser=form.cleaned_data.get("is_superuser", False),
            )
            profile, _ = UserProfile.objects.get_or_create(user=user)
            for field in [
                "faculty",
                "department",
                "phone",
                "is_user",
                "is_inventory_manager",
                "is_inventory_employee",
                "is_administration_manager",
                "is_faculty_manager",
            ]:
                if field in form.cleaned_data:
                    setattr(profile, field, form.cleaned_data[field])
            profile.save()
            messages.success(request, f'تم إنشاء المستخدم "{user.username}" بنجاح')
            return redirect("admin_user_list")
    else:
        form = UserAdminForm()
    return render(
        request,
        "administration/admin_form.html",
        {"form": form, "title": "إضافة مستخدم جديد"},
    )


@login_required
@user_passes_test(is_superuser)
def admin_user_edit(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        form = UserAdminForm(request.POST, instance=user)
        if form.is_valid():
            # Handle password change
            if form.cleaned_data.get("password"):
                user.set_password(form.cleaned_data["password"])

            # Save User fields
            user.save()

            # Get or create profile
            profile, _ = UserProfile.objects.get_or_create(user=user)

            # Update profile fields from form
            for field in [
                "faculty",
                "department",
                "phone",
                "is_user",
                "is_inventory_manager",
                "is_inventory_employee",
                "is_administration_manager",
                "is_faculty_manager",
            ]:
                if field in form.cleaned_data:
                    setattr(profile, field, form.cleaned_data[field])

            profile.save()

            messages.success(request, f'تم تحديث المستخدم "{user.username}" بنجاح')
            return redirect("admin_user_list")
    else:
        # Load profile data into form initial values
        profile = getattr(user, "profile", None)

        initial_data = {
            "faculty": profile.faculty.id if profile and profile.faculty else None,
            "department": profile.department.id
            if profile and profile.department
            else None,
            "phone": profile.phone if profile else "",
            "is_user": profile.is_user if profile else True,
            "is_inventory_manager": profile.is_inventory_manager if profile else False,
            "is_inventory_employee": profile.is_inventory_employee
            if profile
            else False,
            "is_administration_manager": profile.is_administration_manager
            if profile
            else False,
            "is_faculty_manager": profile.is_faculty_manager if profile else False,
        }

        form = UserAdminForm(instance=user, initial=initial_data)

    return render(
        request,
        "administration/admin_form.html",
        {"form": form, "title": "تعديل مستخدم", "editing": True},
    )


@require_GET
@login_required
def get_departments_by_faculty(request):
    """Return departments filtered by faculty_id for AJAX dropdown population."""
    faculty_id = request.GET.get("faculty_id")

    if not faculty_id or not faculty_id.isdigit():
        return JsonResponse({"departments": []})

    departments = (
        Department.objects.filter(faculty_id=int(faculty_id))
        .order_by("name")
        .values("id", "name")
    )

    return JsonResponse({"departments": list(departments), "count": len(departments)})


@login_required
@user_passes_test(is_superuser)
def admin_user_delete(request, user_id):
    user = get_object_or_404(User, id=user_id)

    # Prevent deleting self or superusers
    if user == request.user or user.is_superuser:
        messages.error(request, "لا يمكن حذف هذا المستخدم")
        return redirect("admin_user_list")

    # Check for related transactions BEFORE showing confirm page
    has_tx, msg = _has_related_transactions("user", user_id)
    if has_tx:
        messages.error(request, msg)
        return redirect("admin_user_list")

    if request.method == "POST":
        username = user.username
        user.delete()
        messages.success(request, f'تم حذف المستخدم "{username}" بنجاح')
        return redirect("admin_user_list")

    return render(
        request,
        "administration/admin_confirm_delete.html",
        {"object": user, "obj_name": f"المستخدم {user.username}", "entity": "user"},
    )


def is_superuser(user):
    return user.is_superuser


def _parse_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() in ("true", "1", "yes", "y")


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_user_import_excel(request):
    """Handle Excel file upload and batch user creation (Windows-safe)."""
    if request.method != "POST" or not request.FILES.get("excel_file"):
        messages.error(request, "يرجى اختيار ملف إكسل صالح")
        return redirect("admin_user_list")

    excel_file = request.FILES["excel_file"]
    dry_run = request.POST.get("dry_run") == "on"
    update_existing = request.POST.get("update_existing") == "on"
    skip_existing = request.POST.get("skip_existing") == "on"

    tmp_path = None
    created = updated = skipped = errors = 0
    error_list = []

    try:
        # Save to temp file (Windows-safe pattern)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            for chunk in excel_file.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name
        # File is AUTOMATICALLY CLOSED here by the 'with' block

        # Now safe to open with openpyxl (no lock conflict)
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb.active

        headers = [str(cell.value).strip().lower() for cell in ws[1] if cell.value]
        required = ["username", "first_name", "faculty_id", "department_id"]
        missing = [f for f in required if f not in headers]
        if missing:
            messages.error(request, f"أعمدة مفقودة في الملف: {', '.join(missing)}")
            return redirect("admin_user_list")

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(cell for cell in row):
                continue

            row_data = {
                headers[i]: cell
                for i, cell in enumerate(row)
                if i < len(headers) and headers[i]
            }

            try:
                username = str(row_data.get("username", "")).strip()
                first_name = str(row_data.get("first_name", "")).strip()
                password = (
                    str(row_data.get("password", DEFAULT_PASSWORD)).strip()
                    or DEFAULT_PASSWORD
                )

                faculty_id = row_data.get("faculty_id")
                department_id = row_data.get("department_id")
                if not faculty_id or not department_id:
                    raise ValueError("مفقود: faculty_id أو department_id")

                faculty_id, department_id = int(faculty_id), int(department_id)

                is_active = _parse_bool(row_data.get("is_active"), True)
                is_staff = _parse_bool(row_data.get("is_staff"), False)
                is_superuser_flag = _parse_bool(row_data.get("is_superuser"), False)
                is_inventory_manager = _parse_bool(
                    row_data.get("is_inventory_manager"), False
                )
                is_inventory_employee = _parse_bool(
                    row_data.get("is_inventory_employee"), False
                )
                is_administration_manager = _parse_bool(
                    row_data.get("is_administration_manager"), False
                )
                is_faculty_manager = _parse_bool(
                    row_data.get("is_faculty_manager"), False
                )
                is_user = _parse_bool(row_data.get("is_user"), True)

                faculty = Faculty.objects.filter(id=faculty_id).first()
                if not faculty:
                    raise ValueError(f"الكلية {faculty_id} غير موجودة")
                department = Department.objects.filter(id=department_id).first()
                if not department:
                    raise ValueError(f"القسم {department_id} غير موجود")
                if department.faculty_id != faculty_id:
                    raise ValueError(
                        f"القسم {department_id} لا ينتمي للكلية {faculty_id}"
                    )

                with transaction.atomic():
                    user, user_created = User.objects.get_or_create(
                        username=username,
                        defaults={
                            "first_name": first_name,
                            "is_active": is_active,
                            "is_staff": is_staff,
                            "is_superuser": is_superuser_flag,
                        },
                    )

                    if not user_created:
                        if skip_existing:
                            skipped += 1
                            continue
                        elif update_existing:
                            changed = any(
                                getattr(user, k) != v
                                for k, v in {
                                    "first_name": first_name,
                                    "is_active": is_active,
                                    "is_staff": is_staff,
                                    "is_superuser": is_superuser_flag,
                                }.items()
                            )
                            if changed and not dry_run:
                                for k, v in {
                                    "first_name": first_name,
                                    "is_active": is_active,
                                    "is_staff": is_staff,
                                    "is_superuser": is_superuser_flag,
                                }.items():
                                    setattr(user, k, v)
                                user.save()
                            updated += 1
                        else:
                            skipped += 1
                            continue

                    if (user_created or update_existing) and not dry_run:
                        user.set_password(password)
                        user.save(
                            update_fields=["password"] if not user_created else None
                        )

                    profile, _ = UserProfile.objects.get_or_create(
                        user=user,
                        defaults={"faculty": faculty, "department": department},
                    )
                    if not dry_run:
                        for k, v in {
                            "is_inventory_manager": is_inventory_manager,
                            "is_inventory_employee": is_inventory_employee,
                            "is_administration_manager": is_administration_manager,
                            "is_faculty_manager": is_faculty_manager,
                            "is_user": is_user,
                        }.items():
                            setattr(profile, k, v)
                        profile.faculty, profile.department = faculty, department
                        profile.save()
                    if user_created:
                        created += 1

            except Exception as e:
                errors += 1
                error_list.append(f"صف {row_idx}: {e}")
                logger.error(
                    f"Row {row_idx} error: {e} | Username: {row_data.get('username', 'N/A')}"
                )

    except Exception as e:
        logger.error(f"Excel processing error: {e}", exc_info=True)
        messages.error(request, f"خطأ في معالجة الملف: {str(e)}")
        return redirect("admin_user_list")
    finally:
        # Windows-safe cleanup (handles OS/AV locks gracefully)
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except PermissionError:
                # Windows Defender/Antivirus may briefly lock the file after use
                logger.warning(
                    f"Temp file locked by OS/AV, will be auto-cleaned: {tmp_path}"
                )

    summary = f"✅ تم الاستيراد: {created} جديد، {updated} محدث، {skipped} مُتخطى، {errors} أخطاء"
    if dry_run:
        summary = f"🔍 معاينة فقط: {created} سيتم إنشاؤها، {updated} سيتم تحديثها، {skipped} سيتم تخطيها، {errors} أخطاء"
    if errors > 0:
        summary += f"\n⚠️ أول 3 أخطاء: {' | '.join(error_list[:3])}"

    if dry_run:
        messages.warning(request, summary, extra_tags="safe")
    else:
        messages.success(request, summary, extra_tags="safe")

    return redirect("admin_user_list")


@login_required
@user_passes_test(is_superuser)
def admin_faculty_list(request):
    faculties = Faculty.objects.all().order_by("name")
    return render(
        request, "administration/admin_faculty_list.html", {"faculties": faculties}
    )


@login_required
@user_passes_test(is_superuser)
def admin_faculty_create(request):
    if request.method == "POST":
        form = FacultyForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "تم إضافة الكلية بنجاح")
            return redirect("admin_faculty_list")
    else:
        form = FacultyForm()
    return render(
        request, "administration/admin_form.html", {"form": form, "title": "إضافة كلية"}
    )


@login_required
@user_passes_test(is_superuser)
def admin_faculty_edit(request, faculty_id):
    faculty = get_object_or_404(Faculty, id=faculty_id)
    if request.method == "POST":
        form = FacultyForm(request.POST, instance=faculty)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تحديث الكلية بنجاح")
            return redirect("admin_faculty_list")
    else:
        form = FacultyForm(instance=faculty)
    return render(
        request,
        "administration/admin_form.html",
        {"form": form, "title": "تعديل كلية", "editing": True},
    )


@login_required
@user_passes_test(is_superuser)
def admin_faculty_delete(request, faculty_id):
    faculty = get_object_or_404(Faculty, id=faculty_id)

    # Check for related transactions
    has_tx, msg = _has_related_transactions("faculty", faculty_id)
    if has_tx:
        messages.error(request, msg)
        return redirect("admin_faculty_list")

    if request.method == "POST":
        faculty_name = faculty.name
        faculty.delete()
        messages.success(request, f'تم حذف الكلية "{faculty_name}" بنجاح')
        return redirect("admin_faculty_list")

    return render(
        request,
        "administration/admin_confirm_delete.html",
        {"object": faculty, "obj_name": f"الكلية {faculty.name}", "entity": "faculty"},
    )


@login_required
@user_passes_test(is_superuser)
def admin_department_list(request):
    departments = Department.objects.select_related("faculty").order_by(
        "name", "faculty__name"
    )
    return render(
        request,
        "administration/admin_department_list.html",
        {"departments": departments},
    )


@login_required
@user_passes_test(is_superuser)
def admin_department_create(request):
    if request.method == "POST":
        form = FacultyDepartmentForm(request.POST)

        # 🔑 CRITICAL: You MUST check is_valid() before saving
        if form.is_valid():
            form.save()
            messages.success(request, "تم إضافة القسم بنجاح")
            return redirect("admin_department_list")
        else:
            # Errors will automatically show in the template
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج")
    else:
        form = FacultyDepartmentForm()

    return render(
        request, "administration/admin_form.html", {"form": form, "title": "إضافة قسم"}
    )


@login_required
@user_passes_test(is_superuser)
def admin_department_edit(request, dept_id):
    dept = get_object_or_404(Department, id=dept_id)

    if request.method == "POST":
        form = FacultyDepartmentForm(request.POST, instance=dept)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تحديث القسم بنجاح")
            return redirect("admin_department_list")
        else:
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج")
    else:
        form = FacultyDepartmentForm(instance=dept)

    return render(
        request,
        "administration/admin_form.html",
        {"form": form, "title": "تعديل قسم", "editing": True},
    )


@login_required
@user_passes_test(is_superuser)
def admin_department_delete(request, dept_id):
    dept = get_object_or_404(Department, id=dept_id)

    # Check for related transactions
    has_tx, msg = _has_related_transactions("department", dept_id)
    if has_tx:
        messages.error(request, msg)
        return redirect("admin_department_list")

    if request.method == "POST":
        dept_name = dept.name
        faculty_name = dept.faculty.name
        dept.delete()
        messages.success(
            request, f'تم حذف القسم "{dept_name}" من كلية "{faculty_name}" بنجاح'
        )
        return redirect("admin_department_list")

    return render(
        request,
        "administration/admin_confirm_delete.html",
        {"object": dept, "obj_name": f"القسم {dept.name}", "entity": "department"},
    )


@login_required
@user_passes_test(lambda u: u.is_superuser)
def view_logs(request):
    """Display complete system logs, including full multi-line tracebacks."""
    log_path = settings.BASE_DIR / "logs" / "inventory_control.log"
    log_entries = []

    try:
        if log_path.exists():
            with open(log_path, "r", encoding="utf-8") as f:
                lines = f.readlines()

            # Regex to match log line start: YYYY-MM-DD HH:MM:SS | LEVEL |
            log_pattern = re.compile(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2} \|")
            current_entry = None

            # Process last 300 lines to capture full recent tracebacks
            for line in lines[-300:]:
                line = line.rstrip("\n")
                if log_pattern.match(line):
                    parts = line.split(" | ", 3)
                    current_entry = {
                        "time": parts[0].strip(),
                        "level": parts[1].strip() if len(parts) > 1 else "LOG",
                        "module": parts[2].strip() if len(parts) > 2 else "-",
                        "message": parts[3] if len(parts) > 3 else "",
                    }
                    log_entries.append(current_entry)
                elif current_entry is not None:
                    # Continuation line (part of traceback) → append to message
                    current_entry["message"] += "\n" + line

            # Show newest first
            log_entries = list(reversed(log_entries))
        else:
            messages.warning(request, "ملف السجلات غير موجود بعد.")
    except Exception as e:
        logger.error(f"Failed to read logs: {e}", exc_info=True)
        messages.error(request, f"خطأ في قراءة السجلات: {str(e)}")

    return render(request, "administration/logs.html", {"log_entries": log_entries})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_backup_db(request):
    """Stream PostgreSQL backup directly to client browser (no server file)."""
    if request.method != "POST":
        return redirect("home")

    db_config = settings.DATABASES.get("default", {})
    if "postgresql" not in db_config.get("ENGINE", "").lower():
        messages.error(request, "النسخ الاحتياطي متاح فقط لقواعد بيانات PostgreSQL.")
        return redirect("home")

    db_name = db_config["NAME"]
    db_user = db_config.get("USER", os.environ.get("PGUSER", ""))
    db_password = db_config.get("PASSWORD", os.environ.get("PGPASSWORD", ""))
    db_host = db_config.get("HOST", "localhost")
    db_port = db_config.get("PORT", "5432")

    if not db_user or not db_name:
        messages.error(request, "بيانات الاتصال بقاعدة البيانات غير مكتملة.")
        return redirect("home")

    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f"backup_{db_name}_{timestamp}.backup"

    PG_DUMP_PATH = r"C:\PostgreSQL\18\bin\pg_dump.exe"

    # Remove -f flag: pg_dump will output to stdout
    cmd = [
        # PG_DUMP_PATH,
        "pg_dump",
        "-h",
        str(db_host),
        "-p",
        str(db_port),
        "-U",
        str(db_user),
        "-d",
        str(db_name),
        "-F",
        "t",  # Tar format → stdout
    ]

    env = os.environ.copy()
    if db_password:
        env["PGPASSWORD"] = str(db_password)

    # Prevent console popup on Windows
    creation_flags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0

    try:
        # Start process without blocking
        process = subprocess.Popen(
            cmd,
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            creationflags=creation_flags,
        )

        # Stream stdout directly to browser
        response = StreamingHttpResponse(
            process.stdout, content_type="application/x-tar"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        # Attach cleanup to ensure process terminates after download
        response.streaming_content = _stream_backup(process)
        return response

    except FileNotFoundError:
        messages.error(request, f"ملف pg_dump.exe غير موجود في: {PG_DUMP_PATH}")
    except Exception as e:
        logger.error(f"❌ DB backup failed: {e}", exc_info=True)
        messages.error(request, "حدث خطأ أثناء إنشاء النسخة الاحتياطية.")

    return redirect("home")


def _stream_backup(process):
    """Generator that streams stdout and logs stderr if backup fails."""
    try:
        # Yield chunks as they are generated
        for chunk in iter(lambda: process.stdout.read(8192), b""):
            yield chunk
    finally:
        # Wait for process to finish & log errors
        process.stdout.close()
        _, stderr = process.communicate()
        if process.returncode != 0:
            err_msg = stderr.decode("utf-8", errors="replace")[:500]
            logger.error(
                f"❌ pg_dump failed (exit code {process.returncode}): {err_msg}"
            )
