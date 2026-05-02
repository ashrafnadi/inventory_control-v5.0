# administration/management/commands/create_users_from_excel.py
"""
Management command to create users and profiles from Excel file.
✅ FIXED: Properly handles UserProfile.save() logic for faculty/department assignment.
"""

import logging
from pathlib import Path

from django.contrib.auth.models import User
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from administration.models import Department, Faculty, UserProfile

logger = logging.getLogger(__name__)
DEFAULT_PASSWORD = "pun121fate121"


class Command(BaseCommand):
    help = "Create users and profiles from Excel file with role assignments"

    def add_arguments(self, parser):
        parser.add_argument("excel_file", type=str, help="Path to Excel file (.xlsx)")
        parser.add_argument(
            "--dry-run", action="store_true", help="Preview without changes"
        )
        parser.add_argument(
            "--faculty", type=int, help="Override faculty_id for all users"
        )
        parser.add_argument(
            "--department", type=int, help="Override department_id for all users"
        )
        parser.add_argument(
            "--skip-existing", action="store_true", help="Skip existing users"
        )
        parser.add_argument(
            "--update-existing", action="store_true", help="Update existing users"
        )

    def _parse_bool(self, value, default=False):
        """Safely parse boolean from Excel cell (TRUE/FALSE/1/0/yes/no)."""
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        return str(value).strip().lower() in ("true", "1", "yes", "y")

    def handle(self, *args, **options):
        excel_path = Path(options["excel_file"])
        dry_run = options["dry_run"]
        override_faculty = options.get("faculty")
        override_department = options.get("department")
        skip_existing = options["skip_existing"]
        update_existing = options["update_existing"]

        if not excel_path.exists():
            self.stderr.write(self.style.ERROR(f"File not found: {excel_path}"))
            return
        if not excel_path.suffix.lower() == ".xlsx":
            self.stderr.write(self.style.ERROR("Only .xlsx files are supported"))
            return

        self.stdout.write(f"Loading: {excel_path}")
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN MODE"))

        try:
            wb = load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Failed to read Excel: {e}"))
            return

        headers = [str(cell.value).strip().lower() for cell in ws[1] if cell.value]
        required = ["username", "first_name", "faculty_id", "department_id"]
        missing = [f for f in required if f not in headers]
        if missing:
            self.stderr.write(
                self.style.ERROR(f"Missing columns: {', '.join(missing)}")
            )
            return

        created = updated = skipped = errors = 0
        error_list = []

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(cell for cell in row):
                continue

            row_data = {
                headers[i]: cell
                for i, cell in enumerate(row)
                if i < len(headers) and headers[i]
            }

            try:
                username = str(row_data.get("username", "")).strip()
                first_name = str(row_data.get("first_name", "")).strip()
                password = (
                    str(row_data.get("password", DEFAULT_PASSWORD)).strip()
                    or DEFAULT_PASSWORD
                )

                faculty_id = (
                    override_faculty if override_faculty else row_data.get("faculty_id")
                )
                department_id = (
                    override_department
                    if override_department
                    else row_data.get("department_id")
                )

                if not faculty_id or not department_id:
                    raise ValueError("Missing faculty_id or department_id")

                faculty_id = int(faculty_id)
                department_id = int(department_id)

                # Parse User model booleans
                is_active = self._parse_bool(row_data.get("is_active"), True)
                is_staff = self._parse_bool(row_data.get("is_staff"), False)
                is_superuser = self._parse_bool(row_data.get("is_superuser"), False)

                # Parse UserProfile role flags
                is_inventory_manager = self._parse_bool(
                    row_data.get("is_inventory_manager"), False
                )
                is_inventory_employee = self._parse_bool(
                    row_data.get("is_inventory_employee"), False
                )
                is_administration_manager = self._parse_bool(
                    row_data.get("is_administration_manager"), False
                )
                is_faculty_manager = self._parse_bool(
                    row_data.get("is_faculty_manager"), False
                )
                is_user = self._parse_bool(row_data.get("is_user"), True)

                # ✅ Validate FKs exist AND get model instances (not just IDs)
                faculty = Faculty.objects.filter(id=faculty_id).first()
                if not faculty:
                    raise ValueError(f"Faculty {faculty_id} not found")
                department = Department.objects.filter(id=department_id).first()
                if not department:
                    raise ValueError(f"Department {department_id} not found")
                # ✅ Ensure department belongs to faculty (matches model validation)
                if department.faculty_id != faculty_id:
                    raise ValueError(
                        f"Department {department_id} does not belong to Faculty {faculty_id}"
                    )

                with transaction.atomic():
                    # Handle User creation/update
                    user, user_created = User.objects.get_or_create(
                        username=username,
                        defaults={
                            "first_name": first_name,
                            "is_active": is_active,
                            "is_staff": is_staff,
                            "is_superuser": is_superuser,
                        },
                    )

                    if not user_created:
                        if skip_existing:
                            skipped += 1
                            continue
                        elif update_existing:
                            changed = False
                            if user.first_name != first_name:
                                user.first_name = first_name
                                changed = True
                            if user.is_active != is_active:
                                user.is_active = is_active
                                changed = True
                            if user.is_staff != is_staff:
                                user.is_staff = is_staff
                                changed = True
                            if user.is_superuser != is_superuser:
                                user.is_superuser = is_superuser
                                changed = True
                            if changed and not dry_run:
                                user.save()
                            updated += 1
                        else:
                            skipped += 1
                            continue

                    # Set password for new or updated users
                    if (user_created or update_existing) and not dry_run:
                        user.set_password(password)
                        user.save(
                            update_fields=["password"] if not user_created else None
                        )

                    # ✅ FIXED: Handle UserProfile with proper save() logic
                    profile, profile_created = UserProfile.objects.get_or_create(
                        user=user,
                        defaults={
                            "faculty": faculty,
                            "department": department,
                        },  # Pass instances, not IDs
                    )

                    # ✅ Update profile fields AFTER creation to trigger save() logic
                    if not dry_run:
                        # Update role flags
                        role_updates = {
                            "is_inventory_manager": is_inventory_manager,
                            "is_inventory_employee": is_inventory_employee,
                            "is_administration_manager": is_administration_manager,
                            "is_faculty_manager": is_faculty_manager,
                            "is_user": is_user,
                        }
                        for field, value in role_updates.items():
                            if getattr(profile, field) != value:
                                setattr(profile, field, value)

                        # ✅ Explicitly set faculty/department to trigger model.save() logic
                        if profile.faculty_id != faculty_id:
                            profile.faculty = faculty
                        if profile.department_id != department_id:
                            profile.department = department

                        # ✅ Save to trigger clean() and auto-set faculty from department
                        profile.save()

                    if user_created:
                        created += 1
                        status = "Would create" if dry_run else "Created"
                        roles = []
                        if is_inventory_manager:
                            roles.append("INV_MGR")
                        if is_inventory_employee:
                            roles.append("INV_EMP")
                        if is_faculty_manager:
                            roles.append("FAC_MGR")
                        if is_administration_manager:
                            roles.append("ADM_MGR")
                        role_str = f" [{','.join(roles)}]" if roles else ""
                        self.stdout.write(
                            self.style.SUCCESS(f"  {status}: {username}{role_str}")
                        )
                        logger.info(
                            f"Created user {username} | Faculty: {faculty.name} | Dept: {department.name} | Roles: {','.join(roles) or 'none'}"
                        )

            except Exception as e:
                errors += 1
                msg = f"Row {row_idx}: {e}"
                error_list.append(msg)
                self.stderr.write(self.style.ERROR(f"  Error: {msg}"))
                logger.error(
                    f"Row {row_idx} error: {e} | Username: {row_data.get('username', 'N/A')}"
                )

        # Summary
        self.stdout.write("\n" + "=" * 70)
        self.stdout.write(self.style.SUCCESS("SUMMARY"))
        self.stdout.write("=" * 70)
        self.stdout.write(f"Created: {created}")
        self.stdout.write(f"Updated: {updated}")
        self.stdout.write(f"Skipped: {skipped}")
        self.stdout.write(f"Errors: {errors}")

        if error_list:
            self.stdout.write("\nErrors (first 10):")
            for e in error_list[:10]:
                self.stdout.write(f"  {e}")

        if dry_run:
            self.stdout.write(self.style.WARNING("\nDRY RUN - No changes made"))
        elif errors == 0:
            self.stdout.write(
                self.style.SUCCESS(f"\nProcessed {created + updated} users")
            )
