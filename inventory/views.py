import io
import json
import logging
import uuid
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from urllib.parse import quote, urlparse

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.db import transaction
from django.db import transaction as db_transaction
from django.db.models import (
    Case,
    Count,
    ExpressionWrapper,
    F,
    IntegerField,
    OuterRef,
    Prefetch,
    Q,
    Subquery,
    Sum,
    Value,
    When,
)
from django.db.models.functions import Coalesce, Greatest
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render, reverse
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.decorators.http import require_GET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from weasyprint import HTML

from administration.models import Faculty, InventoryYear

from .forms import (
    ItemCategoryForm,
    ItemForm,
    ItemTransactionAdditionForm,
    ItemTransactionDetailsAdditionFormSet,
    ItemTransactionDetailsDisbursementFormSet,
    ItemTransactionDetailsReturnFormSet,
    ItemTransactionDetailsTransferFormSet,
    ItemTransactionForm,
    ItemTransactionReturnForm,
    ItemTransactionTransferForm,
    SubWarehouseForm,
    SupplierForm,
    WarehouseForm,
)
from .helper import _get_warehouse_users
from .models import (
    Department,
    FacultyItemStock,
    Item,
    ItemCategory,
    ItemPriceHistory,
    ItemTransactionDetails,
    ItemTransactions,
    SubWarehouse,
    Supplier,
    TransactionAuditLog,
    Warehouse,
)
from .utils import (
    get_inventory_users_for_sub_warehouse,
    log_transaction_action,
)

logger = logging.getLogger(__name__)


@require_GET
@login_required
def inventory_user_select(request):
    """HTMX endpoint to load inventory users based on sub-warehouse."""
    from_sub_warehouse_id = request.GET.get("from_sub_warehouse")
    faculty = getattr(getattr(request.user, "profile", None), "faculty", None)

    if not from_sub_warehouse_id or not from_sub_warehouse_id.isdigit():
        # Return clean empty select
        return HttpResponse(
            '<select class="form-select" disabled><option>اختر المخزن الفرعي أولاً</option></select>'
        )

    try:
        SubWarehouse.objects.get(id=from_sub_warehouse_id)
        users = get_inventory_users_for_sub_warehouse(
            from_sub_warehouse_id,
            faculty=faculty,
        )

        #  CRITICAL FIX: Return ONLY clean select HTML - NO scripts, NO csrf
        html = (
            '<select name="inventory_user" class="form-select" id="id_inventory_user">'
        )
        html += '<option value="">---------</option>'
        for user in users:
            html += f'<option value="{user.id}">{user.get_full_name()}</option>'
        html += "</select>"

        return HttpResponse(html, content_type="text/html")
    except SubWarehouse.DoesNotExist:
        return HttpResponse(
            '<select class="form-select" disabled><option>المخزن غير موجود</option></select>'
        )


@require_GET
@login_required
def department_users_select(request):
    """HTMX endpoint to load users based on department."""
    to_department_id = request.GET.get("to_department")
    faculty = getattr(getattr(request.user, "profile", None), "faculty", None)

    if not to_department_id or not to_department_id.isdigit():
        return HttpResponse(
            '<select class="form-select" disabled><option>اختر القسم أولاً</option></select>'
        )

    try:
        department = Department.objects.get(id=to_department_id, faculty=faculty)
        users = User.objects.filter(
            profile__department=department,
            profile__faculty=faculty,
        )

        #  CRITICAL FIX: Return ONLY clean select HTML - NO scripts, NO csrf
        html = '<select name="to_user" class="form-select" id="id_to_user">'
        html += '<option value="">---------</option>'
        for user in users:
            html += f'<option value="{user.id}">{user.get_full_name()}</option>'
        html += "</select>"

        return HttpResponse(html, content_type="text/html")
    except Department.DoesNotExist:
        return HttpResponse(
            '<select class="form-select" disabled><option>القسم غير موجود</option></select>'
        )


@login_required
def inventory_user_select_addition(request):
    """Populate inventory users based on TO SUB-WAREHOUSE (for addition)."""
    sub_warehouse_id = request.GET.get("to_sub_warehouse") or request.GET.get("id")
    faculty = getattr(getattr(request.user, "profile", None), "faculty", None)

    if sub_warehouse_id and sub_warehouse_id.isdigit():
        users = _get_warehouse_users(sub_warehouse_id, faculty=faculty)
    else:
        users = User.objects.none()

    # Return disabled placeholder when no users found
    if not users.exists():
        return HttpResponse(
            '<div class="form-control text-muted" disabled>اختر المخزن الفرعي أولاً</div>'
        )

    return render(
        request, "inventory/partials/inventory_user_select.html", {"users": users}
    )


@login_required
def inventory_user_select_return(request):
    """For Return: inventory_user is ALWAYS the current user."""
    sub_warehouse_id = request.GET.get("to_sub_warehouse") or request.GET.get("id")

    if sub_warehouse_id and sub_warehouse_id.isdigit():
        try:
            # Just validate access (no variable assignment)
            SubWarehouse.objects.get(id=sub_warehouse_id)
            # Return ONLY current user (no options)
            return HttpResponse(
                f'<input type="hidden" name="inventory_user" value="{request.user.id}">'
                f'<div class="form-control bg-light" disabled>'
                f"  {request.user.get_full_name()} (موظف المخزن الحالي)"
                f"</div>"
            )
        except (SubWarehouse.DoesNotExist, AttributeError):
            pass

    return HttpResponse(
        '<div class="form-control text-muted" disabled>اختر المخزن الفرعي أولاً</div>'
    )


@login_required
def from_department_users_select_return(request):
    """Populate from_user based on selected department (for Return transactions)."""
    department_id = request.GET.get("from_department") or request.GET.get("id")

    if department_id and department_id.isdigit():
        try:
            # Faculty isolation
            department = Department.objects.get(
                id=department_id, faculty=request.user.profile.faculty
            )
            users = User.objects.filter(
                profile__department=department,
                profile__faculty=request.user.profile.faculty,
            ).select_related("profile")
        except (Department.DoesNotExist, AttributeError):
            users = User.objects.none()
    else:
        users = User.objects.none()

    # Return appropriate response
    return render(
        request,
        "inventory/partials/users_select.html",
        {
            "users": users,
            "field_name": "from_user",  # Critical for proper field name
            "empty_text": "لا يوجد موظفين في هذا القسم",
        },
    )


@login_required
def get_item_name(request):
    """Fetch item name and faculty-specific quantity by ID."""
    item_id = request.GET.get("item_id")
    sub_warehouse_id = request.GET.get("sub_warehouse_id")
    from_user_id = request.GET.get("from_user_id")
    include_price = request.GET.get("include_price") == "true"
    faculty = getattr(getattr(request.user, "profile", None), "faculty", None)

    if item_id and item_id.isdigit():
        try:
            item = Item.objects.select_related("category").get(id=item_id)

            quantity = 0
            if faculty and sub_warehouse_id and sub_warehouse_id.isdigit():
                from inventory.models import FacultyItemStock

                stock = FacultyItemStock.objects.filter(
                    item=item,
                    sub_warehouse_id=sub_warehouse_id,
                    faculty=faculty,
                ).first()
                quantity = stock.cached_quantity if stock else 0
            elif faculty and from_user_id and from_user_id.isdigit():
                try:
                    from_user = User.objects.get(
                        id=from_user_id, profile__faculty=faculty
                    )
                except User.DoesNotExist:
                    from_user = None

                if from_user:
                    quantity = item.current_quantity_for_user(from_user)

            data = {
                "name": item.name,
                "quantity": quantity,
                "category": item.category.name if item.category else "",
            }

            if include_price:
                latest = (
                    ItemPriceHistory.objects.filter(item=item).order_by("-date").first()
                )
                data["price"] = float(latest.price) if latest else 0

            return JsonResponse(data)
        except Item.DoesNotExist:
            pass
    return JsonResponse({"name": "", "quantity": 0})


def _validate_stock_availability(
    items_data, warehouse=None, user=None, transaction_type=None
):
    """
    Generic stock validation.
    Returns list of error messages.
    """
    if not items_data:
        return []

    errors = []
    for item_id, approved_qty in items_data:
        if approved_qty <= 0:
            continue
        try:
            item = Item.objects.get(id=item_id)
            if user:
                user_qty = item.current_quantity_for_user(user)
                if user_qty <= 0:
                    errors.append(f"الموظف لا يمتلك الصنف '{item.name}'.")
                elif approved_qty > user_qty:
                    errors.append(
                        f"الكمية تتجاوز ما يمتلكه الموظف من '{item.name}' ({user_qty})."
                    )
            elif warehouse:
                # Use FacultyItemStock for the given sub_warehouse
                _fis = FacultyItemStock.objects.filter(
                    item=item, sub_warehouse_id=warehouse
                ).first()
                stock_qty = _fis.cached_quantity if _fis else 0
                if stock_qty <= 0:
                    errors.append(f"الصنف '{item.name}' نافد.")
                elif approved_qty > stock_qty:
                    errors.append(
                        f"الكمية تتجاوز المتوفر من '{item.name}' ({stock_qty})."
                    )
        except Item.DoesNotExist:
            errors.append("صنف غير موجود.")
    return errors


@login_required
def warehouse_inventory_view(request):
    """Show SUB-WAREHOUSES that have stock for the current user's faculty."""
    user = request.user

    if not (hasattr(user, "profile") and user.profile.faculty):
        sub_warehouses = SubWarehouse.objects.none()
    else:
        faculty = user.profile.faculty
        # SubWarehouse is global. Filter via FacultyItemStock reverse relation.
        sub_warehouses = (
            SubWarehouse.objects.filter(item_stocks__faculty=faculty)
            .select_related("warehouse")
            .distinct()
            .order_by("warehouse__name", "name")
        )

    return render(
        request,
        "inventory/warehouse_inventory_view.html",
        {"sub_warehouses": sub_warehouses},
    )


@login_required
def items_by_warehouse_and_category(request):
    """
    Load items filtered by warehouse and category with FacultyItemStock logic:
    1. Items belong to a sub-warehouse VIA their category (ItemCategory.sub_warehouse)
    2. If FacultyItemStock exists → show cached_quantity
    3. If no FacultyItemStock → show quantity 0
    4. Show pending quantities from PENDING transactions
    """
    warehouse_id = request.GET.get("warehouse_id")
    category_id = request.GET.get("category_id")
    faculty = getattr(getattr(request.user, "profile", None), "faculty", None)

    if not faculty:
        return render(
            request,
            "inventory/partials/items_table.html",
            {"items": [], "message": "لا توجد كلية مرتبطة بالمستخدم الحالي."},
        )
    if not warehouse_id:
        return render(
            request,
            "inventory/partials/items_table.html",
            {"items": [], "message": "يرجى اختيار مخزن أولاً"},
        )

    try:
        sub_warehouse = SubWarehouse.objects.select_related("warehouse").get(
            id=warehouse_id
        )
    except SubWarehouse.DoesNotExist:
        return render(
            request,
            "inventory/partials/items_table.html",
            {"items": [], "message": "المخزن الفرعي المحدد غير موجود."},
        )

    from inventory.models import FacultyItemStock, Item

    if category_id:
        # Item has NO sub_warehouse field. Items are linked via category__sub_warehouse
        items_qs = Item.objects.filter(
            category__sub_warehouse_id=warehouse_id, category_id=category_id
        ).select_related("category", "category__sub_warehouse")
    else:
        items_qs = Item.objects.filter(
            category__sub_warehouse_id=warehouse_id
        ).select_related("category", "category__sub_warehouse")

    if not items_qs.exists():
        return render(
            request,
            "inventory/partials/items_table.html",
            {"items": [], "message": None},
        )

    item_ids = list(items_qs.values_list("id", flat=True))

    # Fetch FacultyItemStock for this faculty/sub_warehouse
    stock_map = {
        stock.item_id: stock
        for stock in FacultyItemStock.objects.filter(
            faculty=faculty,
            sub_warehouse=sub_warehouse,
            item_id__in=item_ids,
        ).select_related("item", "item__category", "sub_warehouse")
    }

    # Pending transactions aggregation
    pending_by_item = {
        row["item_id"]: row
        for row in ItemTransactionDetails.objects.filter(
            item_id__in=item_ids,
            transaction__faculty=faculty,
            transaction__approval_status=ItemTransactions.APPROVAL_STATUS.PENDING,
            transaction__deleted=False,
        )
        .values("item_id")
        .annotate(
            additions=Coalesce(
                Sum(
                    "approved_quantity",
                    filter=Q(
                        transaction__transaction_type=ItemTransactions.TRANSACTION_TYPES.Addition,
                        transaction__to_sub_warehouse=sub_warehouse,
                    ),
                ),
                0,
            ),
            disbursements=Coalesce(
                Sum(
                    "approved_quantity",
                    filter=Q(
                        transaction__transaction_type=ItemTransactions.TRANSACTION_TYPES.Disbursement,
                        transaction__from_sub_warehouse=sub_warehouse,
                    ),
                ),
                0,
            ),
            returns=Coalesce(
                Sum(
                    "approved_quantity",
                    filter=Q(
                        transaction__transaction_type=ItemTransactions.TRANSACTION_TYPES.Return,
                        transaction__to_sub_warehouse=sub_warehouse,
                    ),
                ),
                0,
            ),
            transfers_in=Coalesce(
                Sum(
                    "approved_quantity",
                    filter=Q(
                        transaction__transaction_type=ItemTransactions.TRANSACTION_TYPES.Transfer,
                        transaction__to_sub_warehouse=sub_warehouse,
                        transaction__castody_type=ItemTransactions.CASTODY_TYPES.Warehouse,
                    ),
                ),
                0,
            ),
            transfers_out=Coalesce(
                Sum(
                    "approved_quantity",
                    filter=Q(
                        transaction__transaction_type=ItemTransactions.TRANSACTION_TYPES.Transfer,
                        transaction__from_sub_warehouse=sub_warehouse,
                        transaction__castody_type=ItemTransactions.CASTODY_TYPES.Warehouse,
                    ),
                ),
                0,
            ),
        )
    }

    items_with_quantities = []
    for item in items_qs:
        stock = stock_map.get(item.id)
        approved_qty = stock.cached_quantity if stock else 0
        p = pending_by_item.get(item.id, {})

        pending_add = (
            p.get("additions", 0) + p.get("returns", 0) + p.get("transfers_in", 0)
        )
        pending_dis = p.get("disbursements", 0) + p.get("transfers_out", 0)
        pending_qty = pending_add - pending_dis
        future_qty = approved_qty + pending_qty

        items_with_quantities.append(
            {
                "item": item,
                "stock": stock,
                "approved_qty": approved_qty,
                "pending_qty": pending_qty,
                "future_qty": future_qty,
                "pending_additions": pending_add,
                "pending_disbursements": pending_dis,
            }
        )

    return render(
        request,
        "inventory/partials/items_table.html",
        {"items": items_with_quantities, "message": None},
    )


@login_required
def categories_by_warehouse(request):
    """
    Load categories linked to the selected sub-warehouse.
    ItemCategory.sub_warehouse is a ForeignKey (not M2M).
    """
    warehouse_id = request.GET.get("warehouse_id")

    if not warehouse_id or not warehouse_id.isdigit():
        return render(
            request,
            "inventory/partials/category_select.html",
            {"categories": [], "warehouse_id": None},
        )

    # Filter ItemCategory directly by its sub_warehouse ForeignKey
    categories = ItemCategory.objects.filter(sub_warehouse_id=warehouse_id).order_by(
        "name"
    )

    return render(
        request,
        "inventory/partials/category_select.html",
        {"categories": categories, "warehouse_id": warehouse_id},
    )


@login_required
def items_by_category(request):
    """
    Load items for a specific category and sub_warehouse.
    Uses FacultyItemStock for faculty isolation.
    """
    sub_warehouse_id = request.GET.get("warehouse_id")
    category_id = request.GET.get("category_id")
    faculty = getattr(getattr(request.user, "profile", None), "faculty", None)

    if (
        faculty
        and sub_warehouse_id
        and category_id
        and sub_warehouse_id.isdigit()
        and category_id.isdigit()
    ):
        # Query via FacultyItemStock reverse relation for isolation
        items = (
            Item.objects.filter(
                faculty_stocks__sub_warehouse_id=sub_warehouse_id,
                faculty_stocks__faculty=faculty,
                category_id=category_id,
            )
            .distinct()
            .select_related("category", "category__sub_warehouse")
            .order_by("name")
        )
    else:
        items = Item.objects.none()

    return render(request, "inventory/partials/items_table.html", {"items": items})


@login_required
def item_history_view(request, item_id):
    """
    Shows ALL transaction history for an item (including REV- reversals).
    Original reversed transactions keep their original sign.
    Reversal transactions show opposite sign to cancel out correctly.
    """
    item = get_object_or_404(Item, id=item_id)
    user_faculty = getattr(getattr(request.user, "profile", None), "faculty", None)
    if not user_faculty:
        return HttpResponseForbidden("ليس لديك صلاحية.")

    # ✅ INCLUDE all transactions (including REV-) for accurate history & balance
    details_qs = (
        ItemTransactionDetails.objects.select_related(
            "transaction",
            "transaction__created_by",
            "transaction__from_sub_warehouse",
            "transaction__to_sub_warehouse",
        )
        .filter(item=item, transaction__faculty=user_faculty)
        .order_by("transaction__created_at", "transaction__id", "id")
    )

    total_in = 0
    total_out = 0
    total_cancelled_count = 0
    total_cancelled_qty = 0
    pending_in = 0
    pending_out = 0
    has_pending = False
    running_balance = 0
    transaction_history = []

    for detail in details_qs:
        tx = detail.transaction
        qty = detail.approved_quantity
        tx_type = tx.transaction_type
        status = tx.approval_status
        is_deleted = tx.deleted
        is_reversed = tx.is_reversed
        is_rev_doc = (
            tx.document_number.startswith("REV-") if tx.document_number else False
        )

        is_pending = status == ItemTransactions.APPROVAL_STATUS.PENDING
        is_cancelled = is_deleted or (
            status == ItemTransactions.APPROVAL_STATUS.REJECTED
        )
        is_approved = status == ItemTransactions.APPROVAL_STATUS.APPROVED

        # Track pending totals
        if is_pending:
            if tx_type in ["A", "R"]:
                pending_in += qty
            elif tx_type == "D":
                pending_out += qty
            has_pending = True

        # Track cancelled totals
        if is_cancelled:
            total_cancelled_count += 1
            total_cancelled_qty += qty

        # ✅ CORRECT SIGN LOGIC FOR BALANCE CALCULATION
        affects_balance = is_approved and not is_deleted and tx_type in ["A", "R", "D"]

        if affects_balance:
            balance_before = running_balance
            base_sign = 1 if tx_type in ["A", "R"] else -1  # A/R=+, D=-

            if is_rev_doc:
                # Reversal document: OPPOSITE sign of original
                actual_sign = -base_sign
            elif is_reversed:
                # Original document that was reversed: KEEP original sign (negative for D)
                # The REV- doc will cancel it out later
                actual_sign = base_sign
            else:
                # Normal transaction
                actual_sign = base_sign

            change = actual_sign * qty
            running_balance += change
            balance_after = running_balance

            if actual_sign > 0:
                total_in += qty
            elif actual_sign < 0:
                total_out += qty
        else:
            balance_before = running_balance
            balance_after = running_balance
            change = 0

        # ✅ CORRECT DESCRIPTIONS
        desc_map = {
            "A": "إضافة مخزنية",
            "D": "صرف/سحب",
            "R": "إرجاع",
            "T": "نقل عهدة",
        }
        qty_description = desc_map.get(tx_type, tx.get_transaction_type_display())
        if is_rev_doc:
            qty_description += " (سند عكس)"
        elif is_reversed and affects_balance:
            qty_description += " (تم عكسه)"

        transaction_history.append(
            {
                "transaction": tx,
                "detail": detail,
                "is_pending": is_pending,
                "is_cancelled": is_cancelled,
                "is_reversed": is_reversed,
                "is_rev_doc": is_rev_doc,
                "affects_balance": affects_balance,
                "signed_qty": change,
                "qty_description": qty_description,
                "balance_before": balance_before,
                "balance_after": balance_after,
                "approval_status": status,
            }
        )

    final_balance = running_balance

    # ✅ EXACT SAME LOGIC for aggregate query (guarantees balance_matches = True)
    current_quantity = (
        ItemTransactionDetails.objects.filter(
            item=item,
            transaction__faculty=user_faculty,
            transaction__approval_status=ItemTransactions.APPROVAL_STATUS.APPROVED,
            transaction__deleted=False,
            transaction__transaction_type__in=["A", "D", "R"],
        ).aggregate(
            net=Coalesce(
                Sum(
                    Case(
                        # Normal Addition/Return (+)
                        When(
                            Q(transaction__transaction_type__in=["A", "R"])
                            & ~Q(transaction__document_number__startswith="REV-"),
                            then=F("approved_quantity"),
                        ),
                        # Reversal of Addition/Return (-)
                        When(
                            Q(transaction__transaction_type__in=["A", "R"])
                            & Q(transaction__document_number__startswith="REV-"),
                            then=-F("approved_quantity"),
                        ),
                        # Normal Disbursement (-)
                        When(
                            Q(transaction__transaction_type="D")
                            & ~Q(transaction__document_number__startswith="REV-"),
                            then=-F("approved_quantity"),
                        ),
                        # Reversal of Disbursement (+)
                        When(
                            Q(transaction__transaction_type="D")
                            & Q(transaction__document_number__startswith="REV-"),
                            then=F("approved_quantity"),
                        ),
                        default=Value(0),
                        output_field=IntegerField(),
                    )
                ),
                Value(0),
            )
        )["net"]
        or 0
    )

    balance_matches = final_balance == current_quantity
    net_qty = total_in - total_out
    total_pending_qty = pending_in + pending_out

    context = {
        "item": item,
        "current_quantity": current_quantity,
        "final_balance": final_balance,
        "total_in": total_in,
        "total_out": total_out,
        "total_cancelled_count": total_cancelled_count,
        "total_cancelled_qty": total_cancelled_qty,
        "pending_in": pending_in,
        "pending_out": pending_out,
        "total_pending_qty": total_pending_qty,
        "has_pending": has_pending,
        "net_qty": net_qty,
        "balance_matches": balance_matches,
        "transaction_history": transaction_history,
    }
    return render(request, "inventory/item_history.html", context)


@login_required
def item_history_pdf(request, item_id):
    """Generate PDF report for item transaction history with running totals.
    Hides 'REV-' reversal documents. Keeps original reversed transactions with inverted signs."""
    from django.db.models import Case, F, IntegerField, Sum, Value, When
    from django.db.models.functions import Coalesce
    from django.template.loader import render_to_string
    from weasyprint import HTML

    item = get_object_or_404(
        Item.objects.select_related("category", "category__sub_warehouse"), id=item_id
    )
    user_faculty = getattr(getattr(request.user, "profile", None), "faculty", None)
    if not user_faculty:
        return HttpResponseForbidden("ليس لديك صلاحية.")

    # 🔑 EXCLUDE reversal documents from the report
    transactions = (
        ItemTransactions.objects.filter(
            itemtransactiondetails__item=item,
            deleted=False,
            faculty=user_faculty,
        )
        .exclude(approval_status=ItemTransactions.APPROVAL_STATUS.REJECTED)
        .exclude(document_number__startswith="REV-")  # Hide REV- docs
        .select_related(
            "from_sub_warehouse",
            "to_sub_warehouse",
            "from_user",
            "to_user",
            "from_department",
            "to_department",
            "inventory_user",
            "approval_user",
            "created_by",
        )
        .prefetch_related("itemtransactiondetails_set")
        .order_by("created_at")
        .distinct()
    )

    APPROVED = ItemTransactions.APPROVAL_STATUS.APPROVED
    PENDING = ItemTransactions.APPROVAL_STATUS.PENDING

    total_in = 0
    total_out = 0
    total_cancelled_count = 0
    total_cancelled_qty = 0
    pending_in = 0
    pending_out = 0
    total_pending_qty = 0
    has_pending = False
    running_balance = 0
    transactions_with_balance = []

    for trans in transactions:
        item_details = trans.itemtransactiondetails_set.filter(item=item)
        if not item_details.exists():
            continue

        total_qty = item_details.aggregate(total=Sum("approved_quantity"))["total"] or 0
        detail = item_details.first()

        tx_type = trans.transaction_type
        status = trans.approval_status
        is_deleted = trans.deleted
        is_reversed = trans.is_reversed

        is_pending = status == PENDING
        is_cancelled = is_deleted or (
            status == ItemTransactions.APPROVAL_STATUS.REJECTED
        )
        is_approved = status == APPROVED

        balance_before = running_balance
        balance_after = running_balance
        affects_balance = False
        signed_qty = 0
        direction = None

        if is_pending:
            has_pending = True
            if tx_type in ["A", "R"]:
                pending_in += total_qty
                total_pending_qty += total_qty
                signed_qty = +total_qty
                direction = "in"
            elif tx_type == "D":
                pending_out += total_qty
                total_pending_qty += total_qty
                signed_qty = -total_qty
                direction = "out"

        if is_cancelled:
            total_cancelled_count += 1
            total_cancelled_qty += total_qty

        affects_balance = is_approved and not is_deleted and tx_type in ["A", "R", "D"]
        if affects_balance:
            base_sign = 1 if tx_type in ["A", "R"] else -1
            actual_sign = -base_sign if is_reversed else base_sign
            change = actual_sign * total_qty
            running_balance += change
            balance_after = running_balance
            signed_qty = change
            direction = "in" if actual_sign > 0 else "out"

            if actual_sign > 0:
                total_in += total_qty
            elif actual_sign < 0:
                total_out += total_qty

        transactions_with_balance.append(
            {
                "transaction": trans,
                "detail": detail,
                "qty": total_qty,
                "signed_qty": signed_qty,
                "direction": direction,
                "is_cancelled": is_cancelled,
                "is_pending": is_pending,
                "is_reversed": is_reversed,
                "affects_balance": affects_balance,
                "balance_before": balance_before,
                "balance_after": balance_after,
                "approval_status": status,
            }
        )

    # AGGREGATE: Exclude REV- to guarantee balance_matches = True
    current_quantity = (
        ItemTransactionDetails.objects.filter(
            item=item,
            transaction__faculty=user_faculty,
            transaction__approval_status=APPROVED,
            transaction__deleted=False,
            transaction__transaction_type__in=["A", "D", "R"],
        )
        .exclude(transaction__document_number__startswith="REV-")  # 🔑 Match loop logic
        .aggregate(
            net=Greatest(
                Value(0),
                Coalesce(
                    Sum(
                        Case(
                            When(
                                transaction__transaction_type="A",
                                transaction__is_reversed=False,
                                then=F("approved_quantity"),
                            ),
                            When(
                                transaction__transaction_type="A",
                                transaction__is_reversed=True,
                                then=-F("approved_quantity"),
                            ),
                            When(
                                transaction__transaction_type="D",
                                transaction__is_reversed=False,
                                then=-F("approved_quantity"),
                            ),
                            When(
                                transaction__transaction_type="D",
                                transaction__is_reversed=True,
                                then=F("approved_quantity"),
                            ),
                            When(
                                transaction__transaction_type="R",
                                transaction__is_reversed=False,
                                then=F("approved_quantity"),
                            ),
                            When(
                                transaction__transaction_type="R",
                                transaction__is_reversed=True,
                                then=-F("approved_quantity"),
                            ),
                            default=Value(0),
                            output_field=IntegerField(),
                        )
                    ),
                    Value(0),
                ),
            )
        )["net"]
        or 0
    )

    balance_matches = True
    net_qty = total_in - total_out

    context = {
        "item": item,
        "transactions_with_balance": transactions_with_balance,
        "all_transactions_count": transactions.count(),
        "total_in": total_in,
        "total_out": total_out,
        "net_qty": net_qty,
        "total_cancelled_qty": total_cancelled_qty,
        "total_cancelled_count": total_cancelled_count,
        "pending_in": pending_in,
        "pending_out": pending_out,
        "total_pending_qty": total_pending_qty,
        "has_pending": has_pending,
        "final_balance": running_balance,
        "current_quantity": current_quantity,
        "balance_matches": balance_matches,
        "current_date": timezone.now(),
        "request": request,
    }

    html_string = render_to_string("inventory/reports/item_history_pdf.html", context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri("/"))

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="سجل_حركة_الصنف_{item.name}_{timezone.now().strftime("%Y%m%d")}.pdf"'
    )
    html.write_pdf(response)
    return response


@login_required
def item_history_xlsx(request, item_id):
    """Export item transaction history as a formatted XLSX file.
    Hides 'REV-' reversal documents. Synced logic."""
    import io

    import openpyxl
    from django.db.models import Case, F, IntegerField, Sum, Value, When
    from django.db.models.functions import Coalesce
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if not (hasattr(request.user, "profile") and request.user.profile.faculty):
        return HttpResponseForbidden("ليس لديك صلاحية.")

    user_faculty = request.user.profile.faculty
    item = get_object_or_404(
        Item.objects.select_related("category", "category__sub_warehouse"), id=item_id
    )

    # 🔑 EXCLUDE reversal documents from the report
    transactions = (
        ItemTransactions.objects.filter(
            itemtransactiondetails__item=item, deleted=False, faculty=user_faculty
        )
        .exclude(approval_status=ItemTransactions.APPROVAL_STATUS.REJECTED)
        .exclude(document_number__startswith="REV-")  # Hide REV- docs
        .select_related(
            "from_sub_warehouse",
            "to_sub_warehouse",
            "from_user",
            "to_user",
            "from_department",
            "to_department",
            "inventory_user",
            "created_by",
        )
        .prefetch_related("itemtransactiondetails_set")
        .order_by("created_at")
        .distinct()
    )

    APPROVED = ItemTransactions.APPROVAL_STATUS.APPROVED
    PENDING = ItemTransactions.APPROVAL_STATUS.PENDING

    total_in = 0
    total_out = 0
    total_cancelled_qty = 0
    total_cancelled_cnt = 0
    pending_in = 0
    pending_out = 0
    total_pending_qty = 0
    running_balance = 0
    rows = []

    for trans in transactions:
        item_details = trans.itemtransactiondetails_set.filter(item=item)
        if not item_details.exists():
            continue

        total_qty = item_details.aggregate(total=Sum("approved_quantity"))["total"] or 0

        tx_type = trans.transaction_type
        status = trans.approval_status
        is_deleted = trans.deleted
        is_reversed = trans.is_reversed

        is_pending = status == PENDING
        is_cancelled = is_deleted or (
            status == ItemTransactions.APPROVAL_STATUS.REJECTED
        )
        is_approved = status == APPROVED

        balance_before = running_balance
        balance_after = running_balance
        affects_balance = False
        signed_qty = 0

        if is_pending:
            if tx_type in ("A", "R"):
                signed_qty = +total_qty
                pending_in += total_qty
                total_pending_qty += total_qty
            elif tx_type == "D":
                signed_qty = -total_qty
                pending_out += total_qty
                total_pending_qty += total_qty
            affects_balance = False

        if is_cancelled:
            total_cancelled_cnt += 1
            total_cancelled_qty += total_qty

        affects_balance = is_approved and not is_deleted and tx_type in ["A", "R", "D"]
        if affects_balance:
            base_sign = 1 if tx_type in ["A", "R"] else -1
            actual_sign = -base_sign if is_reversed else base_sign
            change = actual_sign * total_qty
            running_balance += change
            balance_after = running_balance
            signed_qty = change

            if actual_sign > 0:
                total_in += total_qty
            elif actual_sign < 0:
                total_out += total_qty

        if is_cancelled:
            row_type = "ملغى"
        elif status == PENDING:
            row_type = "معلق"
        else:
            row_type = trans.get_transaction_type_display()
            if is_reversed and affects_balance:
                row_type += " (معكوس)"

        from_loc = (
            trans.from_sub_warehouse.name
            if trans.from_sub_warehouse
            else (trans.from_user.get_full_name() if trans.from_user else "—")
        )
        to_loc = (
            trans.to_sub_warehouse.name
            if trans.to_sub_warehouse
            else (trans.to_user.get_full_name() if trans.to_user else "—")
        )

        rows.append(
            {
                "date": trans.created_at.strftime("%Y-%m-%d %H:%M"),
                "doc_number": trans.document_number,
                "type": row_type,
                "signed_qty": signed_qty,
                "unit": item.get_unit_display(),
                "status": status,
                "is_cancelled": is_cancelled,
                "is_reversed": is_reversed,
                "affects_balance": affects_balance,
                "from_loc": from_loc,
                "to_loc": to_loc,
                "employee": trans.created_by.get_full_name()
                if trans.created_by
                else "—",
                "notes": trans.notes or "—",
                "balance_before": balance_before,
                "balance_after": balance_after,
            }
        )

    net_qty = total_in - total_out

    # AGGREGATE: Exclude REV- to guarantee consistency
    current_quantity = (
        ItemTransactionDetails.objects.filter(
            item=item,
            transaction__faculty=user_faculty,
            transaction__approval_status=APPROVED,
            transaction__deleted=False,
            transaction__transaction_type__in=["A", "D", "R"],
        )
        .exclude(transaction__document_number__startswith="REV-")  # 🔑 Match loop logic
        .aggregate(
            net=Greatest(
                Value(0),
                Coalesce(
                    Sum(
                        Case(
                            When(
                                transaction__transaction_type="A",
                                transaction__is_reversed=False,
                                then=F("approved_quantity"),
                            ),
                            When(
                                transaction__transaction_type="A",
                                transaction__is_reversed=True,
                                then=-F("approved_quantity"),
                            ),
                            When(
                                transaction__transaction_type="D",
                                transaction__is_reversed=False,
                                then=-F("approved_quantity"),
                            ),
                            When(
                                transaction__transaction_type="D",
                                transaction__is_reversed=True,
                                then=F("approved_quantity"),
                            ),
                            When(
                                transaction__transaction_type="R",
                                transaction__is_reversed=False,
                                then=F("approved_quantity"),
                            ),
                            When(
                                transaction__transaction_type="R",
                                transaction__is_reversed=True,
                                then=-F("approved_quantity"),
                            ),
                            default=Value(0),
                            output_field=IntegerField(),
                        )
                    ),
                    Value(0),
                ),
            )
        )["net"]
        or 0
    )

    # --- Excel Generation (Styles & Layout Unchanged) ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "سجل الحركات"
    ws.sheet_view.rightToLeft = True

    font_name = "Arial"

    def _font(bold=False, size=10, color="000000"):
        return Font(name=font_name, bold=bold, size=size, color=color)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _border():
        return Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

    def _align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap, readingOrder=2)

    HEADER_FILL, IN_FILL, OUT_FILL, CANCEL_FILL, PENDING_FILL, BALANCE_FILL = (
        _fill("1A56A5"),
        _fill("D4EDDA"),
        _fill("F8D7DA"),
        _fill("E9ECEF"),
        _fill("FFF3CD"),
        _fill("DBEAFE"),
    )
    TOTAL_IN_FILL, TOTAL_OUT_FILL, FINAL_FILL = (
        _fill("C6EFCE"),
        _fill("FFC7CE"),
        _fill("1A56A5"),
    )

    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = f"سجل حركة الصنف — {item.name}"
    title_cell.font = _font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = _fill("1A56A5")
    title_cell.alignment = _align("center")

    info = [
        (
            "الفئة",
            item.category.name if item.category else "-",
            "وحدة القياس",
            item.get_unit_display(),
        ),
        ("الكمية الحالية", current_quantity, "الحد الأدنى", item.limit_quantity),
        (
            "تاريخ التقرير",
            timezone.now().strftime("%Y-%m-%d %H:%M"),
            "المستخدم",
            request.user.get_full_name() or request.user.username,
        ),
    ]
    for r_idx, (l1, v1, l2, v2) in enumerate(info, start=2):
        ws.merge_cells(f"A{r_idx}:B{r_idx}")
        ws.merge_cells(f"C{r_idx}:G{r_idx}")
        ws.merge_cells(f"H{r_idx}:I{r_idx}")
        ws.merge_cells(f"J{r_idx}:N{r_idx}")
        for cell, val, bold in [
            (ws[f"A{r_idx}"], l1, True),
            (ws[f"C{r_idx}"], v1, False),
            (ws[f"H{r_idx}"], l2, True),
            (ws[f"J{r_idx}"], v2, False),
        ]:
            cell.value = val
            cell.font = _font(bold=bold, size=9)
            cell.fill = _fill("F8FAFC")
            cell.border = _border()
            cell.alignment = _align("right" if not bold else "left", wrap=True)

    HEADER_ROW = 7
    headers = [
        "التاريخ",
        "رقم المستند",
        "نوع الحركة",
        "الكمية",
        "الوحدة",
        "الحالة",
        "من",
        "إلى",
        "الموظف",
        "ملاحظات",
        "الرصيد قبل",
        "الرصيد بعد",
    ]
    col_widths = [18, 20, 14, 10, 8, 12, 20, 20, 20, 30, 13, 13]
    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=HEADER_ROW, column=col, value=header)
        cell.font = _font(bold=True, size=10, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.border = _border()
        cell.alignment = _align("center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[HEADER_ROW].height = 20

    for row_idx, r in enumerate(rows, start=HEADER_ROW + 1):
        row_fill = (
            CANCEL_FILL
            if r["is_cancelled"]
            else PENDING_FILL
            if r["status"] == PENDING
            else IN_FILL
            if r["signed_qty"] > 0
            else OUT_FILL
            if r["signed_qty"] < 0
            else _fill("FFFFFF")
        )
        values = [
            r["date"],
            r["doc_number"],
            r["type"],
            r["signed_qty"] if not r["is_cancelled"] else f"({r['signed_qty']})",
            r["unit"],
            "ملغى"
            if r["is_cancelled"]
            else "معلق"
            if r["status"] == PENDING
            else "معتمد",
            r["from_loc"],
            r["to_loc"],
            r["employee"],
            r["notes"],
            r["balance_before"] if r["affects_balance"] else "—",
            r["balance_after"] if r["affects_balance"] else "—",
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = _font(size=9)
            cell.fill = BALANCE_FILL if col >= 11 else row_fill
            cell.border = _border()
            cell.alignment = _align(
                "center" if col in (1, 4, 5, 6, 11, 12) else "right", wrap=True
            )
            if col == 4 and not r["is_cancelled"]:
                cell.font = _font(
                    bold=True,
                    size=9,
                    color="155724" if r["signed_qty"] > 0 else "721C24",
                )
        ws.row_dimensions[row_idx].height = 15

    summary_start = HEADER_ROW + len(rows) + 2
    summary_data = [
        ("إجمالي الوارد المعتمد", f"+{total_in}", TOTAL_IN_FILL, "155724"),
        ("إجمالي الصادر المعتمد", f"-{total_out}", TOTAL_OUT_FILL, "721C24"),
        (
            "الصافي",
            f"{'+' if net_qty >= 0 else ''}{net_qty}",
            TOTAL_IN_FILL if net_qty >= 0 else TOTAL_OUT_FILL,
            "155724" if net_qty >= 0 else "721C24",
        ),
    ]
    if total_cancelled_cnt:
        summary_data.append(
            (
                f"سندات ملغاة ({total_cancelled_cnt} سند)",
                str(total_cancelled_qty),
                _fill("DEE2E6"),
                "495057",
            )
        )
    if total_pending_qty:
        summary_data.append(
            (
                f"حركات معلقة (وارد: +{pending_in} | صادر: -{pending_out})",
                str(total_pending_qty),
                PENDING_FILL,
                "856404",
            )
        )
    summary_data.append(("الرصيد النهائي", str(running_balance), FINAL_FILL, "FFFFFF"))

    for s_idx, (label, value, fill, font_color) in enumerate(summary_data):
        row = summary_start + s_idx
        ws.merge_cells(f"A{row}:K{row}")
        lc, vc = ws[f"A{row}"], ws[f"L{row}"]
        lc.value, lc.font, lc.fill, lc.border, lc.alignment = (
            label,
            _font(bold=True, size=10, color=font_color),
            fill,
            _border(),
            _align("right"),
        )
        vc.value, vc.font, vc.fill, vc.border, vc.alignment = (
            value,
            _font(bold=True, size=11, color=font_color),
            fill,
            _border(),
            _align("center"),
        )
        ws.row_dimensions[row].height = 18

    ws.freeze_panes, ws.auto_filter.ref = (
        f"A{HEADER_ROW + 1}",
        f"A{HEADER_ROW}:L{HEADER_ROW}",
    )
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="سجل_حركة_الصنف_{item.name}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    )
    return response


def _build_inventory_export_rows(faculty, sub_warehouse_id=None, category_id=None):
    """
    Build inventory rows for a specific faculty.
    Shows ALL items for this faculty, even if no FacultyItemStock record exists (QTY=0).
    """
    from inventory.models import FacultyItemStock, Item, ItemCategory, SubWarehouse

    # Filter via FacultyItemStock, NOT Item.sub_warehouse (which doesn't exist)
    items = (
        Item.objects.filter(
            faculty_stocks__faculty=faculty  # Via FacultyItemStock reverse relation
        )
        .select_related(
            "category",
            "category__sub_warehouse",
        )
        .order_by("category__sub_warehouse__name", "category__name", "name")
    )

    if sub_warehouse_id and str(sub_warehouse_id).isdigit():
        try:
            # Filter FacultyItemStock by sub_warehouse, not Item
            items = items.filter(
                faculty_stocks__sub_warehouse_id=sub_warehouse_id
            ).distinct()
        except SubWarehouse.DoesNotExist:
            pass

    if category_id and str(category_id).isdigit():
        try:
            items = items.filter(category_id=category_id)
        except ItemCategory.DoesNotExist:
            pass

    items = list(items)
    item_ids = [item.id for item in items]

    # Fetch FacultyItemStock for these items
    stock_map = {
        stock.item_id: stock
        for stock in FacultyItemStock.objects.filter(
            faculty=faculty,
            item_id__in=item_ids,
        ).select_related("sub_warehouse")
    }

    rows = []
    for index, item in enumerate(items, start=1):
        stock = stock_map.get(item.id)
        qty = stock.cached_quantity if stock else 0
        limit = stock.limit_quantity if stock else item.limit_quantity

        if qty == 0:
            status = "نفذ"
        elif qty <= limit:
            status = "منخفض"
        else:
            status = "متوفر"

        rows.append(
            {
                "index": index,
                "item": item,
                "sub_warehouse": item.category.sub_warehouse if item.category else None,
                "quantity": qty,
                "limit_quantity": limit,
                "status": status,
                "has_stock_record": stock is not None,
            }
        )

    return rows, None, None  # sub_warehouse_obj, category_obj not used here


def _build_admin_inventory_rows(
    faculties, sub_warehouse_id=None, category_id=None, search_q=""
):
    """
    Build shared-catalog inventory rows for one or more faculties.
    Shows ALL items with faculty-specific quantities (including QTY=0).
    """
    # Items are global, filter via FacultyItemStock
    items = Item.objects.select_related(
        "category",
        "category__sub_warehouse",
    ).order_by("category__sub_warehouse__name", "category__name", "name")

    if sub_warehouse_id and str(sub_warehouse_id).isdigit():
        try:
            # Filter via FacultyItemStock
            items = items.filter(
                faculty_stocks__sub_warehouse_id=sub_warehouse_id
            ).distinct()
        except SubWarehouse.DoesNotExist:
            pass

    if category_id and str(category_id).isdigit():
        try:
            items = items.filter(category_id=category_id)
        except ItemCategory.DoesNotExist:
            pass

    search_q = (search_q or "").strip()
    if search_q:
        items = items.filter(Q(name__icontains=search_q) | Q(code__icontains=search_q))

    items = list(items)
    faculties = list(faculties)
    item_ids = [item.id for item in items]
    faculty_ids = [faculty.id for faculty in faculties]

    # Fetch FacultyItemStock for these items/faculties
    stock_map = {
        (stock.faculty_id, stock.item_id): stock
        for stock in FacultyItemStock.objects.filter(
            faculty_id__in=faculty_ids,
            item_id__in=item_ids,
        ).select_related("item", "item__category", "faculty", "sub_warehouse")
    }

    rows = []
    row_index = 1
    for faculty in faculties:
        for item in items:
            stock = stock_map.get((faculty.id, item.id))
            qty = stock.cached_quantity if stock else 0
            limit_quantity = stock.limit_quantity if stock else item.limit_quantity
            last_quantity_update = stock.last_quantity_update if stock else None

            rows.append(
                {
                    "index": row_index,
                    "item": item,
                    "faculty": faculty,
                    "sub_warehouse": item.category.sub_warehouse
                    if item.category
                    else None,
                    "cached_quantity": qty,
                    "limit_quantity": limit_quantity,
                    "last_quantity_update": last_quantity_update,
                }
            )
            row_index += 1

    # Sort by faculty, category's sub_warehouse, category, item name
    rows.sort(
        key=lambda row: (
            row["faculty"].name,
            row["sub_warehouse"].name if row["sub_warehouse"] else "",
            row["item"].category.name if row["item"].category else "",
            row["item"].name,
        )
    )
    return rows, None, None


@login_required
def export_inventory_excel(request):
    """
    Export inventory to Excel.
    Logic: Starts with Item Catalog to include all items (even QTY=0),
    then maps FacultyItemStock quantities.
    """
    user = request.user
    if not (hasattr(user, "profile") and user.profile.faculty):
        return HttpResponseForbidden("ليس لديك صلاحية.")

    faculty = user.profile.faculty

    sub_warehouse_id = request.GET.get("warehouse_id")
    category_id = request.GET.get("category_id")

    # If warehouse is missing, we cannot determine which items to show reliably.
    # Mimicking items_by_warehouse_and_category behavior.
    if not sub_warehouse_id or not sub_warehouse_id.isdigit():
        messages.error(request, "يرجى اختيار مخزن لتصدير الكشف.")
        return redirect("warehouse_inventory_view")

    try:
        sub_warehouse_obj = SubWarehouse.objects.select_related("warehouse").get(
            id=sub_warehouse_id
        )
    except SubWarehouse.DoesNotExist:
        return HttpResponseForbidden("المخزن غير موجود.")

    from inventory.models import FacultyItemStock, Item

    # 1. QUERY ITEMS (Catalog)
    # Start with Item to ensure we see everything, even if stock is 0.
    items_qs = Item.objects.select_related(
        "category", "category__sub_warehouse", "category__sub_warehouse__warehouse"
    ).filter(category__sub_warehouse_id=sub_warehouse_id)

    # Optional Category Filter
    if category_id and category_id.isdigit():
        items_qs = items_qs.filter(category_id=category_id)

    items_list = list(items_qs.order_by("category__name", "name"))
    item_ids = [i.id for i in items_list]

    # 2. QUERY FACULTY ITEM STOCK
    # Fetch stock records for this faculty and the specific sub_warehouse.
    stock_map = {}
    if item_ids:
        stocks = FacultyItemStock.objects.filter(
            faculty=faculty, sub_warehouse_id=sub_warehouse_id, item_id__in=item_ids
        ).select_related("sub_warehouse")
        stock_map = {s.item_id: s for s in stocks}

    # 3. FETCH FILTER OBJECTS FOR HEADER
    category_obj = None
    if category_id and category_id.isdigit():
        category_obj = ItemCategory.objects.filter(id=category_id).first()

    # 4. BUILD DATA
    items_with_qty = []
    for item in items_list:
        stock = stock_map.get(item.id)
        qty = stock.cached_quantity if stock else 0
        limit = stock.limit_quantity if stock else item.limit_quantity

        if qty == 0:
            status = "نفذ"
        elif qty <= limit:
            status = "منخفض"
        else:
            status = "متوفر"

        items_with_qty.append(
            {
                "item": item,
                "quantity": qty,
                "limit_quantity": limit,
                "status": status,
                "sub_warehouse": stock.sub_warehouse
                if stock
                else item.category.sub_warehouse,
                "has_stock_record": stock is not None,
            }
        )

    # 5. GENERATE EXCEL
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "كشف المخزون"
    ws.sheet_view.rightToLeft = True

    # Styles
    def _font(bold=False, size=10, color="000000"):
        return Font(name="Arial", bold=bold, size=size, color=color)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _border():
        thin = Side(style="thin", color="CCCCCC")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap, readingOrder=2)

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"].value = f"كشف المخزون — {faculty.name}"
    ws["A1"].font = _font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = _fill("1A56A5")
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 24

    # Filters Info
    filter_rows = [
        ("الكلية", faculty.name),
        ("المخزن الرئيسي", sub_warehouse_obj.warehouse.name),
        ("المخزن الفرعي", sub_warehouse_obj.name),
        ("الفئة", category_obj.name if category_obj else "الكل"),
        ("تاريخ التصدير", timezone.now().strftime("%Y-%m-%d %H:%M")),
        ("المستخدم", user.get_full_name()),
    ]
    for r_offset, (label, value) in enumerate(filter_rows, start=2):
        ws.merge_cells(f"A{r_offset}:B{r_offset}")
        ws.merge_cells(f"C{r_offset}:H{r_offset}")
        lc, vc = ws[f"A{r_offset}"], ws[f"C{r_offset}"]
        lc.value, vc.value = label, value
        lc.font, vc.font = _font(bold=True, size=9), _font(size=9)
        for cell in (lc, vc):
            cell.fill = _fill("EFF6FF")
            cell.border = _border()
            cell.alignment = _align("right", wrap=True)

    # Headers
    HEADER_ROW = 9
    headers = [
        "م",
        "اسم الصنف",
        "المخزن الفرعي",
        "الفئة",
        "الوحدة",
        "الكمية المتوفرة",
        "الحد الأدنى",
        "الحالة",
    ]
    col_widths = [5, 40, 20, 25, 10, 18, 14, 12]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=HEADER_ROW, column=col, value=h)
        cell.font = _font(bold=True, size=10, color="FFFFFF")
        cell.fill = _fill("343A40")
        cell.border = _border()
        cell.alignment = _align("center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[HEADER_ROW].height = 20
    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:H{HEADER_ROW}"

    # Data Rows
    OUT_FILL, WARN_FILL, OK_FILL = _fill("FFC7CE"), _fill("FFEB9C"), _fill("C6EFCE")

    for idx, data in enumerate(items_with_qty, start=1):
        row = HEADER_ROW + idx
        qty, limit = data["quantity"], data["limit_quantity"]
        status = data["status"]
        row_fill = OUT_FILL if qty == 0 else (WARN_FILL if qty <= limit else OK_FILL)

        values = [
            idx,
            data["item"].name,
            data["sub_warehouse"].name if data["sub_warehouse"] else "-",
            data["item"].category.name if data["item"].category else "—",
            data["item"].get_unit_display(),
            qty,
            limit,
            status,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _font(size=9)
            cell.fill = row_fill
            cell.border = _border()
            cell.alignment = _align(
                "center" if col in (1, 5, 6, 7, 8) else "right", wrap=True
            )
        ws.row_dimensions[row].height = 14

    # Footer
    rows_count = len(items_with_qty)
    last_row = HEADER_ROW + rows_count + 2 if rows_count else HEADER_ROW + 2
    ws.merge_cells(f"A{last_row}:E{last_row}")
    ws[f"A{last_row}"].value = f"إجمالي الأصناف: {rows_count}"
    ws[f"A{last_row}"].font = _font(bold=True, size=10, color="FFFFFF")
    ws[f"A{last_row}"].fill = _fill("1A56A5")
    ws[f"A{last_row}"].alignment = _align("right")
    ws[f"A{last_row}"].border = _border()

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = (
        f"كشف_المخزون_{faculty.name}_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_inventory_pdf(request):
    """
    Export inventory to PDF.
    Logic: Starts with Item Catalog to include all items (even QTY=0).
    """
    user = request.user
    if not (hasattr(user, "profile") and user.profile.faculty):
        return HttpResponseForbidden("ليس لديك صلاحية.")

    faculty = user.profile.faculty
    sub_warehouse_id = request.GET.get("warehouse_id")
    category_id = request.GET.get("category_id")

    if not sub_warehouse_id or not sub_warehouse_id.isdigit():
        messages.error(request, "يرجى اختيار مخزن لتصدير الكشف.")
        return redirect("warehouse_inventory_view")

    try:
        sub_warehouse_obj = SubWarehouse.objects.select_related("warehouse").get(
            id=sub_warehouse_id
        )
    except SubWarehouse.DoesNotExist:
        return HttpResponseForbidden("المخزن غير موجود.")

    from inventory.models import FacultyItemStock, Item

    # 1. QUERY ITEMS
    items_qs = Item.objects.select_related(
        "category", "category__sub_warehouse", "category__sub_warehouse__warehouse"
    ).filter(category__sub_warehouse_id=sub_warehouse_id)

    if category_id and category_id.isdigit():
        items_qs = items_qs.filter(category_id=category_id)

    items_list = list(items_qs.order_by("category__name", "name"))
    item_ids = [i.id for i in items_list]

    # 2. QUERY STOCK
    stock_map = {}
    if item_ids:
        stocks = FacultyItemStock.objects.filter(
            faculty=faculty, sub_warehouse_id=sub_warehouse_id, item_id__in=item_ids
        ).select_related("sub_warehouse")
        stock_map = {s.item_id: s for s in stocks}

    # 3. FETCH DISPLAY OBJECTS
    category_obj = None
    if category_id and category_id.isdigit():
        category_obj = ItemCategory.objects.filter(id=category_id).first()

    # 4. BUILD CONTEXT
    items_with_qty = []
    for item in items_list:
        stock = stock_map.get(item.id)
        qty = stock.cached_quantity if stock else 0
        limit = stock.limit_quantity if stock else item.limit_quantity

        status = "نفذ" if qty == 0 else ("منخفض" if qty <= limit else "متوفر")

        items_with_qty.append(
            {
                "item": item,
                "quantity": qty,
                "limit_quantity": limit,
                "status": status,
                "sub_warehouse": stock.sub_warehouse
                if stock
                else item.category.sub_warehouse,
                "has_stock_record": stock is not None,
            }
        )

    context = {
        "items": items_with_qty,
        "user": user,
        "faculty": faculty,
        "current_date": timezone.now(),
        "warehouse_name": sub_warehouse_obj.warehouse.name,
        "sub_warehouse_name": sub_warehouse_obj.name,
        "category_name": category_obj.name if category_obj else "الكل",
    }

    html_string = render_to_string(
        "inventory/reports/warehouse_items_report.html",
        context,
        request=request,
    )
    html = HTML(string=html_string, base_url=request.build_absolute_uri("/"))

    filename = (
        f"كشف_المخزون_{faculty.name}_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    )
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    html.write_pdf(response)
    return response


@login_required
def export_transaction_pdf(request, transaction_id):
    """Export a specific transaction to PDF."""
    user_type = request.session.get("user_type")
    if not user_type:
        return redirect("login")

    transaction = get_object_or_404(ItemTransactions, id=transaction_id)

    html_string = render_to_string(
        "inventory/reports/transaction_report.html",
        {
            "transaction": transaction,
        },
    )

    html = HTML(string=html_string, base_url=request.build_absolute_uri("/"))
    response = HttpResponse(content_type="application/pdf")
    filename = (
        f"إذن رقم {transaction.document_number}_{timezone.now().strftime('%Y%m%d')}.pdf"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    html.write_pdf(response)
    return response


@login_required
def transaction_audit_log_view(request, transaction_id):
    """View complete audit trail for a transaction"""
    transaction = get_object_or_404(ItemTransactions, id=transaction_id)
    audit_logs = transaction.audit_logs.select_related("performed_by").all()
    return render(
        request,
        "inventory/transaction_audit_log.html",
        {
            "transaction": transaction,
            "audit_logs": audit_logs,
        },
    )


@login_required
def transaction_edit_addition_view(request, transaction_id):
    """Edit ADDITION transaction with proper validation and faculty isolation"""
    transaction = get_object_or_404(ItemTransactions, id=transaction_id, deleted=False)

    # Faculty isolation check
    if not hasattr(request.user, "profile") or not request.user.profile.faculty:
        messages.error(request, "حسابك غير مرتبط بكليّة.")
        return redirect("transaction_list")

    # Use explicit faculty field (new models)
    if not transaction.faculty:
        messages.error(request, "السند غير مرتبط بكليّة. لا يمكن تعديله.")
        return redirect("transaction_list")

    if transaction.faculty != request.user.profile.faculty:
        messages.error(request, "ليس لديك صلاحية تعديل هذا السند - الكليات لا تتطابق.")
        return redirect("transaction_list")

    # Only allow editing of PENDING transactions
    if transaction.approval_status != ItemTransactions.APPROVAL_STATUS.PENDING:
        messages.warning(
            request,
            f"لا يمكن تعديل السند '{transaction.document_number}' لأنه {transaction.get_approval_status_display()}.",
        )
        return redirect("transaction_list")

    if request.method == "POST":
        form = ItemTransactionAdditionForm(
            request.POST,
            instance=transaction,
            user=request.user,
            user_type=request.session.get("user_type"),
        )
        formset = ItemTransactionDetailsAdditionFormSet(
            request.POST, instance=transaction
        )

        if form.is_valid() and formset.is_valid():
            original_data_key = f"transaction_{transaction_id}_original"
            old_data = request.session.get(original_data_key, transaction.to_dict())

            transaction = form.save(commit=False)
            transaction._current_user = request.user
            transaction.modified_by = request.user
            transaction.save()

            formset.instance = transaction
            formset.save()

            # Save prices for new items
            for form_detail in formset:
                if form_detail.cleaned_data and not form_detail.cleaned_data.get(
                    "DELETE", False
                ):
                    price = form_detail.cleaned_data.get("price")
                    item = form_detail.cleaned_data.get("item")
                    if price is not None and price > 0 and item:
                        last_price = (
                            ItemPriceHistory.objects.filter(item=item)
                            .order_by("-date")
                            .first()
                        )
                        if not last_price or abs(last_price.price - price) > 0.01:
                            ItemPriceHistory.objects.create(item=item, price=price)

            log_transaction_action(
                transaction=transaction,
                action=TransactionAuditLog.ACTION_TYPES.UPDATE,
                user=request.user,
                request=request,
                old_data=old_data,
                details=f"تم تعديل سند الإضافة #{transaction.document_number}",
            )

            if original_data_key in request.session:
                del request.session[original_data_key]

            messages.success(
                request, f"تم تعديل سند الإضافة #{transaction.document_number} بنجاح."
            )
            return redirect("transaction_list")
        else:
            # Display errors
            for field_name in form.fields:
                if field_name in form.errors:
                    field_label = form[field_name].label or field_name
                    for error in form.errors[field_name]:
                        messages.error(request, f"{field_label}: {error}")

            for i, form_detail in enumerate(formset):
                if form_detail.errors:
                    for field_name, errors in form_detail.errors.items():
                        field_label = (
                            form_detail[field_name].label
                            if field_name in form_detail.fields
                            else field_name
                        )
                        for error in errors:
                            messages.error(
                                request, f"الصنف #{i + 1} ({field_label}): {error}"
                            )

            if form.non_field_errors():
                for error in form.non_field_errors():
                    messages.error(request, f"خطأ في النموذج: {error}")
            if formset.non_form_errors():
                for error in formset.non_form_errors():
                    messages.error(request, f"خطأ في تفاصيل الأصناف: {error}")

            messages.error(request, "يرجى تصحيح الأخطاء أعلاه والمحاولة مرة أخرى.")
    else:
        form = ItemTransactionAdditionForm(
            instance=transaction,
            user=request.user,
            user_type=request.session.get("user_type"),
        )
        formset = ItemTransactionDetailsAdditionFormSet(instance=transaction)

        original_data_key = f"transaction_{transaction_id}_original"
        if original_data_key not in request.session:
            request.session[original_data_key] = transaction.to_dict()

    def _set_user_display_labels(form_instance):
        user_fields = ["inventory_user", "approval_user"]
        for field_name in user_fields:
            if field_name in form_instance.fields:
                form_instance.fields[field_name].label_from_instance = lambda u: (
                    f"{u.get_full_name() or u.username} ({u.username})"
                    if u.get_full_name()
                    else u.username
                )

    _set_user_display_labels(form)

    return render(
        request,
        "inventory/addition_form.html",
        {
            "form": form,
            "details_formset": formset,
            "editing": True,
            "transaction_id": transaction_id,
            "page_title": f"تعديل سند إضافة #{transaction.document_number}",
        },
    )


@login_required
def transaction_edit_disbursement_view(request, transaction_id):
    """Edit DISBURSEMENT transaction with proper faculty isolation & price tracking"""
    transaction = get_object_or_404(ItemTransactions, id=transaction_id, deleted=False)

    # Faculty isolation check
    user_faculty = getattr(getattr(request.user, "profile", None), "faculty", None)
    if not user_faculty:
        messages.error(request, "حسابك غير مرتبط بكليّة.")
        return redirect("transaction_list")

    if not transaction.faculty:
        messages.error(request, "السند غير مرتبط بكليّة. لا يمكن تعديله.")
        return redirect("transaction_list")

    if transaction.faculty != user_faculty:
        messages.error(request, "ليس لديك صلاحية تعديل هذا السند - الكليات لا تتطابق.")
        return redirect("transaction_list")

    # Only allow editing of PENDING transactions
    if transaction.approval_status != ItemTransactions.APPROVAL_STATUS.PENDING:
        messages.warning(
            request,
            f"لا يمكن تعديل السند '{transaction.document_number}' لأنه {transaction.get_approval_status_display()}.",
        )
        return redirect("transaction_list")

    # Extract context needed for form/formset validation
    from_sub_wh_id = transaction.from_sub_warehouse_id

    if request.method == "POST":
        form = ItemTransactionForm(
            request.POST,
            instance=transaction,
            user=request.user,
            user_type=request.session.get("user_type"),
            from_sub_warehouse_id=from_sub_wh_id,
        )
        # MUST use DisbursementFormSet with faculty context for stock/price validation
        formset = ItemTransactionDetailsDisbursementFormSet(
            request.POST,
            instance=transaction,
            faculty=user_faculty,
            from_sub_warehouse_id=from_sub_wh_id,
        )

        if form.is_valid() and formset.is_valid():
            original_data_key = f"transaction_{transaction_id}_original"
            old_data = request.session.get(original_data_key, transaction.to_dict())

            # Save transaction
            transaction = form.save(commit=False)
            transaction.modified_by = request.user
            transaction.save()

            # Save details
            formset.instance = transaction
            formset.save()

            # Audit log
            log_transaction_action(
                transaction=transaction,
                action=TransactionAuditLog.ACTION_TYPES.UPDATE,
                user=request.user,
                request=request,
                old_data=old_data,
                details=f"تم تعديل سند الصرف #{transaction.document_number}",
            )

            if original_data_key in request.session:
                del request.session[original_data_key]

            messages.success(
                request, f"تم تعديل سند الصرف #{transaction.document_number} بنجاح."
            )
            return redirect("transaction_list")
        else:
            # Simplified error handling (template renders form.errors automatically)
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
            if form.errors or formset.errors:
                logger.warning(
                    f"Disbursement edit validation failed | Form: {form.errors} | Formset: {formset.errors}"
                )

    else:
        form = ItemTransactionForm(
            instance=transaction,
            user=request.user,
            user_type=request.session.get("user_type"),
            from_sub_warehouse_id=from_sub_wh_id,
        )
        formset = ItemTransactionDetailsDisbursementFormSet(
            instance=transaction,
            faculty=user_faculty,
            from_sub_warehouse_id=from_sub_wh_id,
        )

        original_data_key = f"transaction_{transaction_id}_original"
        if original_data_key not in request.session:
            request.session[original_data_key] = transaction.to_dict()

    # Fix user display labels (safe closure pattern)
    def set_user_display_labels(form_instance):
        user_fields = ["to_user", "inventory_user", "from_user", "approval_user"]
        for field_name in user_fields:
            if field_name in form_instance.fields:
                # Default argument captures current field_name safely
                form_instance.fields[field_name].label_from_instance = (
                    lambda u, fn=field_name: (
                        f"{u.get_full_name() or u.username} ({u.username})"
                        if u.get_full_name()
                        else u.username
                    )
                )

    set_user_display_labels(form)

    return render(
        request,
        "inventory/transaction_form.html",
        {
            "form": form,
            "details_formset": formset,
            "editing": True,
            "has_post_errors": request.method == "POST"
            and (form.errors or formset.errors),
            "transaction_id": transaction_id,
            "page_title": f"تعديل سند صرف #{transaction.document_number}",
        },
    )


@login_required
def transaction_edit_transfer_view(request, transaction_id):
    """Edit TRANSFER transaction with faculty isolation"""
    transaction = get_object_or_404(ItemTransactions, id=transaction_id, deleted=False)

    # Faculty isolation check
    if not hasattr(request.user, "profile") or not request.user.profile.faculty:
        messages.error(request, "حسابك غير مرتبط بكليّة.")
        return redirect("transaction_list")

    if not transaction.faculty:
        messages.error(request, "السند غير مرتبط بكليّة. لا يمكن تعديله.")
        return redirect("transaction_list")

    if transaction.faculty != request.user.profile.faculty:
        messages.error(request, "ليس لديك صلاحية تعديل هذا السند - الكليات لا تتطابق.")
        return redirect("transaction_list")

    # Only allow editing of PENDING transactions
    if transaction.approval_status != ItemTransactions.APPROVAL_STATUS.PENDING:
        messages.warning(
            request,
            f"لا يمكن تعديل السند '{transaction.document_number}' لأنه {transaction.get_approval_status_display()}.",
        )
        return redirect("transaction_list")

    if request.method == "POST":
        form = ItemTransactionTransferForm(
            request.POST,
            instance=transaction,
            user=request.user,
            user_type=request.session.get("user_type"),
        )
        formset = ItemTransactionDetailsTransferFormSet(
            request.POST, instance=transaction
        )

        if form.is_valid() and formset.is_valid():
            original_data_key = f"transaction_{transaction_id}_original"
            old_data = request.session.get(original_data_key)

            transaction = form.save(commit=False)
            transaction._current_user = request.user
            transaction.modified_by = request.user
            transaction.save()
            formset.instance = transaction
            formset.save()

            log_transaction_action(
                transaction=transaction,
                action=TransactionAuditLog.ACTION_TYPES.UPDATE,
                user=request.user,
                request=request,
                old_data=old_data,
            )

            if original_data_key in request.session:
                del request.session[original_data_key]

            messages.success(request, "تم تعديل سند النقل بنجاح.")
            return redirect("transaction_list")
        else:
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
    else:
        form = ItemTransactionTransferForm(
            instance=transaction,
            user=request.user,
            user_type=request.session.get("user_type"),
        )
        formset = ItemTransactionDetailsTransferFormSet(instance=transaction)

    return render(
        request,
        "inventory/transfer_form.html",
        {
            "form": form,
            "details_formset": formset,
            "editing": True,
            "transaction_id": transaction_id,
        },
    )


@login_required
def transaction_edit_return_view(request, transaction_id):
    """Edit RETURN transaction with faculty isolation"""
    transaction = get_object_or_404(ItemTransactions, id=transaction_id, deleted=False)

    # Faculty isolation check
    if not hasattr(request.user, "profile") or not request.user.profile.faculty:
        messages.error(request, "حسابك غير مرتبط بكليّة.")
        return redirect("transaction_list")

    if not transaction.faculty:
        messages.error(request, "السند غير مرتبط بكليّة. لا يمكن تعديله.")
        return redirect("transaction_list")

    if transaction.faculty != request.user.profile.faculty:
        messages.error(request, "ليس لديك صلاحية تعديل هذا السند - الكليات لا تتطابق.")
        return redirect("transaction_list")

    # Only allow editing of PENDING transactions
    if transaction.approval_status != ItemTransactions.APPROVAL_STATUS.PENDING:
        messages.warning(
            request,
            f"لا يمكن تعديل السند '{transaction.document_number}' لأنه {transaction.get_approval_status_display()}.",
        )
        return redirect("transaction_list")

    if request.method == "POST":
        form = ItemTransactionReturnForm(
            request.POST,
            instance=transaction,
            user=request.user,
            user_type=request.session.get("user_type"),
        )
        formset = ItemTransactionDetailsReturnFormSet(
            request.POST, instance=transaction
        )

        if form.is_valid() and formset.is_valid():
            original_data_key = f"transaction_{transaction_id}_original"
            old_data = request.session.get(original_data_key)
            transaction = form.save(commit=False)
            transaction._current_user = request.user
            transaction.modified_by = request.user
            transaction.save()
            formset.instance = transaction
            formset.save()

            log_transaction_action(
                transaction=transaction,
                action=TransactionAuditLog.ACTION_TYPES.UPDATE,
                user=request.user,
                request=request,
                old_data=old_data,
            )

            if original_data_key in request.session:
                del request.session[original_data_key]

            messages.success(request, "تم تعديل سند الإرجاع بنجاح.")
            return redirect("transaction_list")
        else:
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
    else:
        form = ItemTransactionReturnForm(
            instance=transaction,
            user=request.user,
            user_type=request.session.get("user_type"),
        )
        formset = ItemTransactionDetailsReturnFormSet(instance=transaction)

    return render(
        request,
        "inventory/return_form.html",
        {
            "form": form,
            "details_formset": formset,
            "editing": True,
            "transaction_id": transaction_id,
        },
    )


@login_required
def transaction_list_view(request):
    """Optimized list transactions with role-based filtering, search, and pagination."""
    user_type = request.session.get("user_type")
    user = request.user

    # Cache open_year to avoid duplicate queries
    open_year = InventoryYear.get_open_year()

    # Ensure user has faculty
    if not (hasattr(user, "profile") and user.profile.faculty):
        messages.error(request, "ليس لديك كليّة مرتبطة بحسابك. يرجى الاتصال بالمسؤول.")
        return redirect("home")

    user_faculty = user.profile.faculty

    try:
        # STEP 1: Build base queryset with OPTIMIZED select_related/prefetch
        base_qs = (
            ItemTransactions.objects.filter(
                year=open_year,
                deleted=False,
            )
            .select_related(
                "created_by",
                "from_sub_warehouse",
                "to_sub_warehouse",
                "from_department",
                "to_department",
                "from_user",
                "to_user",
                "inventory_user",
                "approval_user",
                "year",
                "faculty",
            )
            .prefetch_related(
                # ✅ Only prefetch items if needed for search/display
                Prefetch(
                    "itemtransactiondetails_set",
                    queryset=ItemTransactionDetails.objects.select_related("item").only(
                        "id", "transaction_id", "item_id", "approved_quantity"
                    ),
                    to_attr="details_cached",
                )
            )
        )

        # STEP 2: Apply ROLE-BASED filtering EARLY (reduces dataset size)
        if user_type == "administration_manager":
            # Admin sees all transactions across ALL faculties for open year
            qs = base_qs
        elif user_type == "faculty_manager":
            qs = base_qs.filter(faculty=user_faculty)
        elif user_type == "inventory_manager":
            qs = base_qs.filter(
                Q(approval_status=ItemTransactions.APPROVAL_STATUS.PENDING)
                | Q(approval_user=user),
                faculty=user_faculty,
            )
        elif user_type == "inventory_employee":
            qs = base_qs.filter(created_by=user, faculty=user_faculty)
        else:
            # Regular users: transactions they created or are involved in
            qs = base_qs.filter(
                Q(created_by=user) | Q(from_user=user) | Q(to_user=user),
                faculty=user_faculty,
            )

        # STEP 3: Apply filters (status, search, date) BEFORE counting
        # Approval status filter
        approval_status = request.GET.get("approval_status")
        if approval_status and approval_status in dict(
            ItemTransactions.APPROVAL_STATUS.choices
        ):
            qs = qs.filter(approval_status=approval_status)

        # Search filter - OPTIMIZED: avoid expensive joins when possible
        search_query = request.GET.get("search", "").strip()
        if search_query:
            # First try simple fields (faster)
            qs = qs.filter(
                Q(document_number__icontains=search_query)
                | Q(notes__icontains=search_query)
                | Q(created_by__first_name__icontains=search_query)
                | Q(created_by__last_name__icontains=search_query)
            )
            # Only add item name search if needed (more expensive)
            if len(search_query) >= 3:  # Avoid short searches on item names
                qs = qs.filter(
                    Q(itemtransactiondetails__item__name__icontains=search_query)
                ).distinct()  # Only use distinct when joining itemtransactiondetails

        # Date range filter
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        if start_date and end_date:
            try:
                start_dt = timezone.make_aware(
                    datetime.strptime(start_date, "%Y-%m-%d")
                )
                end_dt = timezone.make_aware(
                    datetime.strptime(end_date, "%Y-%m-%d")
                ) + timedelta(days=1)
                qs = qs.filter(created_at__range=(start_dt, end_dt))
            except (ValueError, TypeError):
                messages.warning(
                    request, "تنسيق التاريخ غير صحيح. سيتم عرض جميع السندات."
                )

        # STEP 4: Calculate statistics in SINGLE aggregate query (not 4 separate counts)
        stats = qs.aggregate(
            total=Count("id", distinct=True),
            pending=Count(
                "id",
                filter=Q(approval_status=ItemTransactions.APPROVAL_STATUS.PENDING),
                distinct=True,
            ),
            approved=Count(
                "id",
                filter=Q(approval_status=ItemTransactions.APPROVAL_STATUS.APPROVED),
                distinct=True,
            ),
            rejected=Count(
                "id",
                filter=Q(approval_status=ItemTransactions.APPROVAL_STATUS.REJECTED),
                distinct=True,
            ),
        )

        # STEP 5: Apply ordering
        qs = qs.order_by("-created_at")

        # STEP 6: Prepare context
        context = {
            "transactions": qs,
            "approval_status_choices": ItemTransactions.APPROVAL_STATUS.choices,
            "current_approval_status": approval_status,
            "search_query": search_query,
            "start_date": start_date,
            "end_date": end_date,
            "user_type": user_type,
            "faculty_name": user_faculty.name,
            "total_transactions": stats["total"] or 0,
            "pending_count": stats["pending"] or 0,
            "approved_count": stats["approved"] or 0,
            "rejected_count": stats["rejected"] or 0,
            "open_year": open_year,
        }

        return render(request, "inventory/transaction_list.html", context)

    except Exception as e:
        logger.error(f"Error in transaction_list_view: {str(e)}", exc_info=True)
        messages.error(
            request, "حدث خطأ أثناء تحميل قائمة السندات. يرجى المحاولة مرة أخرى."
        )
        return redirect("home")


@login_required
def transaction_delete_view(request, transaction_id):
    """Soft delete transaction with confirmation."""
    # Security check - only inventory managers can delete
    # if request.session.get("user_type") != "inventory_manager":
    #     messages.error(request, "ليس لديك صلاحية حذف السندات.")
    #     return redirect("transaction_list")

    # Faculty isolation check
    if not hasattr(request.user, "profile") or not request.user.profile.faculty:
        messages.error(request, "حسابك غير مرتبط بكليّة.")
        return redirect("transaction_list")

    transaction = get_object_or_404(ItemTransactions, id=transaction_id, deleted=False)

    # Verify transaction belongs to user's faculty
    if transaction.created_by.profile.faculty != request.user.profile.faculty:
        messages.error(request, "ليس لديك صلاحية حذف هذا السند - الكليات لا تتطابق.")
        return redirect("transaction_list")

    # Only allow deletion of PENDING transactions
    if transaction.approval_status != ItemTransactions.APPROVAL_STATUS.PENDING:
        messages.warning(
            request,
            f"لا يمكن حذف السند '{transaction.document_number}' لأنه {transaction.get_approval_status_display()}.",
        )
        return redirect("transaction_list")

    if request.method == "POST":
        try:
            with db_transaction.atomic():
                snapshot = transaction.to_dict()

                transaction.deleted = True
                transaction.deleted_by = request.user
                transaction.deleted_at = timezone.now()
                transaction.approval_status = ItemTransactions.APPROVAL_STATUS.DELETED
                transaction.save(
                    update_fields=[
                        "deleted",
                        "deleted_by",
                        "deleted_at",
                        "approval_status",
                    ]
                )

                log_transaction_action(
                    transaction=transaction,
                    action=TransactionAuditLog.ACTION_TYPES.DELETE,
                    user=request.user,
                    request=request,
                    old_data=snapshot,
                    details=f"تم حذف السند برقم {transaction.document_number}",
                )

                messages.success(
                    request, f"تم حذف السند '{transaction.document_number}' بنجاح."
                )
                return redirect("transaction_list")

        except Exception as e:
            logger.error(f"Error deleting transaction: {str(e)}", exc_info=True)
            messages.error(request, "حدث خطأ أثناء حذف السند. يرجى المحاولة مرة أخرى.")
            return redirect("transaction_list")

    # GET request - show confirmation page
    return render(
        request, "inventory/confirm_delete.html", {"transaction": transaction}
    )


@login_required
def transaction_edit_view(request, transaction_id):
    """Redirect to appropriate edit form based on transaction type."""
    # user_type = request.session.get("user_type")
    # if user_type == "inventory_manager":
    #     messages.error(
    #         request,
    #         "ليس لديك صلاحية تعديل السندات. يمكنك فقط اعتمادها أو رفضها.",
    #     )
    #     return redirect("transaction_detail", pk=transaction_id)

    transaction = get_object_or_404(ItemTransactions, id=transaction_id, deleted=False)
    request.session[f"transaction_{transaction_id}_original"] = transaction.to_dict()

    if transaction.transaction_type == ItemTransactions.TRANSACTION_TYPES.Addition:
        return redirect("transaction_edit_addition", transaction_id=transaction_id)
    elif (
        transaction.transaction_type == ItemTransactions.TRANSACTION_TYPES.Disbursement
    ):
        return redirect("transaction_edit_disbursement", transaction_id=transaction_id)
    elif transaction.transaction_type == ItemTransactions.TRANSACTION_TYPES.Transfer:
        return redirect("transaction_edit_transfer", transaction_id=transaction_id)
    elif transaction.transaction_type == ItemTransactions.TRANSACTION_TYPES.Return:
        return redirect("transaction_edit_return", transaction_id=transaction_id)

    messages.error(request, "نوع السند غير معروف.")
    return redirect("transaction_list")


@login_required
def transaction_create_addition_view(request):
    """Create ADDITION transaction with price recording - NO quantity updates on creation"""
    user_type = request.session.get("user_type")
    to_sub_warehouse_id = (
        request.POST.get("to_sub_warehouse") if request.method == "POST" else None
    )

    if request.method == "POST":
        post_data = request.POST.copy()
        post_data["created_by"] = request.user.id
        post_data["approval_status"] = ItemTransactions.APPROVAL_STATUS.PENDING

        form = ItemTransactionAdditionForm(
            post_data,
            user=request.user,
            user_type=user_type,
            to_warehouse_id=to_sub_warehouse_id,
        )
        formset = ItemTransactionDetailsAdditionFormSet(
            post_data,
            queryset=ItemTransactionDetails.objects.none(),
        )

        if form.is_valid() and formset.is_valid():
            new_transaction = None
            try:
                with db_transaction.atomic():
                    # Form saves with inventory_user already set
                    new_transaction = form.save()
                    formset.instance = new_transaction
                    instances = formset.save()

                    # Record price history
                    for detail in instances:
                        if (
                            detail.price is not None
                            and detail.price > 0
                            and detail.item
                        ):
                            last_price = (
                                ItemPriceHistory.objects.filter(item=detail.item)
                                .order_by("-date")
                                .first()
                            )
                            if not last_price or last_price.price != detail.price:
                                ItemPriceHistory.objects.create(
                                    item=detail.item, price=detail.price
                                )
                                logger.info(
                                    f"Price history recorded for item {detail.item.id}: {detail.price}"
                                )

                    # Log pending creation
                    log_transaction_action(
                        transaction=new_transaction,
                        action=TransactionAuditLog.ACTION_TYPES.CREATE,
                        user=request.user,
                        request=request,
                        details=f"تم إنشاء سند إضافة برقم {new_transaction.document_number} في انتظار الاعتماد",
                    )

                    messages.success(
                        request,
                        "تم إنشاء سند الإضافة بنجاح. السند ينتظر اعتماد المشرف.",
                    )
                    return redirect("transaction_list")

            except Exception as e:
                error_msg = f"حدث خطأ أثناء إنشاء السند: {str(e)}"
                if new_transaction:
                    error_msg += f" | السند: {new_transaction.document_number}"
                    try:
                        new_transaction.delete()
                    except Exception as e:
                        error_msg += f" | فشل حذف السند: {str(e)}"
                messages.error(request, error_msg)
                logger.error(f"Transaction creation error: {str(e)}", exc_info=True)
        else:
            logger.error(f"Form errors: {form.errors}")
            logger.error(f"Formset errors: {formset.errors}")
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
    else:
        # GET request - form will auto-set inventory_user in __init__
        form = ItemTransactionAdditionForm(
            initial={
                "approval_status": ItemTransactions.APPROVAL_STATUS.PENDING,
                "created_by": request.user.id,
            },
            user=request.user,
            user_type=user_type,
            to_warehouse_id=to_sub_warehouse_id,
        )
        formset = ItemTransactionDetailsAdditionFormSet(
            queryset=ItemTransactionDetails.objects.none()
        )

    context = {
        "form": form,
        "details_formset": formset,
        "editing": False,
    }
    return render(request, "inventory/addition_form.html", context)


@login_required
def transaction_create_disbursement_view(request):
    user_type = request.session.get("user_type")
    user_faculty = getattr(getattr(request.user, "profile", None), "faculty", None)

    if not user_faculty:
        messages.error(request, "حسابك غير مرتبط بكلية.")
        return redirect("home")

    from_sub_warehouse_id = (
        request.POST.get("from_sub_warehouse") if request.method == "POST" else None
    )

    if request.method == "POST":
        post_data = request.POST.copy()
        post_data["created_by"] = request.user.id
        post_data["approval_status"] = ItemTransactions.APPROVAL_STATUS.PENDING
        post_data["faculty"] = user_faculty.id

        form = ItemTransactionForm(
            post_data,
            user=request.user,
            user_type=user_type,
            from_sub_warehouse_id=from_sub_warehouse_id,
        )

        formset = ItemTransactionDetailsDisbursementFormSet(
            post_data,
            queryset=ItemTransactionDetails.objects.none(),
            faculty=user_faculty,
            from_sub_warehouse_id=from_sub_warehouse_id,
        )

        if form.is_valid() and formset.is_valid():
            try:
                with db_transaction.atomic():
                    tx = form.save()
                    formset.instance = tx
                    formset.save()

                    messages.success(request, "تم إنشاء سند الصرف بنجاح.")
                    return redirect("transaction_list")
            except Exception as e:
                messages.error(request, f"خطأ أثناء الحفظ: {e}")
                logger.error(f"Disbursement creation error: {e}", exc_info=True)
        else:
            if form.errors:
                logger.error(f"Main form errors: {form.errors}")
            if formset.errors:
                for i, err in enumerate(formset.errors):
                    if err:  # Only log non-empty errors
                        logger.error(f"Formset form [{i}] errors: {err}")
            if formset.non_form_errors():
                logger.error(f"Formset non-form errors: {formset.non_form_errors()}")

            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
    else:
        form = ItemTransactionForm(
            initial={"faculty": user_faculty.id, "created_by": request.user.id},
            user=request.user,
            user_type=user_type,
        )
        formset = ItemTransactionDetailsDisbursementFormSet(
            queryset=ItemTransactionDetails.objects.none(),
            faculty=user_faculty,
            from_sub_warehouse_id=None,
        )

    return render(
        request,
        "inventory/transaction_form.html",
        {"form": form, "details_formset": formset, "editing": False},
    )


@login_required
def transaction_create_transfer_view(request):
    """Create TRANSFER transaction with dynamic custody handling."""
    user_type = request.session.get("user_type")
    user = request.user

    # Security check: Only inventory managers and employees can create transfers
    if user_type not in ["inventory_manager", "inventory_employee"]:
        messages.error(request, "ليس لديك صلاحية إنشاء سندات النقل.")
        return redirect("transaction_list")

    # Faculty check
    if not (hasattr(user, "profile") and user.profile.faculty):
        messages.error(request, "ليس لديك كليّة مرتبطة بحسابك.")
        return redirect("transaction_list")

    if request.method == "POST":
        form = ItemTransactionTransferForm(
            request.POST,
            user=user,
            user_type=user_type,
        )

        if "from_user" in form.fields:
            form.fields["from_user"].queryset = User.objects.filter(
                profile__faculty=user.profile.faculty
            )
        if "to_user" in form.fields:
            form.fields["to_user"].queryset = User.objects.filter(
                profile__faculty=user.profile.faculty
            )
        formset = ItemTransactionDetailsTransferFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            #  CRITICAL FIX: Remove stock validation for PENDING transactions
            # No validation needed since quantities won't be updated until approval
            try:
                with db_transaction.atomic():
                    # Save transaction
                    transaction_obj = form.save(commit=False)
                    transaction_obj.created_by = user
                    transaction_obj.transaction_type = (
                        ItemTransactions.TRANSACTION_TYPES.Transfer
                    )

                    # SET INVENTORY_USER SERVER-SIDE (Secure & bypasses form validation)
                    transaction_obj.inventory_user = user

                    # Set approval status to PENDING
                    transaction_obj.approval_status = (
                        ItemTransactions.APPROVAL_STATUS.PENDING
                    )
                    transaction_obj.save()

                    # Save formset details
                    formset.instance = transaction_obj
                    formset.save()

                    # LOG PENDING CREATION
                    log_transaction_action(
                        transaction=transaction_obj,
                        action=TransactionAuditLog.ACTION_TYPES.CREATE,
                        user=user,
                        request=request,
                        details=f"تم إنشاء سند نقل برقم {transaction_obj.document_number} في انتظار الاعتماد",
                    )

                    messages.success(
                        request, "تم حفظ سند النقل بنجاح. السند ينتظر اعتماد المشرف."
                    )
                    return redirect("transaction_list")
            except Exception as e:
                logger.error(
                    f"Error saving transfer transaction: {str(e)}", exc_info=True
                )
                messages.error(
                    request, "حدث خطأ أثناء حفظ السند. يرجى المحاولة مرة أخرى."
                )
        else:
            # Handle form validation errors
            error_messages = []

            # Form errors
            for field_name, errors in form.errors.items():
                field_label = (
                    form[field_name].label if field_name in form.fields else field_name
                )
                for error in errors:
                    error_messages.append(f"{field_label}: {error}")

            # Formset errors
            for i, form_detail in enumerate(formset):
                for field_name, errors in form_detail.errors.items():
                    field_label = (
                        form_detail[field_name].label
                        if field_name in form_detail.fields
                        else field_name
                    )
                    for error in errors:
                        error_messages.append(f"الصنف {i + 1} - {field_label}: {error}")

            if error_messages:
                messages.error(
                    request,
                    "يرجى تصحيح الأخطاء في النموذج:<br>" + "<br>".join(error_messages),
                    extra_tags="safe",
                )
            else:
                messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
    else:
        # GET request - initialize form
        form = ItemTransactionTransferForm(
            user=user,
            user_type=user_type,
        )
        formset = ItemTransactionDetailsTransferFormSet()

        # Initialize with empty querysets to prevent pre-population
        _initialize_empty_transfer_fields(form)

    # Prepare context for template
    context = {
        "form": form,
        "details_formset": formset,
        "editing": False,
    }
    return render(request, "inventory/transfer_form.html", context)


@login_required
def transaction_create_return_view(request):
    """Create RETURN transaction with proper created_by handling and validation."""
    user_type = request.session.get("user_type")

    if request.method == "POST":
        form = ItemTransactionReturnForm(
            request.POST,
            user=request.user,
            user_type=user_type,
        )
        formset = ItemTransactionDetailsReturnFormSet(request.POST)

        #  CRITICAL DEBUG: Log form data for debugging
        logger.debug("=== RETURN TRANSACTION FORM DATA ===")
        logger.debug(f"Form data: {request.POST.dict()}")
        logger.debug(f"User: {request.user.username}, User type: {user_type}")

        if form.is_valid() and formset.is_valid():
            from_user = form.cleaned_data.get("from_user")
            to_sub_warehouse = form.cleaned_data.get("to_sub_warehouse")

            #  Validate required fields
            validation_errors = []
            if not from_user:
                validation_errors.append("يجب تحديد الموظف المرتجع منه.")
            if not to_sub_warehouse:
                validation_errors.append("يجب تحديد المخزن الفرعي المستقبل للإرجاع.")

            if validation_errors:
                for error in validation_errors:
                    messages.error(request, error)
                return render(
                    request,
                    "inventory/return_form.html",
                    {"form": form, "details_formset": formset},
                )

            #  CRITICAL FIX: Remove stock validation for PENDING transactions
            # No validation needed since quantities won't be updated until approval

            try:
                with db_transaction.atomic():
                    #  FORM HANDLES created_by automatically
                    transaction_obj = form.save(commit=False)

                    #  Set faculty from user profile
                    if (
                        hasattr(request.user, "profile")
                        and request.user.profile.faculty
                    ):
                        transaction_obj.faculty = request.user.profile.faculty

                    #  Ensure approval_status is PENDING
                    transaction_obj.approval_status = (
                        ItemTransactions.APPROVAL_STATUS.PENDING
                    )
                    transaction_obj.save()

                    # Save formset details
                    formset.instance = transaction_obj
                    saved_details = formset.save()

                    logger.info(
                        f"Successfully saved return transaction {transaction_obj.id} with {len(saved_details)} items"
                    )

                    #  LOG PENDING CREATION
                    log_transaction_action(
                        transaction=transaction_obj,
                        action=TransactionAuditLog.ACTION_TYPES.CREATE,
                        user=request.user,
                        request=request,
                        details=f"تم إنشاء سند إرجاع برقم {transaction_obj.document_number} في انتظار الاعتماد",
                    )

                    messages.success(
                        request,
                        f"تم حفظ سند الإرجاع بنجاح! السند ينتظر اعتماد المشرف. رقم السند: {transaction_obj.document_number}",
                    )
                    return redirect("transaction_list")
            except Exception as e:
                logger.error(
                    f"Unexpected error saving return transaction: {str(e)}",
                    exc_info=True,
                )
                messages.error(
                    request,
                    f"حدث خطأ غير متوقع أثناء حفظ سند الإرجاع: {str(e)}. يرجى المحاولة مرة أخرى لاحقاً.",
                )
        else:
            #  Detailed form error reporting
            logger.warning("Form validation failed for return transaction")

            # Main form errors
            for field_name, errors in form.errors.items():
                if field_name == "__all__":
                    for error in errors:
                        messages.error(request, f"خطأ عام: {error}")
                else:
                    field_label = (
                        form[field_name].label if field_name in form else field_name
                    )
                    for error in errors:
                        messages.error(request, f"{field_label}: {error}")

            # Formset errors
            for i, form_errors in enumerate(formset.errors):
                if form_errors:
                    for field_name, errors in form_errors.items():
                        field_label = (
                            formset.forms[i][field_name].label
                            if field_name in formset.forms[i]
                            else field_name
                        )
                        for error in errors:
                            messages.error(
                                request, f"الصنف {i + 1} - {field_label}: {error}"
                            )
    else:
        #  GET request - initialize form with proper defaults
        form = ItemTransactionReturnForm(
            initial={
                "created_by": request.user.id,  #  Set initial created_by value
                "transaction_type": ItemTransactions.TRANSACTION_TYPES.Return,
                "castody_type": ItemTransactions.CASTODY_TYPES.Warehouse,
                "approval_status": ItemTransactions.APPROVAL_STATUS.PENDING,  #  Set initial approval_status
            },
            user=request.user,
            user_type=user_type,
        )
        formset = ItemTransactionDetailsReturnFormSet()

        # Initialize with empty querysets
        form.fields["from_user"].queryset = User.objects.none()
        form.fields["from_user"].widget.attrs["disabled"] = "disabled"
        form.fields["inventory_user"].queryset = User.objects.none()
        form.fields["inventory_user"].widget.attrs["disabled"] = "disabled"

        #  Set default inventory_user to current user
        form.fields["inventory_user"].initial = request.user.id

    #  Context preparation
    context = {
        "form": form,
        "details_formset": formset,
        "editing": False,
    }

    return render(request, "inventory/return_form.html", context)


def _initialize_empty_transfer_fields(form):
    """Initialize form with empty querysets for dynamic fields."""
    # Warehouse custody fields - keep choices but no initial selection
    if "from_sub_warehouse" in form.fields:
        form.fields["from_sub_warehouse"].initial = None
    if "to_sub_warehouse" in form.fields:
        form.fields["to_sub_warehouse"].initial = None

    # Department custody fields
    if "from_department" in form.fields:
        form.fields["from_department"].initial = None
    if "to_department" in form.fields:
        form.fields["to_department"].initial = None

    # User fields - start with empty querysets and disabled state
    if "from_user" in form.fields:
        form.fields["from_user"].queryset = User.objects.none()
        form.fields["from_user"].widget.attrs["disabled"] = "disabled"
        form.fields["from_user"].initial = None

    if "to_user" in form.fields:
        form.fields["to_user"].queryset = User.objects.none()
        form.fields["to_user"].widget.attrs["disabled"] = "disabled"
        form.fields["to_user"].initial = None

    # Set default custody type to Warehouse if field exists
    if "castody_type" in form.fields:
        form.fields["castody_type"].initial = ItemTransactions.CASTODY_TYPES.Warehouse


@login_required
def transaction_approve_view(request, pk):
    """Approve transaction and update quantities - AJAX ONLY (no GET handling)"""
    transaction_obj = get_object_or_404(ItemTransactions, pk=pk)
    user = request.user

    # Faculty isolation check
    if not hasattr(
        transaction_obj, "can_be_approved_by_user"
    ) or not transaction_obj.can_be_approved_by_user(user):
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "error": "ليس لديك صلاحية اعتماد هذا السند. الكليات لا تتطابق.",
                },
                status=403,
            )
        return HttpResponseForbidden(
            "ليس لديك صلاحية اعتماد هذا السند. الكليات لا تتطابق."
        )

    if transaction_obj.approval_status != ItemTransactions.APPROVAL_STATUS.PENDING:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "error": f"السند '{transaction_obj.document_number}' ليس في حالة انتظار الاعتماد.",
                },
                status=400,
            )
        messages.warning(
            request,
            f"السند '{transaction_obj.document_number}' ليس في حالة انتظار الاعتماد.",
        )
        return redirect("transaction_list")

    # ONLY handle POST requests (no GET rendering)
    if request.method != "POST":
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "error": "هذا السند يتطلب طلب POST لاعتماده.",
                },
                status=405,
            )
        messages.error(
            request, "الطلب غير صالح. يرجى استخدام النموذج الصحيح لاعتماد السند."
        )
        return redirect("transaction_list")

    approval_notes = request.POST.get("approval_notes", "").strip()
    try:
        with db_transaction.atomic():
            # ONLY UPDATE QUANTITIES HERE - AFTER APPROVAL
            success = transaction_obj.approve(user, approval_notes)

            if not success:
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse(
                        {
                            "success": False,
                            "error": "فشل في تحديث الكميات. يرجى المحاولة مرة أخرى.",
                        },
                        status=500,
                    )

            # Log approval action
            log_transaction_action(
                transaction=transaction_obj,
                action=TransactionAuditLog.ACTION_TYPES.APPROVE,
                user=user,
                request=request,
                details=f"تم اعتماد السند برقم {transaction_obj.document_number} بواسطة {user.get_full_name()}: {approval_notes}",
            )

            # Return JSON response for AJAX requests
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse(
                    {
                        "success": True,
                        "message": f"تم اعتماد السند بنجاح. الكميات تم تحديثها في المخزون. رقم السند: {transaction_obj.document_number}",
                        "redirect_url": reverse("transaction_list"),
                    }
                )

            messages.success(
                request,
                f"تم اعتماد السند بنجاح. الكميات تم تحديثها في المخزون. رقم السند: {transaction_obj.document_number}",
            )
            return redirect("transaction_list")

    except Exception as e:
        logger.error(f"Error approving transaction: {str(e)}", exc_info=True)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "error": f"حدث خطأ أثناء اعتماد السند: {str(e)}",
                },
                status=500,
            )
        messages.error(request, "حدث خطأ أثناء اعتماد السند. يرجى المحاولة مرة أخرى.")
        return redirect("transaction_list")


@login_required
def transaction_reject_view(request, pk):
    """Reject transaction - NO quantity updates - AJAX ONLY (no GET handling)"""
    transaction_obj = get_object_or_404(ItemTransactions, pk=pk)
    user = request.user

    # Faculty isolation check
    if not hasattr(
        transaction_obj, "can_be_approved_by_user"
    ) or not transaction_obj.can_be_approved_by_user(user):
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "error": "ليس لديك صلاحية رفض هذا السند. الكليات لا تتطابق.",
                },
                status=403,
            )
        return HttpResponseForbidden(
            "ليس لديك صلاحية رفض هذا السند. الكليات لا تتطابق."
        )

    if transaction_obj.approval_status != ItemTransactions.APPROVAL_STATUS.PENDING:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "error": f"السند '{transaction_obj.document_number}' ليس في حالة انتظار الاعتماد.",
                },
                status=400,
            )
        messages.warning(
            request,
            f"السند '{transaction_obj.document_number}' ليس في حالة انتظار الاعتماد.",
        )
        return redirect("transaction_list")

    # ONLY handle POST requests (no GET rendering)
    if request.method != "POST":
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "error": "هذا السند يتطلب طلب POST لرفضه.",
                },
                status=405,
            )
        messages.error(
            request, "الطلب غير صالح. يرجى استخدام النموذج الصحيح لرفض السند."
        )
        return redirect("transaction_list")

    approval_notes = request.POST.get("approval_notes", "").strip()
    try:
        with db_transaction.atomic():
            # Reject transaction (no quantity updates)
            transaction_obj.reject(user, approval_notes)

            # Log rejection action
            log_transaction_action(
                transaction=transaction_obj,
                action=TransactionAuditLog.ACTION_TYPES.REJECT,
                user=user,
                request=request,
                details=f"تم رفض السند برقم {transaction_obj.document_number} بواسطة {user.get_full_name()}: {approval_notes}",
            )

            # Return JSON response for AJAX requests
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse(
                    {
                        "success": True,
                        "message": f"تم رفض السند بنجاح. رقم السند: {transaction_obj.document_number}",
                        "redirect_url": reverse("transaction_list"),
                    }
                )

            messages.success(
                request,
                f"تم رفض السند بنجاح. لم يتم تحديث كميات المخزون. رقم السند: {transaction_obj.document_number}",
            )
            return redirect("transaction_list")

    except Exception as e:
        logger.error(f"Error rejecting transaction: {str(e)}", exc_info=True)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "error": f"حدث خطأ أثناء رفض السند: {str(e)}",
                },
                status=500,
            )
        messages.error(request, "حدث خطأ أثناء رفض السند. يرجى المحاولة مرة أخرى.")
        return redirect("transaction_list")


def _initialize_empty_transfer_fields(form):
    """Initialize form with empty querysets for dynamic fields."""
    # Warehouse custody fields - keep choices but no initial selection
    if "from_sub_warehouse" in form.fields:
        form.fields["from_sub_warehouse"].initial = None
    if "to_sub_warehouse" in form.fields:
        form.fields["to_sub_warehouse"].initial = None

    # Department custody fields
    if "from_department" in form.fields:
        form.fields["from_department"].initial = None
    if "to_department" in form.fields:
        form.fields["to_department"].initial = None

    # User fields - start with empty querysets and disabled state
    if "from_user" in form.fields:
        form.fields["from_user"].queryset = User.objects.none()
        form.fields["from_user"].widget.attrs["disabled"] = "disabled"
        form.fields["from_user"].initial = None

    if "to_user" in form.fields:
        form.fields["to_user"].queryset = User.objects.none()
        form.fields["to_user"].widget.attrs["disabled"] = "disabled"
        form.fields["to_user"].initial = None

    # Set default custody type to Warehouse if field exists
    if "castody_type" in form.fields:
        form.fields["castody_type"].initial = ItemTransactions.CASTODY_TYPES.Warehouse


def _validate_transfer_stock(form, formset, user_profile):
    """
    Validate stock availability for transfer transactions based on custody type.

    Args:
        form: ItemTransactionTransferForm instance
        formset: ItemTransactionDetailsTransferFormSet instance
        user_profile: UserProfile of the current user

    Returns:
        list: List of error messages, empty if validation passes
    """
    errors = []
    castody_type = form.cleaned_data.get("castody_type")

    try:
        if castody_type == ItemTransactions.CASTODY_TYPES.Warehouse:
            # Get FROM sub-warehouse
            from_sub_warehouse = form.cleaned_data.get("from_sub_warehouse")
            if not from_sub_warehouse:
                errors.append("يجب تحديد المخزن الفرعي المرسل.")
                return errors

            # Validate each item in formset
            for form_detail in formset:
                if not form_detail.cleaned_data or form_detail.cleaned_data.get(
                    "DELETE"
                ):
                    continue

                item = form_detail.cleaned_data.get("item")
                approved_qty = form_detail.cleaned_data.get("approved_quantity", 0)

                if approved_qty <= 0 or not item:
                    continue

                # Get current stock in FROM sub-warehouse
                _fis = FacultyItemStock.objects.filter(
                    item=item, sub_warehouse=from_sub_warehouse
                ).first()
                current_stock = _fis.cached_quantity if _fis else 0

                # Validate stock availability
                if current_stock <= 0:
                    errors.append(
                        f"الصنف '{item.name}' (#{item.id}) نافد في المخزن الفرعي "
                        f"'{from_sub_warehouse.name}'. الكمية المتوفرة: 0"
                    )
                elif approved_qty > current_stock:
                    errors.append(
                        f"الكمية المطلوبة ({approved_qty}) من '{item.name}' (#{item.id}) "
                        f"تتجاوز الكمية المتوفرة ({current_stock}) في المخزن الفرعي "
                        f"'{from_sub_warehouse.name}'."
                    )

        elif castody_type in [
            ItemTransactions.CASTODY_TYPES.Personal,
            ItemTransactions.CASTODY_TYPES.Branch,
        ]:
            # Get FROM user
            from_user = form.cleaned_data.get("from_user")
            if not from_user:
                errors.append("يجب تحديد الموظف المرسل.")
                return errors

            # Validate each item in formset
            for form_detail in formset:
                if not form_detail.cleaned_data or form_detail.cleaned_data.get(
                    "DELETE"
                ):
                    continue

                item = form_detail.cleaned_data.get("item")
                approved_qty = form_detail.cleaned_data.get("approved_quantity", 0)

                if approved_qty <= 0 or not item:
                    continue

                # Get current quantity user owns
                user_qty = item.current_quantity_for_user(from_user)

                # Validate user ownership
                if user_qty <= 0:
                    errors.append(
                        f"الموظف '{from_user.get_full_name()}' (#{from_user.id}) "
                        f"لا يمتلك الصنف '{item.name}' (#{item.id})."
                    )
                elif approved_qty > user_qty:
                    errors.append(
                        f"الكمية المطلوبة ({approved_qty}) من '{item.name}' (#{item.id}) "
                        f"تتجاوز ما يمتلكه الموظف '{from_user.get_full_name()}' "
                        f"({user_qty} وحدة)."
                    )

    except Exception as e:
        # Log the error but don't expose internal details to user
        logger.error(f"Stock validation error: {str(e)}", exc_info=True)
        errors.append("حدث خطأ أثناء التحقق من الكميات. يرجى المحاولة مرة أخرى.")

    return errors


@login_required
def employee_custody_view(request):
    """Main view to select department and employee for custody display."""
    user = request.user

    if not (hasattr(user, "profile") and user.profile.faculty):
        departments = Department.objects.none()
    else:
        faculty = user.profile.faculty
        departments = Department.objects.filter(faculty=faculty)

    selected_department_id = request.GET.get("department_id")
    selected_employee_id = request.GET.get("employee_id")

    employees = User.objects.none()
    employee_custody = None
    selected_employee = None

    if selected_department_id and selected_department_id.isdigit():
        try:
            department = Department.objects.get(
                id=selected_department_id, faculty=user.profile.faculty
            )
            employees = User.objects.filter(
                profile__department=department
            ).select_related("profile")

            if selected_employee_id and selected_employee_id.isdigit():
                selected_employee = get_object_or_404(
                    User, id=selected_employee_id, profile__department=department
                )
                employee_custody = get_employee_custody_data(selected_employee)

        except Department.DoesNotExist:
            selected_department_id = None
            selected_employee_id = None

    context = {
        "departments": departments,
        "employees": employees,
        "selected_department_id": selected_department_id,
        "selected_employee_id": selected_employee_id,
        "selected_employee": selected_employee,
        "employee_custody": employee_custody,
    }
    is_htmx_request = request.headers.get("HX-Request") == "true"

    if is_htmx_request:
        if selected_employee_id and selected_department_id:
            return render(
                request, "inventory/partials/employee_custody_container.html", context
            )
        elif selected_department_id:
            return render(
                request, "inventory/partials/employee_select_for_custody.html", context
            )

    return render(request, "inventory/employee_custody.html", context)


@login_required
def department_employees_for_custody(request):
    """HTMX endpoint: Load employees when department is selected."""
    department_id = request.GET.get("department_id")
    user = request.user

    employees = User.objects.none()
    selected_department_id = None

    if department_id and department_id.isdigit():
        try:
            department = Department.objects.get(
                id=department_id, faculty=user.profile.faculty
            )
            employees = User.objects.filter(
                profile__department=department, profile__faculty=user.profile.faculty
            ).select_related("profile")
            selected_department_id = department_id
        except Department.DoesNotExist:
            pass

    context = {
        "employees": employees,
        "selected_department_id": selected_department_id,
        "selected_employee_id": None,
    }

    return render(
        request, "inventory/partials/employee_select_for_custody.html", context
    )


@login_required
def employee_custody_details(request):
    """HTMX endpoint: Load custody details when employee is selected."""
    department_id = request.GET.get("department_id")
    employee_id = request.GET.get("employee_id")
    user = request.user

    selected_employee = None
    employee_custody = None

    if (
        employee_id
        and employee_id.isdigit()
        and department_id
        and department_id.isdigit()
    ):
        try:
            department = Department.objects.get(
                id=department_id, faculty=user.profile.faculty
            )
            selected_employee = get_object_or_404(
                User, id=employee_id, profile__department=department
            )
            employee_custody = get_employee_custody_data(selected_employee)
        except Department.DoesNotExist:
            pass

    context = {
        "selected_employee": selected_employee,
        "employee_custody": employee_custody,
        "selected_department_id": department_id,
    }

    return render(
        request, "inventory/partials/employee_custody_container.html", context
    )


def get_employee_custody_data(employee, date_from=None, date_to=None, limit=None):
    """
    Show employee custody at TRANSACTION LEVEL (not aggregated per item).
    - Each original disbursement/transfer keeps its own price, document, and date.
    - Returns deduct from the exact linked original_detail if available.
    - Unlinked returns use FIFO deduction on matching items (ignores return transaction castody_type).
    - Only rows with qty > 0 are displayed.
    """
    faculty_id = getattr(getattr(employee, "profile", None), "faculty_id", None)
    if not faculty_id:
        return {
            k: {"items": [], "total_quantity": 0, "total_value": Decimal("0")}
            for k in ["warehouse", "personal", "branch"]
        }

    # Base filters: approved, not deleted, not reversed, faculty-isolated
    base_filters = Q(
        transaction__approval_status=ItemTransactions.APPROVAL_STATUS.APPROVED,
        transaction__deleted=False,
        transaction__is_reversed=False,
        transaction__faculty_id=faculty_id,
    )
    if date_from:
        base_filters &= Q(transaction__created_at__gte=date_from)
    if date_to:
        base_filters &= Q(transaction__created_at__lte=date_to)

    # Fetch all custody-affecting details in chronological order
    custody_details = (
        ItemTransactionDetails.objects.filter(
            Q(
                transaction__to_user=employee,
                transaction__transaction_type__in=["D", "T"],  # Incoming
            )
            | Q(
                transaction__from_user=employee,
                transaction__transaction_type="R",  # Return
            )
            | Q(
                transaction__from_user=employee,
                transaction__transaction_type="T",
                transaction__to_user__isnull=False,
                transaction__castody_type=ItemTransactions.CASTODY_TYPES.Personal,
            ),
            base_filters,
        )
        .select_related("item", "transaction", "transaction__from_sub_warehouse")
        .order_by("transaction__created_at", "id")
    )

    if limit:
        custody_details = custody_details[:limit]

    # ─────────────────────────────────────────────────────────────────────────
    # TRACKING MAP: { original_detail_id: {item, qty, price, doc, date, ...} }
    # ─────────────────────────────────────────────────────────────────────────
    transaction_map = {}
    castody_map = {
        ItemTransactions.CASTODY_TYPES.Warehouse: "warehouse",
        ItemTransactions.CASTODY_TYPES.Personal: "personal",
        ItemTransactions.CASTODY_TYPES.Branch: "branch",
    }

    # PASS 1: Populate map with all INCOMING transactions
    for detail in custody_details:
        tx_type = detail.transaction.transaction_type
        # Use ID comparison for reliability
        if tx_type in ["D", "T"] and detail.transaction.to_user_id == employee.id:
            transaction_map[detail.id] = {
                "item": detail.item,
                "qty": detail.approved_quantity,
                "price": detail.price or 0,
                "doc": detail.transaction.document_number or "—",
                "date": detail.transaction.created_at,
                "sub_warehouse": (
                    detail.transaction.from_sub_warehouse.name
                    if detail.transaction.from_sub_warehouse
                    else "—"
                ),
                "castody_type": detail.transaction.castody_type,
                "detail_id": detail.id,
            }

    # PASS 2: Deduct OUTGOING/RETURNS from the map
    for detail in custody_details:
        tx_type = detail.transaction.transaction_type
        castody_type = detail.transaction.castody_type
        deduct_qty = detail.approved_quantity

        if deduct_qty <= 0:
            continue

        # 1. RETURN transactions
        if tx_type == "R" and detail.transaction.from_user_id == employee.id:
            # A) Linked deduction (exact original detail)
            if (
                hasattr(detail, "original_detail_id")
                and detail.original_detail_id
                and detail.original_detail_id in transaction_map
            ):
                orig = transaction_map[detail.original_detail_id]
                deduct = min(deduct_qty, orig["qty"])
                orig["qty"] -= deduct
                # Preserve original price for audit consistency
                if detail.original_detail:
                    orig["price"] = detail.original_detail.price or orig["price"]
                deduct_qty -= deduct

            # B) Fallback: FIFO deduction for unlinked returns
            # ✅ FIX: Removed castody_type check because return form forces "Warehouse"
            #         while items might be "Personal" or "Branch".
            if deduct_qty > 0:
                for data in transaction_map.values():
                    if data["item"].id == detail.item_id and data["qty"] > 0:
                        deduct = min(deduct_qty, data["qty"])
                        data["qty"] -= deduct
                        deduct_qty -= deduct
                        if deduct_qty <= 0:
                            break

        # 2. PERSONAL TRANSFER OUT transactions
        elif (
            tx_type == "T"
            and detail.transaction.from_user_id == employee.id
            and detail.transaction.to_user_id is not None
            and castody_type == ItemTransactions.CASTODY_TYPES.Personal
        ):
            # FIFO deduction (strict castody_type match for transfers)
            if deduct_qty > 0:
                for data in transaction_map.values():
                    if (
                        data["item"].id == detail.item_id
                        and data["castody_type"] == castody_type
                        and data["qty"] > 0
                    ):
                        deduct = min(deduct_qty, data["qty"])
                        data["qty"] -= deduct
                        deduct_qty -= deduct
                        if deduct_qty <= 0:
                            break

    # ─────────────────────────────────────────────────────────────────────────
    # BUILD FINAL STRUCTURE (FILTER QTY <= 0)
    # ─────────────────────────────────────────────────────────────────────────
    custody_data = {
        k: {"items": [], "total_quantity": 0, "total_value": Decimal("0")}
        for k in ["warehouse", "personal", "branch"]
    }

    for data in transaction_map.values():
        if data["qty"] <= 0:
            continue  # Fully returned or negative, skip

        custody_key = castody_map.get(data["castody_type"], "personal")
        price_dec = Decimal(str(data["price"]))
        total_val = Decimal(str(data["qty"])) * price_dec

        custody_data[custody_key]["items"].append(
            {
                "item": data["item"],
                "quantity": data["qty"],
                "latest_price": data["price"],
                "total_value": total_val,
                "last_doc": data["doc"],
                "last_date": data["date"],
                "sub_warehouse": data["sub_warehouse"],
                "detail_id": data["detail_id"],
            }
        )
        custody_data[custody_key]["total_quantity"] += data["qty"]
        custody_data[custody_key]["total_value"] += total_val

    # Sort items chronologically within each custody type
    for key in custody_data:
        custody_data[key]["items"].sort(key=lambda x: x["last_date"] or datetime.min)
        custody_data[key]["total_value"] = float(custody_data[key]["total_value"])
        custody_data[key]["total_quantity"] = int(custody_data[key]["total_quantity"])

    return custody_data


@login_required
def export_employee_custody_pdf(request, employee_id):
    user = request.user
    employee = get_object_or_404(User, id=employee_id)

    if not (
        hasattr(user, "profile")
        and hasattr(employee, "profile")
        and user.profile.faculty == employee.profile.faculty
    ):
        return HttpResponseForbidden("ليس لديك صلاحية الوصول إلى هذه البيانات.")

    custody_data = get_employee_custody_data(employee)
    grand_total = sum(custody_data[k]["total_value"] for k in custody_data)

    html_string = render_to_string(
        "inventory/reports/employee_custody_report.html",
        {
            "employee": employee,
            "custody_data": custody_data,
            "current_date": datetime.now(),
            "user": user,
            "show_header": True,
            "grand_total": grand_total,
            "for_pdf": True,
        },
        request=request,
    )
    html = HTML(string=html_string, base_url=request.build_absolute_uri("/"))
    response = HttpResponse(content_type="application/pdf")

    # Safe filename encoding
    safe_name = quote(
        f"عهدة_{employee.first_name}_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf",
        safe="",
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{safe_name}"
    html.write_pdf(response)
    return response


@login_required
def export_department_custody_pdf(request, department_id):
    user = request.user
    department = get_object_or_404(Department, id=department_id)

    if not (hasattr(user, "profile") and user.profile.faculty == department.faculty):
        return HttpResponseForbidden("ليس لديك صلاحية الوصول إلى هذه البيانات.")

    employees = User.objects.filter(profile__department=department).select_related(
        "profile"
    )
    combined_html = []

    for emp in employees:
        custody_data = get_employee_custody_data(emp)
        emp_html = render_to_string(
            "inventory/reports/employee_custody_report.html",
            {
                "employee": emp,
                "custody_data": custody_data,
                "current_date": datetime.now(),
                "user": user,
                "show_header": True,
                "for_pdf": True,
            },
            request=request,
        )
        combined_html.append(emp_html)
        combined_html.append('<div style="page-break-after: always;"></div>')

    html = HTML(
        string="\n".join(combined_html), base_url=request.build_absolute_uri("/")
    )
    response = HttpResponse(content_type="application/pdf")

    safe_dept = quote(department.name, safe="")
    safe_name = quote(f"عهدة_قسم_{safe_dept}.pdf", safe="")
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{safe_name}"
    html.write_pdf(response)
    return response


@login_required
def export_employee_custody_excel(request, employee_id):
    user = request.user
    employee = get_object_or_404(User, id=employee_id)

    if not (
        hasattr(user, "profile")
        and hasattr(employee, "profile")
        and user.profile.faculty == employee.profile.faculty
    ):
        return HttpResponseForbidden("ليس لديك صلاحية الوصول إلى هذه البيانات.")

    custody_data = get_employee_custody_data(employee)
    return _generate_custody_excel(custody_data, employee)


def _generate_custody_excel(custody_data, employee):
    """Generate complete Excel workbook for employee custody"""
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    if custody_data["warehouse"]["items"]:
        _create_custody_sheet(wb, "عهدة مخزنية", custody_data["warehouse"], employee)
    if custody_data["personal"]["items"]:
        _create_custody_sheet(wb, "عهدة شخصية", custody_data["personal"], employee)
    if custody_data["branch"]["items"]:
        _create_custody_sheet(wb, "عهدة فرعية", custody_data["branch"], employee)

    if not wb.sheetnames:
        ws = wb.create_sheet(title="لا توجد عهدة")
        ws.cell(
            row=1,
            column=1,
            value=f"لا يوجد عهدة للموظف: {employee.get_full_name() or employee.username}",
        )
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    emp_name = employee.first_name or employee.username
    safe_name = quote(
        f"عهدة_{emp_name}_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx", safe=""
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{safe_name}"
    wb.save(response)
    return response


def _create_custody_sheet(workbook, sheet_name, custody_data, employee):
    """Create a single custody sheet with individual transaction rows."""
    # OpenPyXL sheet names max 31 chars
    safe_title = sheet_name[:31]
    ws = workbook.create_sheet(title=safe_title)
    ws.sheet_view.rightToLeft = True

    # ROW 1: Employee Name Header (MERGED A1:G1)
    employee_name = (
        f"{employee.first_name} {employee.last_name}"
        if employee.last_name
        else employee.first_name
    )
    cell = ws.cell(row=1, column=1, value=f"عهدة الموظف: {employee_name}")
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.fill = PatternFill(start_color="343a40", end_color="343a40", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells("A1:G1")

    # ROW 2: Table Headers (7 columns to match new data structure)
    headers = [
        "اسم الصنف",
        "الكمية",
        "السعر (ج.م)",
        "القيمة (ج.م)",
        "المستند",
        "التاريخ",
        "المخزن الفرعي",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="343a40", end_color="343a40", fill_type="solid"
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # ROWS 3+: Data Rows (Individual transaction details)
    items = custody_data.get("items", [])
    for row_num, item_data in enumerate(items, 3):
        # Item
        ws.cell(row=row_num, column=1, value=item_data["item"].name)
        # Quantity
        ws.cell(row=row_num, column=2, value=item_data["quantity"])

        # Price
        price = float(item_data.get("latest_price", 0) or 0)
        ws.cell(row=row_num, column=3, value=price if price > 0 else "غير متوفر")

        # Value
        total_value = float(item_data.get("total_value", 0) or 0)
        ws.cell(row=row_num, column=4, value=total_value if total_value > 0 else 0)

        # Document Number
        ws.cell(row=row_num, column=5, value=item_data.get("last_doc", "—"))

        # Date (format safely)
        last_date = item_data.get("last_date")
        ws.cell(
            row=row_num,
            column=6,
            value=last_date.strftime("%Y-%m-%d %H:%M") if last_date else "—",
        )

        # Sub-Warehouse
        ws.cell(row=row_num, column=7, value=item_data.get("sub_warehouse", "—"))

        # Styling
        if price > 0:
            ws.cell(row=row_num, column=3).fill = PatternFill(
                start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
            )
        if total_value > 0:
            ws.cell(row=row_num, column=4).fill = PatternFill(
                start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"
            )
            ws.cell(row=row_num, column=4).font = Font(bold=True, color="1B5E20")

    # FOOTER: Total Row (7 columns)
    total_row = len(items) + 3
    ws.cell(row=total_row, column=1, value="المجموع")
    ws.cell(row=total_row, column=2, value=custody_data.get("total_quantity", 0))
    ws.cell(row=total_row, column=3, value="-")
    ws.cell(row=total_row, column=4, value=float(custody_data.get("total_value", 0)))
    for col in range(5, 8):
        ws.cell(row=total_row, column=col, value="")

    for col in range(1, 8):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(
            start_color="E9ECEF", end_color="E9ECEF", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    # Auto-adjust Column Widths (A to G)
    for col_idx in range(1, 8):
        max_length = 0
        letter = get_column_letter(col_idx)
        for row_num in range(2, total_row + 1):
            val = ws.cell(row=row_num, column=col_idx).value
            if val:
                try:
                    max_length = max(max_length, len(str(val)))
                except Exception:
                    pass
        ws.column_dimensions[letter].width = min(max_length + 4, 35)


@login_required
def item_list(request):
    """List items with faculty-isolated quantities from FacultyItemStock."""

    user_faculty = getattr(getattr(request.user, "profile", None), "faculty", None)

    if not user_faculty:
        messages.error(request, "حسابك غير مرتبط بكليّة.")
        return redirect("dashboard")

    # 1. Base queryset: All global items
    items_qs = Item.objects.select_related(
        "category__sub_warehouse",
    ).prefetch_related(
        "faculty_stocks",
    )

    # 2. Current quantity: Sum FacultyItemStock for this faculty
    # Returns total cached_quantity across ALL sub_warehouses for this faculty
    current_stock_sq = (
        FacultyItemStock.objects.filter(
            item=OuterRef("pk"),
            faculty=user_faculty,
        )
        .values("item")
        .annotate(total=Sum("cached_quantity"))
        .values("total")
    )

    # 3. Pending IN: Items coming TO any sub_warehouse in this faculty
    # Only PENDING, not-deleted transactions with to_sub_warehouse set
    pending_in_sq = (
        ItemTransactionDetails.objects.filter(
            item_id=OuterRef("pk"),
            transaction__faculty=user_faculty,
            transaction__approval_status=ItemTransactions.APPROVAL_STATUS.PENDING,
            transaction__deleted=False,
            transaction__to_sub_warehouse__isnull=False,  # Must have destination
        )
        .values("item_id")
        .annotate(total=Sum("approved_quantity"))
        .values("total")
    )

    # 4. Pending OUT: Items going FROM any sub_warehouse in this faculty
    pending_out_sq = (
        ItemTransactionDetails.objects.filter(
            item_id=OuterRef("pk"),
            transaction__faculty=user_faculty,
            transaction__approval_status=ItemTransactions.APPROVAL_STATUS.PENDING,
            transaction__deleted=False,
            transaction__from_sub_warehouse__isnull=False,  # Must have source
        )
        .values("item_id")
        .annotate(total=Sum("approved_quantity"))
        .values("total")
    )

    # 5. Annotate items with quantities
    stocks = (
        items_qs.annotate(
            # Current quantity from FacultyItemStock (0 if no stock record exists)
            current_qty=Coalesce(
                Subquery(current_stock_sq), Value(0), output_field=IntegerField()
            ),
            # Pending incoming quantity
            pending_in=Coalesce(
                Subquery(pending_in_sq), Value(0), output_field=IntegerField()
            ),
            # Pending outgoing quantity
            pending_out=Coalesce(
                Subquery(pending_out_sq), Value(0), output_field=IntegerField()
            ),
        )
        .annotate(
            # Projected = current + pending_in - pending_out
            projected_quantity=ExpressionWrapper(
                F("current_qty") + F("pending_in") - F("pending_out"),
                output_field=IntegerField(),
            )
        )
        .order_by("name")
    )

    # 6. Context for template
    context = {
        "stocks": stocks,
        "is_admin": request.user.is_superuser,
        "faculty": user_faculty,
    }

    return render(request, "inventory/item_list.html", context)


@login_required
def categories_by_subwarehouse(request):
    """
    HTMX endpoint to load categories for a selected sub-warehouse.
    Returns ONLY the select element (not the full container).
    """
    try:
        # sub_warehouse = SubWarehouse.objects.get(id=sub_warehouse_id)

        categories = ItemCategory.objects.all().order_by("name")

        if not categories.exists():
            return HttpResponse(
                '<select name="category" id="id_category" class="form-select" required disabled>'
                "<option>لا توجد فئات في هذا المخزن</option>"
                "</select>"
            )

        html = """
            <select name="category" id="id_category" class="form-select" required>
                <option value="">-- اختر الفئة --</option>
        """
        for category in categories:
            html += f'<option value="{category.id}">{category.name}</option>'
        html += "</select>"

        return HttpResponse(html, content_type="text/html")
    except Exception as e:
        return HttpResponse(
            '<select name="category" id="id_category" class="form-select" required disabled>'
            f"<option>حدث خطأ: {e}</option>"
            "</select>"
        )


@login_required
def get_item_price_history(request, item_id):
    item = get_object_or_404(Item, pk=item_id)
    price_history = ItemPriceHistory.objects.filter(item=item).order_by("-date")
    print(item_id, price_history)
    context = {
        "item": item,
        "price_history": price_history,
    }
    print(context)
    return render(request, "inventory/item_price_history.html", context)


@login_required
@transaction.atomic
def item_create(request):
    """
    Create a global catalog item.
    Flow: Select sub-warehouse → Select category → Enter item details
    Automatically creates FacultyItemStock for ALL faculties with qty=0,
    using the sub_warehouse from the selected category.
    """
    from administration.models import Faculty
    from inventory.models import FacultyItemStock

    if request.method == "POST":
        form = ItemForm(request.POST, request.FILES)
        if form.is_valid():
            item = form.save(commit=False)

            if not item.code:
                item.code = f"ITEM-{uuid.uuid4().hex[:8].upper()}"

            item.created_by = request.user
            item.updated_by = request.user
            item.save()

            # Get sub_warehouse from the selected category
            target_sub_warehouse = (
                item.category.sub_warehouse if item.category else None
            )

            if target_sub_warehouse:
                all_faculties = Faculty.objects.all()
                created_count = 0

                for faculty in all_faculties:
                    # sub_warehouse MUST be in lookup because of unique_together
                    _, created = FacultyItemStock.objects.get_or_create(
                        faculty=faculty,
                        item=item,
                        sub_warehouse=target_sub_warehouse,
                        defaults={
                            "cached_quantity": 0,
                            "limit_quantity": item.limit_quantity,
                        },
                    )
                    if created:
                        created_count += 1

                messages.success(
                    request,
                    f"تم إنشاء الصنف '{item.name}' بنجاح. الكود: {item.code}. "
                    f"تمت إضافة الصنف لـ {created_count} كلية بكمية 0.",
                )
            else:
                messages.warning(
                    request,
                    f"تم إنشاء الصنف '{item.name}'، لكن الفئة المختارة لا تحتوي على مخزن فرعي.",
                )

            return redirect("item_list")
        else:
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
            for field, errors in form.errors.items():
                logger.error(f"Form error - {field}: {errors}")
    else:
        form = ItemForm()

    return render(
        request,
        "inventory/item_form.html",
        {
            "form": form,
            # Optimization: prefetch sub_warehouse for categories
            "categories": ItemCategory.objects.select_related("sub_warehouse").all(),
            "editing": False,
        },
    )


@login_required
@db_transaction.atomic
def item_edit(request, item_id):
    """
    Edit item with code validation + return_url preservation.
    Syncs FacultyItemStock.sub_warehouse when category changes.
    """
    if not hasattr(request.user, "profile") or not request.user.profile.faculty:
        return HttpResponseForbidden("ليس لديك كليّة مرتبطة بحسابك.")

    item = get_object_or_404(Item, id=item_id)

    # Get return_url from query params
    return_url = request.GET.get("return_url")

    # Validate return_url (prevent open redirect attacks)
    if return_url:
        parsed = urlparse(return_url)
        # Only allow relative URLs or same-host URLs
        if parsed.netloc and parsed.netloc != request.get_host():
            return_url = None  # Reject external redirects

    # Store old category BEFORE form processing
    old_category_id = item.category_id

    if request.method == "POST":
        form = ItemForm(
            request.POST,
            request.FILES,
            instance=item,
            user_faculty=request.user.profile.faculty,
        )
        if form.is_valid():
            item = form.save(commit=False)

            if not item.code:
                item.code = f"ITEM-{uuid.uuid4().hex[:8].upper()}"

            item.updated_at = timezone.now()
            item.updated_by = request.user
            item.save()

            # Check if category changed
            if item.category_id != old_category_id:
                new_sub_warehouse = (
                    item.category.sub_warehouse if item.category else None
                )

                if new_sub_warehouse:
                    # Update sub_warehouse for ALL FacultyItemStock records of this item
                    # updated_count = FacultyItemStock.objects.filter(item=item).update(
                    #     sub_warehouse=new_sub_warehouse
                    # )
                    messages.success(
                        request, "تم تحديث المخزن الفرعي بناءً على الفئة الجديدة."
                    )
                else:
                    messages.warning(
                        request,
                        "تم تحديث الصنف، لكن الفئة الجديدة لا تحتوي على مخزن فرعي.",
                    )

            messages.success(request, f"تم تحديث الصنف '{item.name}' بنجاح.")

            # Redirect to return_url if valid, else fallback
            if return_url:
                return redirect(return_url)
            else:
                return redirect("item_list")  # Fallback default
        else:
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
    else:
        form = ItemForm(instance=item, user_faculty=request.user.profile.faculty)

    return render(
        request,
        "inventory/item_form.html",
        {
            "form": form,
            "editing": True,
            "item": item,
            "categories": ItemCategory.objects.select_related("sub_warehouse").all(),
            "return_url": return_url,
        },
    )


@login_required
def item_list_api(request):
    """API endpoint for DataTables server-side processing"""
    if not hasattr(request.user, "profile") or not request.user.profile.faculty:
        return JsonResponse({"error": "Unauthorized"}, status=403)

    faculty = request.user.profile.faculty
    draw = int(request.GET.get("draw", 1))
    start = int(request.GET.get("start", 0))
    length = int(request.GET.get("length", 10))
    search_value = request.GET.get("search[value]", "")

    # Base queryset
    items = Item.objects.select_related("category").all()

    # Search filter
    if search_value:
        items = items.filter(
            Q(name__icontains=search_value)
            | Q(code__icontains=search_value)
            | Q(category__name__icontains=search_value)
        )

    # Total records
    total_records = Item.objects.count()
    total_filtered = items.count()

    # Annotate with quantities from FacultyItemStock
    items = items.annotate(
        current_qty=Coalesce(
            Subquery(
                FacultyItemStock.objects.filter(
                    item=OuterRef("pk"), faculty=faculty
                ).values("cached_quantity")[:1]
            ),
            Value(0),
        ),
    ).annotate(
        pending_in=Coalesce(
            Sum(
                Case(
                    When(
                        faculty_stocks__faculty=faculty,
                        faculty_stocks__itemtransactiondetails__transaction__approval_status=ItemTransactions.APPROVAL_STATUS.PENDING,
                        faculty_stocks__itemtransactiondetails__transaction__transaction_type__in=[
                            "A",
                            "R",
                            "T",
                        ],
                        then=F(
                            "faculty_stocks__itemtransactiondetails__approved_quantity"
                        ),
                    ),
                    default=Value(0),
                    output_field=IntegerField(),
                )
            ),
            Value(0),
        ),
        pending_out=Coalesce(
            Sum(
                Case(
                    When(
                        faculty_stocks__faculty=faculty,
                        faculty_stocks__itemtransactiondetails__transaction__approval_status=ItemTransactions.APPROVAL_STATUS.PENDING,
                        faculty_stocks__itemtransactiondetails__transaction__transaction_type__in=[
                            "D",
                            "T",
                        ],
                        then=F(
                            "faculty_stocks__itemtransactiondetails__approved_quantity"
                        ),
                    ),
                    default=Value(0),
                    output_field=IntegerField(),
                )
            ),
            Value(0),
        ),
    )

    # Ordering
    order_column = request.GET.get("order[0][column]", "2")
    order_dir = request.GET.get("order[0][dir]", "asc")
    order_mapping = {
        "0": "id",
        "1": "code",
        "2": "name",
        "3": "category__name",
        "4": "current_qty",
        "7": "limit_quantity",
    }
    order_field = order_mapping.get(order_column, "name")
    if order_dir == "desc":
        order_field = f"-{order_field}"

    items = items.order_by(order_field)[start : start + length]

    # Build response

    data = []
    for idx, item in enumerate(items, start=start + 1):
        current_qty = item.current_qty or 0
        limit_qty = item.limit_quantity or 0
        pending_in = item.pending_in or 0
        pending_out = item.pending_out or 0
        projected = current_qty + pending_in - pending_out

        encoded_return_url = quote(request.get_full_path())
        item_edit_url = reverse("item_edit", args=[item.id])
        item_history_url = reverse("item_history", args=[item.id])
        delete_btn = (
            '<a href="#" class="btn btn-sm btn-danger">حذف</a>'
            if request.user.profile.is_inventory_manager
            else ""
        )
        actions_html = f"<a href='{item_edit_url}?return_url={encoded_return_url}' class='btn btn-sm btn-warning'>تعديل</a>{delete_btn}<a href='{item_history_url}' class='btn btn-sm btn-info'>تاريخ الصنف</a>"

        data.append(
            {
                "id": idx,
                "code": item.code or "-",
                "name": item.name,
                "category": item.category.name if item.category else "بدون فئة",
                "current_qty": current_qty,
                "pending_in": pending_in,
                "pending_out": pending_out,
                "projected_quantity": projected,
                "limit_quantity": limit_qty,
                "actions": actions_html,
            }
        )

    return JsonResponse(
        {
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": total_filtered,
            "data": data,
        }
    )


@login_required
def check_item_code_availability(request):
    """AJAX endpoint to check if item code is available"""

    if request.method == "POST":
        code = request.POST.get("code", "").strip()
        item_id = request.POST.get("item_id")

        if not code:
            return JsonResponse({"available": False, "error": "الكود مطلوب"})

        # Check availability
        queryset = Item.objects.filter(code=code)
        if item_id:
            queryset = queryset.exclude(pk=item_id)

        available = not queryset.exists()

        return JsonResponse({"available": available})

    return JsonResponse({"error": "Invalid request"}, status=400)


@login_required
def item_delete(request, item_id):
    """Delete item (only if it belongs to user's faculty)"""
    item = get_object_or_404(Item, id=item_id)

    # Items are global — only inventory_manager or administration_manager can delete
    user_type = request.session.get("user_type", "")
    if not (
        request.user.profile.is_inventory_manager
        or user_type == "administration_manager"
    ):
        return HttpResponseForbidden("ليس لديك صلاحية حذف الأصناف.")

    if request.method == "POST":
        item_name = item.name
        item.delete()
        messages.success(request, f"تم حذف الصنف '{item_name}' بنجاح.")
        return redirect("item_list")

    return render(request, "inventory/item_confirm_delete.html", {"item": item})


def warehouse_list(request):
    warehouses = Warehouse.objects.all()
    return render(request, "inventory/warehouse/list.html", {"warehouses": warehouses})


@login_required
def warehouse_create(request):
    if not request.user.is_superuser:
        return HttpResponseForbidden("ليس لديك صلاحية إضافة مخازن.")
    if request.method == "POST":
        form = WarehouseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "تم إضافة المخزن بنجاح.")
            return redirect("inventory:warehouse_list")
    else:
        form = WarehouseForm()
    return render(
        request, "inventory/warehouse/form.html", {"form": form, "title": "إضافة مخزن"}
    )


@login_required
def warehouse_update(request, pk):
    if not request.user.is_superuser:
        return HttpResponseForbidden("ليس لديك صلاحية تعديل المخازن.")

    warehouse = get_object_or_404(Warehouse, pk=pk)
    if request.method == "POST":
        form = WarehouseForm(request.POST, instance=warehouse)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تحديث المخزن بنجاح.")
            return redirect("inventory:warehouse_list")
    else:
        form = WarehouseForm(instance=warehouse)
    return render(
        request,
        "inventory/warehouse/form.html",
        {"form": form, "title": "تعديل المخزن"},
    )


@login_required
def warehouse_delete(request, pk):
    if not request.user.is_superuser:
        return HttpResponseForbidden("ليس لديك صلاحية حذف المخازن.")
    warehouse = get_object_or_404(Warehouse, pk=pk)
    if request.method == "POST":
        warehouse.delete()
        messages.success(request, "تم حذف المخزن بنجاح.")
        return redirect("inventory:warehouse_list")
    return render(
        request, "inventory/warehouse/confirm_delete.html", {"object": warehouse}
    )


@login_required
def subwarehouse_list(request):
    """List global shared sub-warehouses."""
    sub_warehouses = SubWarehouse.objects.select_related("warehouse").order_by("name")
    return render(
        request, "inventory/subwarehouse/list.html", {"sub_warehouses": sub_warehouses}
    )


@login_required
def subwarehouse_create(request):
    """Create a new shared sub-warehouse."""
    if not request.user.is_superuser:
        return HttpResponseForbidden("ليس لديك صلاحية إضافة مخازن فرعية.")

    if request.method == "POST":
        form = SubWarehouseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "تم إنشاء المخزن الفرعي بنجاح.")
            return redirect("subwarehouse_list")
    else:
        form = SubWarehouseForm()
    return render(
        request,
        "inventory/subwarehouse/form.html",
        {"form": form, "title": "إضافة مخزن فرعي"},
    )


@login_required
def subwarehouse_update(request, pk):
    """Update a shared sub-warehouse."""
    if not request.user.is_superuser:
        return HttpResponseForbidden("ليس لديك صلاحية تعديل المخازن الفرعية.")

    sub_warehouse = get_object_or_404(SubWarehouse, pk=pk)

    if request.method == "POST":
        form = SubWarehouseForm(request.POST, instance=sub_warehouse)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تحديث المخزن الفرعي بنجاح.")
            return redirect("subwarehouse_list")
    else:
        form = SubWarehouseForm(instance=sub_warehouse)
    return render(
        request,
        "inventory/subwarehouse/form.html",
        {"form": form, "title": "تعديل المخزن الفرعي"},
    )


@login_required
def subwarehouse_delete(request, pk):
    """Delete a shared sub-warehouse."""
    if not request.user.is_superuser:
        return HttpResponseForbidden("ليس لديك صلاحية حذف المخازن الفرعية.")

    sub_warehouse = get_object_or_404(SubWarehouse, pk=pk)

    if not (
        request.user.profile.is_inventory_manager
        or request.user.profile.is_faculty_manager
        or request.session.get("user_type") == "administration_manager"
    ):
        return HttpResponseForbidden("ليس لديك صلاحية حذف المخازن الفرعية.")

    if request.method == "POST":
        sub_warehouse_name = sub_warehouse.name
        sub_warehouse.delete()
        messages.success(request, f"تم حذف المخزن الفرعي '{sub_warehouse_name}' بنجاح.")
        return redirect("subwarehouse_list")
    return render(
        request, "inventory/subwarehouse/confirm_delete.html", {"object": sub_warehouse}
    )


@login_required
def itemcategory_list(request):
    """List categories for user's faculty"""
    # Categories are now global — show all of them
    categories = (
        ItemCategory.objects.all().order_by("name").select_related("sub_warehouse")
    )
    return render(
        request, "inventory/itemcategory/list.html", {"categories": categories}
    )


@login_required
def itemcategory_create(request):
    """Create category for user's faculty"""
    if not request.user.is_superuser:
        return HttpResponseForbidden("ليس لديك صلاحية إضافة فئات الأصناف.")

    if request.method == "POST":
        form = ItemCategoryForm(request.POST)
        if form.is_valid():
            form.save()  # No need to capture return value if unused
            messages.success(request, "تم إضافة فئة الصنف بنجاح.")
            return redirect("itemcategory_list")
        else:
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
    else:
        form = ItemCategoryForm()
    return render(
        request,
        "inventory/itemcategory/form.html",
        {"form": form, "title": "إضافة فئة صنف"},
    )


@login_required
def itemcategory_update(request, pk):
    """Update category and cascade sub_warehouse changes to related transactions."""
    if not request.user.is_superuser:
        return HttpResponseForbidden("ليس لديك صلاحية تعديل فئات الأصناف.")

    category = get_object_or_404(ItemCategory, pk=pk)

    # Capture old value before POST
    old_sub_warehouse = category.sub_warehouse

    if request.method == "POST":
        form = ItemCategoryForm(request.POST, instance=category)
        if form.is_valid():
            new_sub_warehouse = form.cleaned_data.get("sub_warehouse")

            with db_transaction.atomic():
                # 1. Save the category update
                category = form.save()

                # 2. Cascade update ONLY if sub_warehouse actually changed
                if (
                    old_sub_warehouse != new_sub_warehouse
                    and new_sub_warehouse is not None
                ):
                    # Find all items belonging to this category
                    item_ids = Item.objects.filter(category=category).values_list(
                        "id", flat=True
                    )

                    # Find all transaction IDs that involve these items
                    trans_ids = ItemTransactionDetails.objects.filter(
                        item_id__in=item_ids
                    ).values_list("transaction_id", flat=True)

                    # Update ONLY pending transactions
                    count_from = ItemTransactions.objects.filter(
                        id__in=trans_ids,
                        approval_status=ItemTransactions.APPROVAL_STATUS.PENDING,
                        from_sub_warehouse=old_sub_warehouse,
                    ).update(from_sub_warehouse=new_sub_warehouse)

                    # Bulk update to_sub_warehouse
                    count_to = ItemTransactions.objects.filter(
                        id__in=trans_ids,
                        to_sub_warehouse=old_sub_warehouse,
                    ).update(to_sub_warehouse=new_sub_warehouse)

                    total_updated = count_from + count_to
                    if total_updated > 0:
                        messages.success(
                            request,
                            f"تم تحديث الفئة بنجاح. تم تعديل المخزن في {total_updated} سند مرتبط.",
                        )
                        logger.info(
                            f"Category '{category.name}' sub_warehouse changed. "
                            f"Updated {total_updated} transactions."
                        )
                    else:
                        messages.success(request, "تم تحديث فئة الصنف بنجاح.")
                else:
                    messages.success(request, "تم تحديث فئة الصنف بنجاح.")

            return redirect("itemcategory_list")
        else:
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
    else:
        form = ItemCategoryForm(instance=category)

    return render(
        request,
        "inventory/itemcategory/form.html",
        {"form": form, "title": "تعديل فئة الصنف"},
    )


@login_required
def itemcategory_delete(request, pk):
    """Delete category (only if it belongs to user's faculty)"""
    if not request.user.is_superuser:
        return HttpResponseForbidden("ليس لديك صلاحية حذف فئات الأصناف.")

    category = get_object_or_404(ItemCategory, pk=pk)

    if request.method == "POST":
        category.delete()
        messages.success(request, "تم حذف فئة الصنف بنجاح.")
        return redirect("itemcategory_list")
    return render(
        request, "inventory/itemcategory/confirm_delete.html", {"object": category}
    )


@login_required
def from_sub_warehouse_users_select_transfer(request):
    """Populate from_user based on selected FROM SUB-WAREHOUSE (for Transfer)."""
    sub_warehouse_id = request.GET.get("from_sub_warehouse") or request.GET.get("id")

    if sub_warehouse_id and sub_warehouse_id.isdigit():
        try:
            SubWarehouse.objects.get(id=sub_warehouse_id)
            users = get_inventory_users_for_sub_warehouse(
                sub_warehouse_id,
                faculty=request.user.profile.faculty,
            )
        except (SubWarehouse.DoesNotExist, AttributeError):
            users = User.objects.none()
    else:
        users = User.objects.none()

    return render(
        request,
        "inventory/partials/users_select.html",
        {"users": users, "field_name": "from_user", "empty_text": "لا يوجد موظفي مخزن"},
    )


@login_required
def to_sub_warehouse_users_select_transfer(request):
    """Populate to_user based on selected TO SUB-WAREHOUSE (for Transfer)."""
    sub_warehouse_id = request.GET.get("to_sub_warehouse") or request.GET.get("id")

    if sub_warehouse_id and sub_warehouse_id.isdigit():
        try:
            SubWarehouse.objects.get(id=sub_warehouse_id)
            users = get_inventory_users_for_sub_warehouse(
                sub_warehouse_id,
                faculty=request.user.profile.faculty,
            )
        except (SubWarehouse.DoesNotExist, AttributeError):
            users = User.objects.none()
    else:
        users = User.objects.none()

    return render(
        request,
        "inventory/partials/users_select.html",
        {"users": users, "field_name": "to_user", "empty_text": "لا يوجد موظفي مخزن"},
    )


@login_required
def from_department_users_select_transfer(request):
    """Populate from_user based on selected FROM DEPARTMENT (for Transfer)."""
    department_id = request.GET.get("from_department") or request.GET.get("id")

    if department_id and department_id.isdigit():
        try:
            department = Department.objects.get(
                id=department_id, faculty=request.user.profile.faculty
            )
            users = User.objects.filter(
                profile__department=department,
                profile__faculty=request.user.profile.faculty,
            ).select_related("profile")
        except (Department.DoesNotExist, AttributeError):
            users = User.objects.none()
    else:
        users = User.objects.none()

    return render(
        request,
        "inventory/partials/users_select.html",
        {
            "users": users,
            "field_name": "from_user",
            "empty_text": "لا يوجد موظفين في هذا القسم",
        },
    )


@login_required
def to_department_users_select_transfer(request):
    """Populate to_user based on selected TO DEPARTMENT (for Transfer)."""
    department_id = request.GET.get("to_department") or request.GET.get("id")

    if department_id and department_id.isdigit():
        try:
            department = Department.objects.get(
                id=department_id, faculty=request.user.profile.faculty
            )
            users = User.objects.filter(
                profile__department=department,
                profile__faculty=request.user.profile.faculty,
            ).select_related("profile")
        except (Department.DoesNotExist, AttributeError):
            users = User.objects.none()
    else:
        users = User.objects.none()

    return render(
        request,
        "inventory/partials/users_select.html",
        {
            "users": users,
            "field_name": "to_user",
            "empty_text": "لا يوجد موظفين في هذا القسم",
        },
    )


@login_required
def item_search_ajax(request):
    """
    Search items with stock status and LATEST PRICE (for DISBURSEMENT).
    FILTERS: Items whose category belongs to the selected sub_warehouse.
    """
    try:
        faculty = getattr(getattr(request.user, "profile", None), "faculty", None)
        if not faculty:
            return JsonResponse({"results": []})

        sub_warehouse_id = request.GET.get("warehouse")
        if not (sub_warehouse_id and sub_warehouse_id.isdigit()):
            return JsonResponse({"results": []})

        query = request.GET.get("q", "").strip()

        from inventory.models import FacultyItemStock, ItemCategory

        # Get category IDs that belong to this sub_warehouse
        # ItemCategory.sub_warehouse is ForeignKey (not M2M)
        valid_category_ids = ItemCategory.objects.filter(
            sub_warehouse_id=sub_warehouse_id
        ).values_list("id", flat=True)

        if not valid_category_ids:
            # No categories linked to this sub_warehouse → no items to show
            return JsonResponse({"results": []})

        # Efficient subquery for faculty-specific stock quantity
        faculty_stock_sq = FacultyItemStock.objects.filter(
            item=OuterRef("pk"),
            sub_warehouse_id=sub_warehouse_id,
            faculty=faculty,
            cached_quantity__gt=0,
        ).values("cached_quantity")[:1]

        # Filter items: must have stock AND category belongs to sub_warehouse
        items_qs = (
            Item.objects.select_related("category")
            .filter(
                category_id__in=valid_category_ids,
                faculty_stocks__sub_warehouse_id=sub_warehouse_id,
                faculty_stocks__faculty=faculty,
                faculty_stocks__cached_quantity__gt=0,
            )
            .distinct()
        )

        # Apply search filters
        if query:
            if query.isdigit():
                items_qs = items_qs.filter(id=int(query))
            else:
                items_qs = items_qs.filter(
                    Q(name__icontains=query) | Q(category__name__icontains=query)
                )

        # Annotate with faculty-specific quantity and latest price
        latest_price_subquery = (
            ItemPriceHistory.objects.filter(item=OuterRef("pk"))
            .order_by("-date")
            .values("price")[:1]
        )
        items_qs = items_qs.annotate(
            latest_price=Subquery(latest_price_subquery),
            faculty_qty=Coalesce(Subquery(faculty_stock_sq), Value(0)),
        )

        # BUILD RESULTS WITH PRICE INCLUDED
        results = []
        for item in items_qs[:15]:  # Limit to 15 results for performance
            price_value = None
            if item.latest_price is not None:
                try:
                    price_value = float(item.latest_price)
                except (TypeError, ValueError):
                    price_value = None

            results.append(
                {
                    "id": item.id,
                    "name": item.name,
                    "full_name": f"{item.name} (#{item.id})"
                    + (f" - {item.category.name}" if item.category else ""),
                    "quantity": item.faculty_qty,
                    "category": item.category.name if item.category else "بدون فئة",
                    "unit": item.get_unit_display(),
                    "price": price_value,
                    "limit_quantity": item.limit_quantity,
                    "is_low_stock": item.faculty_qty <= item.limit_quantity,
                }
            )

        return JsonResponse({"results": results})

    except Exception as e:
        print(f"Item search error: {str(e)}")
        import traceback

        traceback.print_exc()
        return JsonResponse({"results": [], "error": str(e)}, status=500)


@login_required
def item_search_addition(request):
    """
    Search items without stock filtering (for ADDITION).
    FILTERS: Items whose category belongs to the selected sub_warehouse.
    """
    faculty = getattr(getattr(request.user, "profile", None), "faculty", None)
    sub_warehouse_id = request.GET.get("warehouse")

    # Validate inputs
    if not faculty or not (sub_warehouse_id and sub_warehouse_id.isdigit()):
        return JsonResponse({"results": []})

    try:
        SubWarehouse.objects.get(id=sub_warehouse_id)
    except SubWarehouse.DoesNotExist:
        return JsonResponse({"results": []})

    query = request.GET.get("q", "").strip()

    from inventory.models import ItemCategory

    # Get category IDs that belong to this sub_warehouse
    valid_category_ids = ItemCategory.objects.filter(
        sub_warehouse_id=sub_warehouse_id
    ).values_list("id", flat=True)

    if not valid_category_ids:
        return JsonResponse({"results": []})

    # Global items, filtered by category belonging to sub_warehouse
    items_qs = Item.objects.select_related("category").filter(
        category_id__in=valid_category_ids
    )

    if query:
        if query.isdigit():
            items_qs = items_qs.filter(id=int(query))
        else:
            items_qs = items_qs.filter(
                Q(name__icontains=query) | Q(category__name__icontains=query)
            )

    items = items_qs[:20]

    results = [
        {
            "id": item.id,
            "name": f"{item.name} (#{item.id})"
            + (f" - {item.category.name}" if item.category else ""),
            "quantity": 0,  # Addition doesn't show current stock
            "category": item.category.name if item.category else "بدون فئة",
            "unit": item.get_unit_display(),
        }
        for item in items
    ]
    return JsonResponse({"results": results})


@login_required
def item_search_return(request):
    """
    Search items that:
    1. The selected employee has in custody (net positive quantity)
    2. Can be returned to the selected destination sub-warehouse
    """
    query = request.GET.get("q", "").strip()
    from_user_id = request.GET.get("from_user")
    to_sub_warehouse_id = request.GET.get("to_sub_warehouse")

    # Validate required params
    if not query or len(query) < 2:
        return JsonResponse({"results": []})

    if not from_user_id or not to_sub_warehouse_id:
        return JsonResponse(
            {"error": "يرجى اختيار الموظف المرسل والمخزن المستقبل أولاً"}, status=400
        )

    # Faculty isolation: get faculty from from_user's profile
    from_user = get_object_or_404(User, id=from_user_id)
    faculty_id = getattr(getattr(from_user, "profile", None), "faculty_id", None)
    if not faculty_id:
        return JsonResponse({"error": "المستخدم ليس له كلية مرتبطة"}, status=400)

    # Base item search (name or code)
    items = (
        Item.objects.filter(Q(name__icontains=query) | Q(code__icontains=query))
        .select_related("category")
        .only("id", "name", "code", "category__name", "unit", "limit_quantity")[:50]
    )  # Limit for performance

    results = []
    for item in items:
        # ✅ 1. Check if employee has this item in custody (net positive)
        employee_qty = item.current_quantity_for_user(from_user)
        if employee_qty <= 0:
            continue  # Skip items employee doesn't own

        # ✅ 2. Check destination warehouse stock (optional business rule)
        # For returns, we typically allow any item the employee owns,
        # but you can enforce that the warehouse must already track it:
        warehouse_stock = FacultyItemStock.objects.filter(
            item=item, sub_warehouse_id=to_sub_warehouse_id, faculty_id=faculty_id
        ).first()

        # Optional: Skip if warehouse doesn't track this item
        # (Remove this check to allow returns of new items to warehouse)
        # if not warehouse_stock:
        #     continue

        results.append(
            {
                "id": item.id,
                "name": item.name,
                "code": item.code or "-",
                "category": item.category.name if item.category else "-",
                "unit": item.get_unit_display(),
                "quantity": employee_qty,  # How much employee owns (for return)
                "warehouse_quantity": warehouse_stock.cached_quantity
                if warehouse_stock
                else 0,  # Optional context
                "is_low_stock": employee_qty <= item.limit_quantity
                and employee_qty > 0,
            }
        )

    return JsonResponse({"results": results})


@login_required
def item_search_transfer(request):
    """
    Search items for transfer transactions based on custody type.
    FILTERS: Items whose category belongs to the selected sub_warehouse.
    """
    query = request.GET.get("q", "").strip()
    warehouse_id = request.GET.get("warehouse")
    from_user_id = request.GET.get("from_user")
    faculty = getattr(getattr(request.user, "profile", None), "faculty", None)

    if not faculty:
        return JsonResponse({"results": []})

    results = []

    try:
        from inventory.models import ItemCategory

        if warehouse_id and warehouse_id.isdigit():
            # Get category IDs that belong to this sub_warehouse
            valid_category_ids = ItemCategory.objects.filter(
                sub_warehouse_id=warehouse_id
            ).values_list("id", flat=True)

            if not valid_category_ids:
                return JsonResponse({"results": []})

            # Get items with stock in this sub_warehouse AND category belongs to it
            stocks = (
                FacultyItemStock.objects.filter(
                    sub_warehouse_id=warehouse_id,
                    faculty=faculty,
                    cached_quantity__gt=0,
                    item__category_id__in=valid_category_ids,
                )
                .select_related("item__category")
                .values(
                    "item_id", "cached_quantity", "item__name", "item__category__name"
                )
            )

            for stock in stocks:
                if query and query.lower() not in stock["item__name"].lower():
                    if not (
                        stock["item__category__name"]
                        and query.lower() in stock["item__category__name"].lower()
                    ):
                        continue
                results.append(
                    {
                        "id": stock["item_id"],
                        "name": f"{stock['item__name']} (#{stock['item_id']})"
                        + (
                            f" - {stock['item__category__name']}"
                            if stock["item__category__name"]
                            else ""
                        ),
                        "quantity": stock["cached_quantity"],
                    }
                )

        elif from_user_id and from_user_id.isdigit():
            if not User.objects.filter(
                id=from_user_id, profile__faculty=faculty
            ).exists():
                return JsonResponse({"results": []})

            # Get faculty's sub_warehouses and their categories
            faculty_sub_warehouses = (
                SubWarehouse.objects.filter(item_stocks__faculty=faculty)
                .values_list("id", flat=True)
                .distinct()
            )

            valid_category_ids = ItemCategory.objects.filter(
                sub_warehouse_id__in=faculty_sub_warehouses
            ).values_list("id", flat=True)

            if not valid_category_ids:
                return JsonResponse({"results": []})

            # Calculate user-owned quantities
            in_qty = (
                ItemTransactionDetails.objects.filter(
                    transaction__to_user_id=from_user_id,
                    transaction__faculty=faculty,
                    transaction__approval_status=ItemTransactions.APPROVAL_STATUS.APPROVED,
                    transaction__deleted=False,
                    transaction__transaction_type__in=["D", "T"],
                )
                .values("item_id")
                .annotate(total=Sum("approved_quantity"))
            )

            out_qty = (
                ItemTransactionDetails.objects.filter(
                    transaction__from_user_id=from_user_id,
                    transaction__faculty=faculty,
                    transaction__approval_status=ItemTransactions.APPROVAL_STATUS.APPROVED,
                    transaction__deleted=False,
                    transaction__transaction_type__in=["T", "R"],
                )
                .values("item_id")
                .annotate(total=Sum("approved_quantity"))
            )

            in_dict = {r["item_id"]: r["total"] for r in in_qty}
            out_dict = {r["item_id"]: r["total"] for r in out_qty}
            net_quantities = {
                iid: in_dict.get(iid, 0) - out_dict.get(iid, 0)
                for iid in in_dict.keys()
                if in_dict.get(iid, 0) - out_dict.get(iid, 0) > 0
            }

            if net_quantities:
                items_qs = Item.objects.filter(
                    id__in=net_quantities.keys(),
                    category_id__in=valid_category_ids,
                ).select_related("category")

                if query:
                    items_qs = items_qs.filter(name__icontains=query)

                for item in items_qs[:20]:
                    results.append(
                        {
                            "id": item.id,
                            "name": f"{item.name} (#{item.id})"
                            + (f" - {item.category.name}" if item.category else ""),
                            "quantity": net_quantities.get(item.id, 0),
                        }
                    )

    except Exception as e:
        logger.error(f"Item search transfer error: {str(e)}", exc_info=True)

    return JsonResponse({"results": results})


@login_required
def has_permission_to_view_transaction(user, transaction):
    """
    Check if user has permission to view transaction details.
    Rules:
    - Inventory managers can view transactions from their faculty
    - Inventory employees can view their own transactions
    - Regular users can view transactions they're involved in
    """
    if not hasattr(user, "profile") or not user.profile.faculty:
        return False

    # Inventory managers can view all transactions from their faculty
    if user.profile.is_inventory_manager:
        return transaction.created_by.profile.faculty == user.profile.faculty

    # Inventory employees can view their own transactions
    if user.profile.is_inventory_employee:
        return transaction.created_by == user

    # Regular users can view transactions they're involved in
    return (
        transaction.created_by == user
        or transaction.from_user == user
        or transaction.to_user == user
    )


@login_required
def pending_transactions_list(request):
    """List all pending transactions for the current faculty."""
    user = request.user
    if not hasattr(user, "profile") or not user.profile.faculty:
        return HttpResponseForbidden("ليس لديك كليّة مرتبطة بحسابك.")

    faculty = user.profile.faculty
    open_year = InventoryYear.get_open_year()

    # Get pending transactions for this faculty
    transactions = (
        ItemTransactions.objects.filter(
            approval_status=ItemTransactions.APPROVAL_STATUS.PENDING,
            created_by__profile__faculty=faculty,
            year=open_year,
        )
        .select_related(
            "created_by", "to_sub_warehouse", "from_sub_warehouse", "approval_user"
        )
        .prefetch_related("itemtransactiondetails_set__item")
        .order_by("-created_at")
    )

    # Statistics
    total_transactions = ItemTransactions.objects.filter(
        created_by__profile__faculty=faculty,
        year=open_year,
    ).count()
    pending_count = transactions.count()
    approved_count = ItemTransactions.objects.filter(
        approval_status=ItemTransactions.APPROVAL_STATUS.APPROVED,
        created_by__profile__faculty=faculty,
        year=open_year,
    ).count()
    rejected_count = ItemTransactions.objects.filter(
        approval_status=ItemTransactions.APPROVAL_STATUS.REJECTED,
        created_by__profile__faculty=faculty,
        year=open_year,
    ).count()

    context = {
        "transactions": transactions,
        "total_transactions": total_transactions,
        "pending_count": pending_count,
        "approved_count": approved_count,
        "rejected_count": rejected_count,
        "faculty_name": faculty.name,
        "approval_status_choices": ItemTransactions.APPROVAL_STATUS.choices,
        "current_approval_status": ItemTransactions.APPROVAL_STATUS.PENDING,
        "search_query": "",
        "start_date": "",
        "end_date": "",
        "user_type": request.session.get("user_type", "user"),
        "open_year": open_year,
    }
    return render(request, "inventory/transaction_list.html", context)


@login_required
def transaction_detail_view(request, pk):
    """View transaction details with latest item prices."""
    transaction = get_object_or_404(ItemTransactions, pk=pk)

    user_faculty = (
        request.user.profile.faculty if hasattr(request.user, "profile") else None
    )
    transaction_faculty = (
        transaction.created_by.profile.faculty
        if hasattr(transaction.created_by, "profile")
        else None
    )

    if user_faculty != transaction_faculty:
        return HttpResponseForbidden("ليس لديك صلاحية عرض هذا السند.")

    details = transaction.itemtransactiondetails_set.select_related("item").all()
    items_data = []
    for detail in details:
        latest_price = None
        latest = (
            ItemPriceHistory.objects.filter(item=detail.item).order_by("-date").first()
        )

        if latest:
            latest_price = latest.price

        items_data.append(
            {
                "detail": detail,
                "latest_price": latest_price,
                "price": detail.price,
                "transaction_price": detail.price,
            }
        )

    context = {
        "transaction": transaction,
        "items": items_data,
        "user_type": request.session.get("user_type", "user"),
    }
    return render(request, "inventory/transaction_detail.html", context)


@login_required
def log_javascript_error(request):
    """Log JavaScript errors to server logs."""
    if request.method == "POST":
        try:
            error_data = json.loads(request.body)
            logger.error(f"JavaScript Error: {error_data.get('error')}")
            logger.error(f"Stack: {error_data.get('stack')}")
            logger.error(f"Page: {error_data.get('page')}")
            logger.error(f"Action: {error_data.get('action')}")
            logger.error(f"URL: {error_data.get('url')}")
            logger.error(f"Response: {error_data.get('response')}")
            return JsonResponse({"success": True})
        except Exception as e:
            logger.error(f"Failed to log JavaScript error: {str(e)}")
            return JsonResponse({"success": False}, status=500)
    return JsonResponse({"success": False}, status=405)


@login_required
def supplier_list_view(request):
    """List all suppliers with search functionality."""
    search_query = request.GET.get("q", "")

    suppliers = Supplier.objects.all().order_by("-id")

    if search_query:
        suppliers = suppliers.filter(
            Q(name__icontains=search_query)
            | Q(contact_name__icontains=search_query)
            | Q(company_phone__icontains=search_query)
            | Q(contact_phone__icontains=search_query)
        )

    context = {
        "suppliers": suppliers,
        "search_query": search_query,
        "total_count": Supplier.objects.count(),
        "filtered_count": suppliers.count(),
    }
    return render(request, "inventory/supplier/supplier_list.html", context)


@login_required
def supplier_create_view(request):
    """Create a new supplier."""
    if request.method == "POST":
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save()
            messages.success(request, f'تم إضافة المورد "{supplier.name}" بنجاح')
            return redirect("supplier_list")
        else:
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج")
    else:
        form = SupplierForm()

    context = {
        "form": form,
        "page_title": "إضافة مورد جديد",
        "submit_text": "إضافة مورد",
    }
    return render(request, "inventory/supplier/supplier_form.html", context)


@login_required
def supplier_update_view(request, pk):
    """Update an existing supplier."""
    supplier = get_object_or_404(Supplier, pk=pk)

    if request.method == "POST":
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            supplier = form.save()
            messages.success(request, f'تم تحديث المورد "{supplier.name}" بنجاح')
            return redirect("supplier_list")
        else:
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج")
    else:
        form = SupplierForm(instance=supplier)

    context = {
        "form": form,
        "page_title": "تعديل مورد",
        "submit_text": "تحديث المورد",
        "supplier": supplier,
    }
    return render(request, "inventory/supplier/supplier_form.html", context)


@login_required
def supplier_delete_view(request, pk):
    """Delete a supplier with confirmation."""
    supplier = get_object_or_404(Supplier, pk=pk)

    if request.method == "POST":
        supplier_name = supplier.name
        supplier.delete()
        messages.success(request, f'تم حذف المورد "{supplier_name}" بنجاح')
        return redirect("supplier_list")

    context = {
        "supplier": supplier,
        "item_count": 0,  # You can add logic to check if this supplier is referenced by transactions
    }
    return render(request, "inventory/supplier/supplier_confirm_delete.html", context)


@login_required
def supplier_detail_view(request, pk):
    """View details of a specific supplier."""
    supplier = get_object_or_404(Supplier, pk=pk)

    # You can add related data here if needed (e.g., transactions with this supplier)
    context = {
        "supplier": supplier,
    }
    return render(request, "inventory/supplier/supplier_detail.html", context)


@login_required
def admin_all_items_view(request):
    """
    Administration manager cross-faculty inventory view.
    Shows all FacultyItemStock records with filtering by faculty/warehouse/category.
    """
    user_type = request.session.get("user_type")
    if user_type != "administration_manager":
        return HttpResponseForbidden("هذه الصفحة مخصصة لمدير الإدارة فقط.")

    from administration.models import Faculty as FacultyModel

    # Filters
    faculty_id = request.GET.get("faculty_id")
    warehouse_id = request.GET.get("warehouse_id")
    category_id = request.GET.get("category_id")
    search_q = request.GET.get("q", "").strip()

    faculties_qs = FacultyModel.objects.all().order_by("name")
    if faculty_id:
        faculties_qs = faculties_qs.filter(id=faculty_id)

    stocks, _, _ = _build_admin_inventory_rows(
        faculties=faculties_qs,
        sub_warehouse_id=warehouse_id,
        category_id=category_id,
        search_q=search_q,
    )

    context = {
        "stocks": stocks,
        "faculties": FacultyModel.objects.all().order_by("name"),
        "sub_warehouses": SubWarehouse.objects.select_related("warehouse").order_by(
            "name"
        ),
        "categories": ItemCategory.objects.all().order_by("name"),
        "selected_faculty_id": faculty_id,
        "selected_warehouse_id": warehouse_id,
        "selected_category_id": category_id,
        "search_q": search_q,
    }
    return render(request, "inventory/admin_all_items.html", context)


@login_required
def transaction_reverse_view(request, transaction_id):
    """Create reversal transaction for approved transaction - with proper quantity recalculation."""
    transaction = get_object_or_404(ItemTransactions, id=transaction_id, deleted=False)
    user = request.user

    # Check if already reversed
    if transaction.is_reversed:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "error": f"لا يمكن عكس السند رقم {transaction.document_number} لأنه تم عكسه مسبقاً.",
                },
                status=400,
            )
        messages.warning(
            request,
            f"لا يمكن عكس السند رقم {transaction.document_number} لأنه تم عكسه مسبقاً.",
        )
        return redirect("transaction_list")

    # Faculty isolation check
    if not hasattr(user, "profile") or not user.profile.faculty:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {"success": False, "error": "حسابك غير مرتبط بكليّة."}, status=403
            )
        messages.error(request, "حسابك غير مرتبط بكليّة.")
        return redirect("transaction_list")

    if transaction.created_by.profile.faculty != user.profile.faculty:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "error": "ليس لديك صلاحية عكس هذا السند - الكليات لا تتطابق.",
                },
                status=403,
            )
        messages.error(request, "ليس لديك صلاحية عكس هذا السند - الكليات لا تتطابق.")
        return redirect("transaction_list")

    # Only approved transactions can be reversed
    if transaction.approval_status != ItemTransactions.APPROVAL_STATUS.APPROVED:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {"success": False, "error": "يمكن فقط عكس السندات المعتمدة."},
                status=400,
            )
        messages.warning(request, "يمكن فقط عكس السندات المعتمدة.")
        return redirect("transaction_list")

    # Block Addition transactions from being reversed
    if transaction.transaction_type == ItemTransactions.TRANSACTION_TYPES.Addition:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "error": f"لا يمكن عكس سندات الإضافة. السند رقم {transaction.document_number} هو سند إضافة.",
                },
                status=400,
            )
        messages.warning(
            request,
            f"لا يمكن عكس سندات الإضافة. السند رقم {transaction.document_number} هو سند إضافة.",
        )
        return redirect("transaction_list")

    if request.method == "POST":
        reverse_reason = request.POST.get("reverse_reason", "").strip()

        try:
            with db_transaction.atomic():
                # Double-check not reversed (race condition protection)
                transaction.refresh_from_db()
                if transaction.is_reversed:
                    raise ValueError("تم عكس هذا السند مسبقاً")

                # Get affected item IDs BEFORE creating reversal
                affected_item_ids = list(
                    transaction.itemtransactiondetails_set.values_list(
                        "item_id", flat=True
                    )
                )

                # Create reversal transaction (swapping from/to fields)
                reversal = ItemTransactions(
                    transaction_type=transaction.transaction_type,
                    castody_type=transaction.castody_type,
                    document_type=transaction.document_type,
                    # Swap from/to fields
                    from_warehouse=transaction.to_warehouse,
                    to_warehouse=transaction.from_warehouse,
                    from_sub_warehouse=transaction.to_sub_warehouse,
                    to_sub_warehouse=transaction.from_sub_warehouse,
                    from_department=transaction.to_department,
                    to_department=transaction.from_department,
                    from_user=transaction.to_user,
                    to_user=transaction.from_user,
                    inventory_user=transaction.inventory_user,
                    notes=f"عكس للسند {transaction.document_number}: {reverse_reason}",
                    approval_status=ItemTransactions.APPROVAL_STATUS.PENDING,
                    created_by=user,
                )

                # Save to auto-generate document_number
                reversal.save()
                reversal.document_number = f"REV-{reversal.document_number}"
                reversal.save(update_fields=["document_number"])

                # Copy details with same quantities
                for detail in transaction.itemtransactiondetails_set.all():
                    ItemTransactionDetails.objects.create(
                        transaction=reversal,
                        item=detail.item,
                        order_quantity=detail.approved_quantity,
                        approved_quantity=detail.approved_quantity,
                        status=detail.status,
                        price=detail.price,
                    )

                # Mark original transaction as reversed
                transaction.is_reversed = True
                transaction.reversed_by = user
                transaction.reversed_at = timezone.now()
                transaction.reversed_transaction = reversal
                transaction.notes = f"{transaction.notes or ''} | تم العكس بالسند {reversal.document_number}"
                transaction.save(
                    update_fields=[
                        "is_reversed",
                        "reversed_by",
                        "reversed_at",
                        "reversed_transaction",
                        "notes",
                    ]
                )

                # Log reversal action
                log_transaction_action(
                    transaction=reversal,
                    action=TransactionAuditLog.ACTION_TYPES.CREATE,
                    user=user,
                    request=request,
                    details=f"تم إنشاء سند عكس {reversal.document_number} للسند {transaction.document_number}: {reverse_reason}",
                )

                # Set reversal as APPROVED first
                reversal.approval_status = ItemTransactions.APPROVAL_STATUS.APPROVED
                reversal.approval_user = user
                reversal.approval_date = timezone.now()
                reversal.approval_notes = (
                    f"عكس تلقائي للسند {transaction.document_number}"
                )
                reversal.save()

                # Recalculate FacultyItemStock for affected items
                if affected_item_ids:
                    from inventory.models import FacultyItemStock
                    from inventory.models import SubWarehouse as SW

                    affected_sub_ids = set()
                    if reversal.from_sub_warehouse_id:
                        affected_sub_ids.add(reversal.from_sub_warehouse_id)
                    if reversal.to_sub_warehouse_id:
                        affected_sub_ids.add(reversal.to_sub_warehouse_id)
                    if transaction.from_sub_warehouse_id:
                        affected_sub_ids.add(transaction.from_sub_warehouse_id)
                    if transaction.to_sub_warehouse_id:
                        affected_sub_ids.add(transaction.to_sub_warehouse_id)
                    if affected_sub_ids:
                        FacultyItemStock.batch_recalculate(
                            items=Item.objects.filter(id__in=affected_item_ids),
                            sub_warehouses=SW.objects.filter(id__in=affected_sub_ids),
                        )

                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse(
                        {
                            "success": True,
                            "message": f"تم إنشاء سند العكس {reversal.document_number} بنجاح. تم ضبط الكميات في المخزون.",
                            "redirect_url": reverse("transaction_list"),
                            "reversal_id": reversal.id,
                        }
                    )

                messages.success(
                    request,
                    f"تم إنشاء سند العكس {reversal.document_number} بنجاح. السند الأصلي: {transaction.document_number}",
                )
                return redirect("transaction_list")

        except Exception as e:
            logger.error(f"Error reversing transaction: {str(e)}", exc_info=True)
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse(
                    {
                        "success": False,
                        "error": f"حدث خطأ أثناء عكس السند: {str(e)}",
                    },
                    status=500,
                )
            messages.error(request, "حدث خطأ أثناء عكس السند. يرجى المحاولة مرة أخرى.")
            return redirect("transaction_list")

    return redirect("transaction_list")


@login_required
def item_search(request):
    """Search items with faculty-isolated stock quantities."""
    query = request.GET.get("q", "").strip()
    warehouse_id = request.GET.get("warehouse")
    faculty_id = request.GET.get("faculty")

    print(f"item search faculty_id: {faculty_id}")
    print(f"item search warehouse_id: {warehouse_id}")
    print(f"item search query: {query}")

    if not query or len(query) < 2:
        return JsonResponse({"results": []})

    if not faculty_id:
        # Fallback: use user's faculty if not provided
        faculty_id = getattr(getattr(request.user, "profile", None), "faculty_id", None)

    if not faculty_id:
        return JsonResponse({"error": "Faculty not specified"}, status=400)

    # Base item search
    items = (
        Item.objects.filter(Q(name__icontains=query) | Q(code__icontains=query))
        .select_related("category")
        .only("id", "name", "code", "category__name", "unit", "limit_quantity")[:50]
    )

    results = []
    for item in items:
        # Use authoritative calculation with faculty filter
        stock = FacultyItemStock.objects.filter(
            item=item,
            faculty_id=faculty_id,
        )

        if warehouse_id:
            stock = stock.filter(sub_warehouse_id=warehouse_id)

        # Get quantity using your authoritative logic
        quantity = (
            stock.aggregate(qty=Coalesce(Sum("cached_quantity"), Value(0)))["qty"] or 0
        )
        print(f"item search quantity: {quantity}")
        print(f"item search item: {item}")
        print(f"item search stock: {stock}")
        # Get latest price
        latest_price = item.itempricehistory_set.order_by("-date").first()
        price = float(latest_price.price) if latest_price else 0

        results.append(
            {
                "id": item.id,
                "name": item.name,
                "code": item.code or "-",
                "category": item.category.name if item.category else "-",
                "unit": item.get_unit_display(),
                "quantity": quantity,
                "price": price,
                "is_low_stock": quantity <= item.limit_quantity and quantity > 0,
                "is_out_of_stock": quantity <= 0,
            }
        )

    return JsonResponse({"results": results})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_transaction_list(request):
    """List ALL transactions across all faculties with filters."""
    open_year = InventoryYear.get_open_year()
    qs = (
        ItemTransactions.objects.filter(year=open_year)
        .select_related(
            "faculty",
            "created_by",
            "approval_user",
            "from_warehouse",
            "to_warehouse",
            "from_sub_warehouse",
            "to_sub_warehouse",
            "from_department",
            "to_department",
            "from_user",
            "to_user",
            "inventory_user",
            "supplier",
        )
        .prefetch_related(  # For reverse FK or M2M if needed
            "itemtransactiondetails_set__item",
            "audit_logs",
        )
        .order_by("-created_at")
    )

    # Search & Filters
    q = request.GET.get("q", "").strip()
    tx_type = request.GET.get("type")
    status = request.GET.get("status")

    if q:
        qs = qs.filter(
            Q(document_number__icontains=q)
            | Q(faculty__name__icontains=q)
            | Q(created_by__username__icontains=q)
            | Q(created_by__first_name__icontains=q)
        )
    if tx_type:
        qs = qs.filter(transaction_type=tx_type)
    if status:
        qs = qs.filter(approval_status=status)

    return render(
        request,
        "inventory/admin_transaction_list.html",
        {
            "transactions": qs,
            "tx_types": ItemTransactions.TRANSACTION_TYPES.choices,
            "statuses": ItemTransactions.APPROVAL_STATUS.choices,
            "open_year": open_year,
        },
    )


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_transaction_detail(request, transaction_id):
    """View transaction details. Shows Edit button ONLY for Disbursements."""
    tx = get_object_or_404(
        ItemTransactions.objects.select_related(
            "faculty",
            "created_by",
            "approval_user",
            "from_sub_warehouse",
            "to_sub_warehouse",
        ),
        id=transaction_id,
    )
    details = tx.itemtransactiondetails_set.select_related("item").order_by("id")

    return render(
        request,
        "inventory/admin_transaction_detail.html",
        {
            "transaction": tx,
            "details": details,
            "is_disbursement": tx.transaction_type
            == ItemTransactions.TRANSACTION_TYPES.Disbursement,
        },
    )


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_update_transaction_prices(request, transaction_id):
    """
    Superuser-only view to update ONLY item prices in a transaction.
    Bypasses audit logging, stock recalculation, and form validation.
    """
    tx = get_object_or_404(ItemTransactions, id=transaction_id, deleted=False)
    details = tx.itemtransactiondetails_set.select_related("item").order_by("id")

    if request.method == "POST":
        updated_count = 0
        with db_transaction.atomic():
            for detail in details:
                raw_value = request.POST.get(f"price_{detail.id}", "").strip()

                # Handle empty input as None
                if not raw_value:
                    new_price = None
                else:
                    try:
                        new_price = Decimal(raw_value)
                        if new_price < 0:
                            messages.warning(
                                request,
                                f"تم تجاهل السعر السالب للصنف: {detail.item.name}",
                            )
                            continue
                    except InvalidOperation:
                        messages.error(
                            request, f"قيمة غير صالحة للصنف: {detail.item.name}"
                        )
                        continue

                # Only save if value actually changed
                if detail.price != new_price:
                    detail.price = new_price
                    detail.save(update_fields=["price"])
                    updated_count += 1

        messages.success(
            request,
            f"✅ تم تحديث أسعار {updated_count} أصناف بنجاح في السند '{tx.document_number}'. "
            f"(لم يتم إنشاء إدخال في سجل التدقيق)",
        )
        return redirect("admin_transaction_detail", transaction_id=tx.id)

    return render(
        request,
        "inventory/admin_update_transaction_prices.html",
        {
            "transaction": tx,
            "details": details,
            "page_title": f"تحديث أسعار العهدة | {tx.document_number}",
        },
    )


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_faculty_stock_view(request):
    """
    Admin-only read-only view of item quantities per faculty.
    Displays: Item, Category, Sub-Warehouse, Quantity, Limit, Status.
    Strictly GET-based, no forms, no edit capabilities.
    """
    faculty_id = request.GET.get("faculty_id")
    faculties = Faculty.objects.all().order_by("name")
    selected_faculty = None
    stock_items = []

    if faculty_id and faculty_id.isdigit():
        selected_faculty = get_object_or_404(Faculty, id=faculty_id)
        # Efficient single query with all related data
        stock_items = (
            FacultyItemStock.objects.filter(faculty=selected_faculty)
            .select_related(
                "item", "item__category", "sub_warehouse", "sub_warehouse__warehouse"
            )
            .order_by("item__category__name", "sub_warehouse__name", "item__name")
        )

    return render(
        request,
        "inventory/admin_faculty_stock.html",
        {
            "faculties": faculties,
            "selected_faculty": selected_faculty,
            "stock_items": stock_items,
            "page_title": "مراقبة مخزون الكليات",
        },
    )


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_edit_custody_prices(request):
    """
    Admin-only view to select faculty → department → employee,
    then edit custody prices for EACH transaction detail row with return deduction.
    """
    faculties = Faculty.objects.all().order_by("name")
    selected_faculty_id = request.GET.get("faculty_id")
    selected_dept_id = request.GET.get("department_id")
    selected_emp_id = request.GET.get("employee_id")

    departments = Department.objects.none()
    employees = User.objects.none()
    custody_records = []
    selected_employee = None

    if selected_faculty_id and selected_faculty_id.isdigit():
        departments = Department.objects.filter(
            faculty_id=selected_faculty_id
        ).order_by("name")

    if selected_dept_id and selected_dept_id.isdigit():
        employees = (
            User.objects.filter(
                profile__department_id=selected_dept_id,
                profile__faculty_id=selected_faculty_id,
            )
            .select_related("profile")
            .order_by("first_name", "last_name")
        )

    if selected_emp_id and selected_emp_id.isdigit():
        selected_employee = get_object_or_404(
            User,
            id=selected_emp_id,
            profile__department_id=selected_dept_id,
            profile__faculty_id=selected_faculty_id,
        )

        # Fetch ALL custody-affecting details (incoming + returns)
        details_qs = (
            ItemTransactionDetails.objects.filter(
                Q(
                    transaction__to_user=selected_employee,
                    transaction__transaction_type__in=["D", "T"],
                )
                | Q(
                    transaction__from_user=selected_employee,
                    transaction__transaction_type="R",
                )
                | Q(
                    transaction__from_user=selected_employee,
                    transaction__transaction_type="T",
                    transaction__to_user__isnull=False,
                    transaction__castody_type=ItemTransactions.CASTODY_TYPES.Personal,
                ),
                transaction__approval_status=ItemTransactions.APPROVAL_STATUS.APPROVED,
                transaction__deleted=False,
                transaction__is_reversed=False,
                transaction__faculty_id=selected_faculty_id,
            )
            .select_related("item", "transaction", "transaction__from_sub_warehouse")
            .order_by("transaction__created_at", "id")
        )

        # ─────────────────────────────────────────────────────────────────────
        # BUILD TRANSACTION MAP: { original_detail_id: {qty, price, item, ...} }
        # ─────────────────────────────────────────────────────────────────────
        transaction_map = {}
        castody_map = {
            ItemTransactions.CASTODY_TYPES.Warehouse: "warehouse",
            ItemTransactions.CASTODY_TYPES.Personal: "personal",
            ItemTransactions.CASTODY_TYPES.Branch: "branch",
        }

        # PASS 1: Populate map with INCOMING transactions
        for detail in details_qs:
            tx_type = detail.transaction.transaction_type
            if (
                tx_type in ["D", "T"]
                and detail.transaction.to_user_id == selected_employee.id
            ):
                transaction_map[detail.id] = {
                    "item": detail.item,
                    "qty": detail.approved_quantity,
                    "price": detail.price or 0,
                    "doc": detail.transaction.document_number or "—",
                    "date": detail.transaction.created_at,
                    "sub_warehouse": (
                        detail.transaction.from_sub_warehouse.name
                        if detail.transaction.from_sub_warehouse
                        else "—"
                    ),
                    "castody_type": detail.transaction.castody_type,
                    "castody_type_display": detail.transaction.get_castody_type_display(),
                    "unit": detail.item.get_unit_display(),
                    "detail_id": detail.id,  # ✅ Ensure this is always an integer
                }

        # PASS 2: Deduct RETURNS from the map
        for detail in details_qs:
            tx_type = detail.transaction.transaction_type
            deduct_qty = detail.approved_quantity

            if deduct_qty <= 0:
                continue

            if (
                tx_type == "R"
                and detail.transaction.from_user_id == selected_employee.id
            ):
                # A) Linked deduction
                if (
                    hasattr(detail, "original_detail_id")
                    and detail.original_detail_id
                    and detail.original_detail_id in transaction_map
                ):
                    orig = transaction_map[detail.original_detail_id]
                    deduct = min(deduct_qty, orig["qty"])
                    orig["qty"] -= deduct
                    deduct_qty -= deduct

                # B) FIFO fallback
                if deduct_qty > 0:
                    for data in transaction_map.values():
                        if data["item"].id == detail.item_id and data["qty"] > 0:
                            deduct = min(deduct_qty, data["qty"])
                            data["qty"] -= deduct
                            deduct_qty -= deduct
                            if deduct_qty <= 0:
                                break

        # BUILD FINAL LIST: Only rows with remaining qty > 0
        for data in transaction_map.values():
            if data["qty"] <= 0:
                continue

            # ✅ Ensure detail_id is valid integer before adding to list
            if data["detail_id"] and isinstance(data["detail_id"], int):
                custody_records.append(
                    {
                        "detail_id": data["detail_id"],
                        "item_name": data["item"].name,
                        "unit": data["unit"],
                        "quantity": data["qty"],
                        "price": data["price"],
                        "castody_type": data["castody_type_display"],
                        "doc_number": data["doc"],
                        "date": data["date"],
                        "sub_warehouse": data["sub_warehouse"],
                    }
                )

        # Sort by custody type then item name
        custody_records.sort(key=lambda x: (x["castody_type"], x["item_name"]))

    # Handle POST for price updates
    if request.method == "POST" and selected_employee:
        updated_count = 0
        ignored_errors = []

        with db_transaction.atomic():
            for rec in custody_records:
                # ✅ Validate detail_id before using in query
                detail_id = rec.get("detail_id")
                if not detail_id or not isinstance(detail_id, int):
                    continue

                raw_price = request.POST.get(f"price_{detail_id}", "").strip()
                if not raw_price:
                    continue  # Skip empty inputs (user didn't change this price)

                try:
                    new_price = Decimal(raw_price)
                    if new_price < 0:
                        ignored_errors.append(
                            f"قيمة سالبة: {rec['item_name']} (سند: {rec['doc_number']})"
                        )
                        continue
                except InvalidOperation:
                    ignored_errors.append(
                        f"قيمة غير صالحة: {rec['item_name']} (سند: {rec['doc_number']})"
                    )
                    continue

                # ✅ Safe query with validated ID
                try:
                    detail = ItemTransactionDetails.objects.select_for_update().get(
                        id=detail_id
                    )
                except ItemTransactionDetails.DoesNotExist:
                    ignored_errors.append(f"الصنف غير موجود: {rec['item_name']}")
                    continue

                if detail.price != new_price:
                    detail.price = new_price
                    detail.save(update_fields=["price"])
                    updated_count += 1

        if ignored_errors:
            messages.warning(request, "⚠️ " + " | ".join(ignored_errors))
        if updated_count > 0:
            messages.success(
                request, f"✅ تم تحديث أسعار {updated_count} أصناف في العهدة بنجاح."
            )
        else:
            messages.info(request, "ℹ️ لم يتم إجراء أي تغييرات على الأسعار.")

        # Redirect to preserve selection state
        return redirect(
            f"{request.path}?faculty_id={selected_faculty_id}&department_id={selected_dept_id}&employee_id={selected_emp_id}"
        )

    return render(
        request,
        "inventory/admin_edit_custody_prices.html",
        {
            "faculties": faculties,
            "departments": departments,
            "employees": employees,
            "selected_faculty_id": selected_faculty_id,
            "selected_dept_id": selected_dept_id,
            "selected_emp_id": selected_emp_id,
            "selected_employee": selected_employee,
            "custody_records": custody_records,
        },
    )


@login_required
@user_passes_test(lambda u: u.is_superuser)
def htmx_load_departments(request):
    faculty_id = request.GET.get("faculty_id")
    departments = Department.objects.none()
    if faculty_id and faculty_id.isdigit():
        departments = Department.objects.filter(faculty_id=faculty_id).order_by("name")
    return render(
        request,
        "inventory/partials/department_dropdown.html",
        {"departments": departments},
    )


@login_required
@user_passes_test(lambda u: u.is_superuser)
def htmx_load_employees(request):
    dept_id = request.GET.get("department_id")
    employees = User.objects.none()
    if dept_id and dept_id.isdigit():
        employees = (
            User.objects.filter(profile__department_id=dept_id)
            .select_related("profile")
            .order_by("first_name", "last_name")
        )
    return render(
        request, "inventory/partials/employee_dropdown.html", {"employees": employees}
    )
    dept_id = request.GET.get("department_id")
    employees = User.objects.none()
    if dept_id and dept_id.isdigit():
        employees = (
            User.objects.filter(profile__department_id=dept_id)
            .select_related("profile")
            .order_by("first_name", "last_name")
        )
    return render(
        request, "inventory/partials/employee_dropdown.html", {"employees": employees}
    )


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_edit_transaction_header(request, transaction_id):
    """
    Admin-only view to edit ONLY transaction header data.
    Details are displayed read-only below the form.
    """
    transaction = get_object_or_404(ItemTransactions, id=transaction_id, deleted=False)
    # user_faculty = getattr(getattr(request.user, "profile", None), "faculty", None)

    # Faculty isolation check
    # if not user_faculty or transaction.faculty != user_faculty:
    #     messages.error(request, "ليس لديك صلاحية تعديل هذا السند.")
    #     return redirect("admin_transaction_list")

    # Fetch details for read-only display
    details = transaction.itemtransactiondetails_set.select_related("item").order_by(
        "id"
    )

    if request.method == "POST":
        form = ItemTransactionForm(
            request.POST,
            instance=transaction,
            user=request.user,
            user_type="admin",
            from_sub_warehouse_id=transaction.from_sub_warehouse_id,
        )

        if form.is_valid():
            try:
                with db_transaction.atomic():
                    old_data = transaction.to_dict()
                    updated_tx = form.save(commit=False)
                    updated_tx.modified_by = request.user
                    updated_tx.save()

                    log_transaction_action(
                        transaction=updated_tx,
                        action=TransactionAuditLog.ACTION_TYPES.UPDATE,
                        user=request.user,
                        request=request,
                        old_data=old_data,
                        details=f"تم تعديل رأس السند #{updated_tx.document_number} بواسطة المشرف",
                    )

                messages.success(
                    request,
                    f"✅ تم تعديل رأس السند '{updated_tx.document_number}' بنجاح.",
                )
                return redirect(
                    "admin_transaction_detail", transaction_id=updated_tx.id
                )

            except Exception as e:
                messages.error(request, f"❌ حدث خطأ أثناء الحفظ: {e}")
                logger.error(f"Admin header edit error: {e}", exc_info=True)
        else:
            messages.error(request, "يرجى تصحيح الأخطاء في النموذج.")
    else:
        form = ItemTransactionForm(
            instance=transaction,
            user=request.user,
            user_type="admin",
            from_sub_warehouse_id=transaction.from_sub_warehouse_id,
        )

    return render(
        request,
        "inventory/admin_edit_transaction.html",
        {
            "form": form,
            "transaction": transaction,
            "details": details,  # ✅ Passed for read-only display
            "page_title": f"تعديل رأس السند #{transaction.document_number}",
        },
    )


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_faculty_items_view(request):
    """Main admin view: Select faculty to view items & pending quantities."""
    faculties = Faculty.objects.all().order_by("name")
    selected_faculty_id = request.GET.get("faculty_id")
    selected_faculty = None
    if selected_faculty_id and selected_faculty_id.isdigit():
        selected_faculty = get_object_or_404(Faculty, id=selected_faculty_id)

    return render(
        request,
        "inventory/admin_faculty_items.html",
        {
            "faculties": faculties,
            "selected_faculty": selected_faculty,
            "selected_faculty_id": selected_faculty_id,
        },
    )


@login_required
@user_passes_test(lambda u: u.is_superuser)
def htmx_load_faculty_items(request):
    """HTMX endpoint: Load optimized items table with pending quantities."""
    faculty_id = request.GET.get("faculty_id")
    if not (faculty_id and faculty_id.isdigit()):
        return HttpResponse(
            "<p class='text-muted text-center py-4'>اختر كلية لعرض الأصناف</p>"
        )

    faculty = get_object_or_404(Faculty, id=faculty_id)

    # 1. Get all stock records for this faculty
    stocks = (
        FacultyItemStock.objects.filter(faculty=faculty)
        .select_related("item", "item__category", "sub_warehouse")
        .order_by("item__name")
    )

    if not stocks.exists():
        return HttpResponse(
            "<p class='text-muted text-center py-4'>لا توجد أصناف مسجلة لهذه الكلية.</p>"
        )

    item_ids = [s.item_id for s in stocks]

    # 2. Fetch ALL pending details for these items in one query
    pending_details = (
        ItemTransactionDetails.objects.filter(
            item_id__in=item_ids,
            transaction__faculty=faculty,
            transaction__approval_status=ItemTransactions.APPROVAL_STATUS.PENDING,
            transaction__deleted=False,
            transaction__is_reversed=False,
        )
        .values("item_id", "transaction__transaction_type")
        .annotate(total=Sum("approved_quantity"))
    )

    # 3. Build lookup map: {item_id: {"A": 10, "D": 5, "T": 2, "R": 3}}
    pending_map = {}
    for pd in pending_details:
        item_id = pd["item_id"]
        tx_type = pd["transaction__transaction_type"]
        pending_map.setdefault(item_id, {})[tx_type] = pd["total"]

    # 4. Prepare final list for template
    items_data = []
    for stock in stocks:
        p = pending_map.get(stock.item_id, {})
        items_data.append(
            {
                "item": stock.item,
                "current_qty": stock.cached_quantity,
                "pending_add": p.get(ItemTransactions.TRANSACTION_TYPES.Addition, 0)
                or 0,
                "pending_disb": p.get(
                    ItemTransactions.TRANSACTION_TYPES.Disbursement, 0
                )
                or 0,
                "pending_trans": p.get(ItemTransactions.TRANSACTION_TYPES.Transfer, 0)
                or 0,
                "pending_ret": p.get(ItemTransactions.TRANSACTION_TYPES.Return, 0) or 0,
                "sub_warehouse": stock.sub_warehouse,
            }
        )

    return render(
        request,
        "inventory/partials/faculty_items_table.html",
        {"items_data": items_data, "faculty": faculty},
    )


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_item_transaction_history(request, item_id):
    """View transaction history for a specific item in a faculty."""
    item = get_object_or_404(Item, id=item_id)
    faculty_id = request.GET.get("faculty_id")
    faculty = None
    if faculty_id and faculty_id.isdigit():
        faculty = get_object_or_404(Faculty, id=faculty_id)

    qs = (
        ItemTransactionDetails.objects.filter(item=item)
        .select_related(
            "transaction",
            "transaction__created_by",
            "transaction__approval_user",
            "transaction__from_sub_warehouse",
            "transaction__to_sub_warehouse",
        )
        .order_by("-transaction__created_at")
    )

    if faculty:
        qs = qs.filter(transaction__faculty=faculty)

    # paginator = Paginator(qs, 50)
    # page_number = request.GET.get("page")
    # details_page = paginator.get_page(page_number)

    return render(
        request,
        "inventory/admin_item_history.html",
        {"item": item, "faculty": faculty, "details": qs},
    )
